# Author: Miguel Marina <karel.capek.robotics@gmail.com>
# SPDX-License-Identifier: GPL-3.0-or-later
#
# Copyright (C) 2025 Capek System Safety & Robotic Solutions
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.

from __future__ import annotations

"""ODD library management helpers."""

import csv
import tkinter as tk
from tkinter import filedialog, simpledialog, ttk

from gui.controls import messagebox

try:  # optional dependency
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - openpyxl may not be installed
    load_workbook = None


class OddLibraryManager:
    """Manage Operational Design Domain (ODD) libraries for :class:`AutoMLApp`."""

    def __init__(self, app: "AutoMLApp") -> None:
        self.app = app

    # ------------------------------------------------------------------
    def manage_odd_libraries(self) -> None:
        app = self.app
        if hasattr(app, "_odd_tab") and app._odd_tab.winfo_exists():
            app.doc_nb.select(app._odd_tab)
            return
        app._odd_tab = app.lifecycle_ui._new_tab("ODD Libraries")
        win = app._odd_tab
        lib_lb = tk.Listbox(win, height=8, width=25)
        lib_lb.grid(row=0, column=0, rowspan=4, sticky="nsew")
        elem_tree = ttk.Treeview(win, columns=("cls", "attrs"), show="tree headings")
        elem_tree.heading("#0", text="Name")
        elem_tree.column("#0", width=150)
        elem_tree.heading("cls", text="Class")
        elem_tree.column("cls", width=120)
        elem_tree.heading("attrs", text="Attributes")
        elem_tree.column("attrs", width=200)
        elem_tree.grid(row=0, column=1, columnspan=3, sticky="nsew")
        win.grid_rowconfigure(0, weight=1)
        win.grid_columnconfigure(1, weight=1)

        if not hasattr(app, "odd_elem_icon"):
            app.odd_elem_icon = app._create_icon("rect", "#696969")

        def refresh_libs() -> None:
            lib_lb.delete(0, tk.END)
            for lib in app.odd_libraries:
                lib_lb.insert(tk.END, lib.get("name", ""))
            refresh_elems()

        def refresh_elems(*_args: object) -> None:
            elem_tree.delete(*elem_tree.get_children())
            sel = lib_lb.curselection()
            if not sel:
                return
            lib = app.odd_libraries[sel[0]]
            for el in lib.get("elements", []):
                name = el.get("name") or el.get("element") or el.get("id")
                cls = el.get("class", "")
                attrs = ", ".join(
                    f"{k}={v}" for k, v in el.items() if k not in {"name", "class"}
                )
                opts = {"values": (cls, attrs), "text": name}
                if app.odd_elem_icon:
                    opts["image"] = app.odd_elem_icon
                elem_tree.insert("", tk.END, **opts)

        def import_elements_from_file(path: str) -> list[dict]:
            elems: list[dict] = []
            if path.lower().endswith(".csv"):
                with open(path, newline="") as f:
                    elems = list(csv.DictReader(f))
            elif path.lower().endswith(".xlsx"):
                try:
                    if load_workbook is None:
                        raise ImportError
                    wb = load_workbook(path, read_only=True)
                    ws = wb.active
                    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        elem = {headers[i]: row[i] for i in range(len(headers))}
                        elems.append(elem)
                except Exception:
                    messagebox.showerror(
                        "Import", "Failed to read Excel file. openpyxl required."
                    )
            return elems

        class ElementDialog(simpledialog.Dialog):
            def __init__(self, parent, app, data=None):
                self.app = app
                self.data = data or {"name": "", "class": ""}
                super().__init__(parent, title="Edit Element")

            def add_attr_row(self, key="", val=""):
                r = len(self.attr_rows)
                k_var = tk.StringVar(value=key)
                v_var = tk.StringVar(value=str(val))
                k_entry = ttk.Entry(self.attr_frame, textvariable=k_var)
                v_entry = ttk.Entry(self.attr_frame, textvariable=v_var)
                k_entry.grid(row=r, column=0, padx=2, pady=2)
                v_entry.grid(row=r, column=1, padx=2, pady=2)

                row = {}

                def remove_row():
                    k_entry.destroy()
                    v_entry.destroy()
                    del_btn.destroy()
                    self.attr_rows.remove(row)
                    for i, rdata in enumerate(self.attr_rows):
                        rdata["k_entry"].grid_configure(row=i)
                        rdata["v_entry"].grid_configure(row=i)
                        rdata["del_btn"].grid_configure(row=i)

                del_btn = ttk.Button(self.attr_frame, text="Delete", command=remove_row)
                del_btn.grid(row=r, column=2, padx=2, pady=2)

                row.update(
                    {
                        "k_var": k_var,
                        "v_var": v_var,
                        "k_entry": k_entry,
                        "v_entry": v_entry,
                        "del_btn": del_btn,
                    }
                )
                self.attr_rows.append(row)

            def body(self, master):  # type: ignore[override]
                ttk.Label(master, text="Name").grid(row=0, column=0, sticky="e")
                self.name_var = tk.StringVar(value=self.data.get("name", ""))
                ttk.Entry(master, textvariable=self.name_var).grid(row=0, column=1, sticky="ew")

                ttk.Label(master, text="Class").grid(row=1, column=0, sticky="e")
                self.class_var = tk.StringVar(value=self.data.get("class", ""))
                cls_opts = [
                    "Road",
                    "Infrastructure",
                    "Temporal",
                    "Movable",
                    "Environment",
                ]
                ttk.Combobox(
                    master, textvariable=self.class_var, values=cls_opts, state="readonly"
                ).grid(row=1, column=1, sticky="ew")

                nb = ttk.Notebook(master)
                nb.grid(row=2, column=0, columnspan=2, sticky="nsew")
                master.grid_rowconfigure(2, weight=1)
                master.grid_columnconfigure(1, weight=1)

                # Attributes tab
                self.attr_frame = ttk.Frame(nb)
                nb.add(self.attr_frame, text="Attributes")
                self.attr_rows = []
                for k, v in self.data.items():
                    if k not in {"name", "class", "p", "n", "tp", "fp", "tn", "fn"}:
                        self.add_attr_row(k, v)
                ttk.Button(
                    self.attr_frame, text="Add Attribute", command=self.add_attr_row
                ).grid(row=99, column=0, columnspan=3, pady=5)

                # Confusion matrix tab
                cm_frame = ttk.Frame(nb)
                nb.add(cm_frame, text="Confusion Matrix")
                self.p_var = tk.DoubleVar(value=float(self.data.get("p", 0) or 0))
                self.n_var = tk.DoubleVar(value=float(self.data.get("n", 0) or 0))
                self.tp_var = tk.DoubleVar(value=float(self.data.get("tp", 0) or 0))
                self.fp_var = tk.DoubleVar(value=float(self.data.get("fp", 0) or 0))
                self.tn_var = tk.DoubleVar(value=float(self.data.get("tn", 0) or 0))
                self.fn_var = tk.DoubleVar(value=float(self.data.get("fn", 0) or 0))
                ttk.Label(cm_frame, text="TP").grid(row=0, column=0, sticky="e")
                ttk.Entry(cm_frame, textvariable=self.tp_var).grid(row=0, column=1, sticky="ew")
                ttk.Label(cm_frame, text="FP").grid(row=1, column=0, sticky="e")
                ttk.Entry(cm_frame, textvariable=self.fp_var).grid(row=1, column=1, sticky="ew")
                ttk.Label(cm_frame, text="TN").grid(row=2, column=0, sticky="e")
                ttk.Entry(cm_frame, textvariable=self.tn_var).grid(row=2, column=1, sticky="ew")
                ttk.Label(cm_frame, text="FN").grid(row=3, column=0, sticky="e")
                ttk.Entry(cm_frame, textvariable=self.fn_var).grid(row=3, column=1, sticky="ew")

            def apply(self):  # type: ignore[override]
                new_data = {"name": self.name_var.get(), "class": self.class_var.get()}
                for row in self.attr_rows:
                    key = row["k_var"].get().strip()
                    if key:
                        new_data[key] = row["v_var"].get()
                tp = float(self.tp_var.get())
                fp = float(self.fp_var.get())
                tn = float(self.tn_var.get())
                fn = float(self.fn_var.get())
                from analysis.confusion_matrix import compute_metrics

                metrics = compute_metrics(tp, fp, tn, fn)
                p = tp + fn
                n = tn + fp
                new_data.update(
                    {
                        "tp": tp,
                        "fp": fp,
                        "tn": tn,
                        "fn": fn,
                        "p": p,
                        "n": n,
                    }
                )
                new_data.update(metrics)
                self.data = new_data

        def add_lib() -> None:
            name = simpledialog.askstring("New Library", "Library name:")
            if not name:
                return
            elems: list[dict] = []
            if messagebox.askyesno("Import", "Import elements from file?"):
                path = filedialog.askopenfilename(
                    filetypes=[("CSV/Excel", "*.csv *.xlsx")]
                )
                if path:
                    elems = import_elements_from_file(path)
            app.odd_libraries.append({"name": name, "elements": elems})
            refresh_libs()
            app.update_odd_elements()

        def edit_lib() -> None:
            sel = lib_lb.curselection()
            if not sel:
                return
            lib = app.odd_libraries[sel[0]]
            name = simpledialog.askstring(
                "Edit Library", "Library name:", initialvalue=lib.get("name", "")
            )
            if name:
                lib["name"] = name
                refresh_libs()

        def delete_lib() -> None:
            sel = lib_lb.curselection()
            if sel:
                idx = sel[0]
                del app.odd_libraries[idx]
                refresh_libs()
                app.update_odd_elements()

        def add_elem() -> None:
            sel = lib_lb.curselection()
            if not sel:
                return
            lib = app.odd_libraries[sel[0]]
            dlg = ElementDialog(win, app)
            lib.setdefault("elements", []).append(dlg.data)
            refresh_elems()
            app.update_odd_elements()

        def edit_elem() -> None:
            sel_lib = lib_lb.curselection()
            sel_elem = elem_tree.selection()
            if not sel_lib or not sel_elem:
                return
            lib = app.odd_libraries[sel_lib[0]]
            idx = elem_tree.index(sel_elem[0])
            data = lib.get("elements", [])[idx]
            dlg = ElementDialog(win, app, data)
            lib["elements"][idx] = dlg.data
            refresh_elems()
            app.update_odd_elements()

        def del_elem() -> None:
            sel_lib = lib_lb.curselection()
            sel_elem = elem_tree.selection()
            if not sel_lib or not sel_elem:
                return
            lib = app.odd_libraries[sel_lib[0]]
            idx = elem_tree.index(sel_elem[0])
            del lib.get("elements", [])[idx]
            refresh_elems()
            app.update_odd_elements()

        btnf = ttk.Frame(win)
        btnf.grid(row=1, column=1, columnspan=3, sticky="ew")
        ttk.Button(btnf, text="Add Lib", command=add_lib).pack(side=tk.LEFT)
        ttk.Button(btnf, text="Edit Lib", command=edit_lib).pack(side=tk.LEFT)
        ttk.Button(btnf, text="Del Lib", command=delete_lib).pack(side=tk.LEFT)
        ttk.Button(btnf, text="Add Elem", command=add_elem).pack(side=tk.LEFT, padx=5)
        ttk.Button(btnf, text="Edit Elem", command=edit_elem).pack(side=tk.LEFT)
        ttk.Button(btnf, text="Del Elem", command=del_elem).pack(side=tk.LEFT)

        lib_lb.bind("<<ListboxSelect>>", refresh_elems)
        refresh_libs()
