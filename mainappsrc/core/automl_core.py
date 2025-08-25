#!/usr/bin/env python3
# Author: Miguel Marina <karel.capek.robotics@gmail.com>
# SPDX-License-Identifier: GPL-3.0-or-later
#
# Copyright (C) 2025 Capek System Safety & Robotic Solutions
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.

import re
import math
import sys
import json
from concurrent.futures import ThreadPoolExecutor
import tkinter as tk
import os, sys
base = os.path.dirname(__file__)
if base not in sys.path:
    sys.path.append(base)
parent = os.path.dirname(base)
if parent not in sys.path:
    sys.path.append(parent)
from typing import Any, Optional
from tkinter import ttk, filedialog, simpledialog, scrolledtext
from gui.dialogs.dialog_utils import askstring_fixed
from gui.controls import messagebox
from gui.utils import logger, add_treeview_scrollbars
from gui.controls.button_utils import enable_listbox_hover_highlight
from gui.utils.tooltip import ToolTip
from gui.styles.style_manager import StyleManager
from gui.toolboxes.review_toolbox import ReviewData, ReviewParticipant, ReviewComment
from functools import partial
# Governance helper class
from mainappsrc.managers.paa_manager import PrototypeAssuranceManager
from gui.toolboxes.safety_management_toolbox import SafetyManagementToolbox
from gui.explorers.safety_case_explorer import SafetyCaseExplorer
from gui.windows.gsn_diagram_window import GSN_WINDOWS
from gui.windows.causal_bayesian_network_window import CBN_WINDOWS
from gui.windows.gsn_config_window import GSNElementConfig
from mainappsrc.models.gsn import GSNDiagram, GSNModule
from mainappsrc.models.gsn.nodes import GSNNode, ALLOWED_AWAY_TYPES
from gui.utils.closable_notebook import ClosableNotebook
from gui.controls.mac_button_style import (
    apply_translucid_button_style,
    apply_purplish_button_style,
)
from gui.dialogs.user_select_dialog import UserSelectDialog
from gui.dialogs.decomposition_dialog import DecompositionDialog
from dataclasses import asdict
from pathlib import Path
from .ui_setup import UISetupMixin
from .event_handlers import EventHandlersMixin
from .persistence_wrappers import PersistenceWrappersMixin
from .analysis_utils import AnalysisUtilsMixin
from .service_init_mixin import ServiceInitMixin
from .icon_setup_mixin import IconSetupMixin
from .style_setup_mixin import StyleSetupMixin
from .page_diagram import PageDiagram
from .node_utils import resolve_original as resolve_node_original
from .editors import (
    ItemDefinitionEditorMixin,
    SafetyConceptEditorMixin,
    RequirementsEditorMixin,
)
from .app_initializer import AppInitializer
from analysis.mechanisms import (
    DiagnosticMechanism,
    MechanismLibrary,
    ANNEX_D_MECHANISMS,
    PAS_8800_MECHANISMS,
)
from pathlib import Path
from collections.abc import Mapping
import csv
try:
    from openpyxl import load_workbook
except Exception:  # openpyxl may not be installed
    load_workbook = None
from gui.utils.drawing_helper import FTADrawingHelper, fta_drawing_helper
from mainappsrc.core.event_dispatcher import EventDispatcher
from mainappsrc.core.window_controllers import WindowControllers
from mainappsrc.core.top_event_workflows import Top_Event_Workflows
from mainappsrc.managers.review_manager import ReviewManager
from mainappsrc.managers.drawing_manager import DrawingManager
from .versioning_review import Versioning_Review
from .validation_consistency import Validation_Consistency
from .reporting_export import Reporting_Export
from .node_clone_service import NodeCloneService
from .view_updater import ViewUpdater
from analysis.user_config import (
    load_user_config,
    save_user_config,
    set_current_user,
    load_all_users,
    set_last_user,
    CURRENT_USER_NAME,
    CURRENT_USER_EMAIL,
)
from analysis.risk_assessment import (
    DERIVED_MATURITY_TABLE,
    ASSURANCE_AGGREGATION_AND,
    AND_DECOMPOSITION_TABLE,
    OR_DECOMPOSITION_TABLE,
    boolify,
    AutoMLHelper,
)
from analysis.models import (
    MissionProfile,
    ReliabilityComponent,
    ReliabilityAnalysis,
    HazopEntry,
    HaraEntry,
    HazopDoc,
    HaraDoc,
    StpaEntry,
    StpaDoc,
    FI2TCDoc,
    TC2FIDoc,
    DamageScenario,
    ThreatScenario,
    AttackPath,
    FunctionThreat,
    ThreatEntry,
    ThreatDoc,
    QUALIFICATIONS,
    COMPONENT_ATTR_TEMPLATES,
    RELIABILITY_MODELS,
    component_fit_map,
    ASIL_LEVEL_OPTIONS,
    ASIL_ORDER,
    ASIL_TARGETS,
    ASIL_TABLE,
    ASIL_DECOMP_SCHEMES,
    calc_asil,
    global_requirements,
    ensure_requirement_defaults,
    REQUIREMENT_TYPE_OPTIONS,
    REQUIREMENT_WORK_PRODUCTS,
    CAL_LEVEL_OPTIONS,
    CybersecurityGoal,
)
from gui.utils.safety_case_table import SafetyCaseTable
from gui.windows.architecture import (
    UseCaseDiagramWindow,
    ActivityDiagramWindow,
    BlockDiagramWindow,
    InternalBlockDiagramWindow,
    ControlFlowDiagramWindow,
    GovernanceDiagramWindow,
    ArchitectureManagerDialog,
    link_requirement_to_object,
    unlink_requirement_from_object,
    link_requirements,
    unlink_requirements,
    ARCH_WINDOWS,
)
from mainappsrc.models.sysml.sysml_repository import SysMLRepository
from .undo_manager import UndoRedoManager
from analysis.fmeda_utils import compute_fmeda_metrics
from analysis.scenario_description import template_phrases
from mainappsrc.core.app_lifecycle_ui import AppLifecycleUI
from mainappsrc.core.editing_labels_styling import Editing_Labels_Styling
import copy
import tkinter.font as tkFont
import builtins
from mainappsrc.managers.user_manager import UserManager
from mainappsrc.managers.project_manager import ProjectManager
from mainappsrc.managers.product_goal_manager import ProductGoalManager
from mainappsrc.ui.project_properties_dialog import ProjectPropertiesDialog
from mainappsrc.managers.sotif_manager import SOTIFManager
from mainappsrc.managers.cyber_manager import CyberSecurityManager
from mainappsrc.managers.cta_manager import ControlTreeManager
from config.automl_constants import (
    dynamic_recommendations,
    WORK_PRODUCT_INFO as BASE_WORK_PRODUCT_INFO,
    WORK_PRODUCT_PARENTS as BASE_WORK_PRODUCT_PARENTS,
    PMHF_TARGETS,
)

builtins.REQUIREMENT_WORK_PRODUCTS = REQUIREMENT_WORK_PRODUCTS
builtins.SafetyCaseTable = SafetyCaseTable
try:
    from PIL import Image, ImageDraw, ImageFont
except ModuleNotFoundError:
    Image = ImageDraw = ImageFont = None
import os
import types
os.environ["GS_EXECUTABLE"] = r"C:\Program Files\gs\gs10.04.0\bin\gswin64c.exe"
import networkx as nx
# Import ReportLab for PDF export.
from reportlab.platypus import Table, TableStyle, SimpleDocTemplate, Paragraph, Spacer, Image as RLImage, PageBreak
from gui.styles.style_editor import StyleEditor
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.platypus import Paragraph, Spacer, Table, TableStyle, PageBreak, SimpleDocTemplate, Image as RLImage
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from io import BytesIO, StringIO
from email.utils import make_msgid
import html
import datetime
try:
    import PIL.Image as PILImage
except ModuleNotFoundError:
    PILImage = None
try:
    from reportlab.platypus import LongTable
except Exception:  # pragma: no cover - fallback when reportlab missing
    LongTable = None
from email.message import EmailMessage
import smtplib
import socket

styles = getSampleStyleSheet()  # Create the stylesheet.
preformatted_style = ParagraphStyle(name="Preformatted", fontName="Courier", fontSize=10)
if hasattr(styles, "add"):
    styles.add(preformatted_style)
else:  # pragma: no cover - fallback for minimal stubs
    styles["Preformatted"] = preformatted_style

# Characters used to display pass/fail status in metrics labels.
from analysis.constants import CHECK_MARK, CROSS_MARK
from analysis.utils import (
    append_unique_insensitive,
    EXPOSURE_PROBABILITIES,
    CONTROLLABILITY_PROBABILITIES,
    SEVERITY_PROBABILITIES,
)
from analysis.safety_management import SafetyManagementToolbox, ACTIVE_TOOLBOX
from analysis.causal_bayesian_network import CausalBayesianNetwork, CausalBayesianNetworkDoc
try:  # pragma: no cover - support direct module import
    from .probability_reliability import Probability_Reliability
    from .version import VERSION
except Exception:  # pragma: no cover
    from mainappsrc.core.probability_reliability import Probability_Reliability
    from mainappsrc.version import VERSION
try:  # pragma: no cover
    from .models.fta.fault_tree_node import FaultTreeNode
except Exception:  # pragma: no cover
    import os, sys
    base = os.path.dirname(__file__)
    sys.path.append(base)
    sys.path.append(os.path.dirname(base))
    from models.fta.fault_tree_node import FaultTreeNode

from .structure_tree_operations import Structure_Tree_Operations

from gui.toolboxes import (
    RequirementsExplorerWindow,
    DiagramElementDialog,
    _RequirementRelationDialog,
)


from pathlib import Path
from gui.dialogs.user_info_dialog import UserInfoDialog

from . import config_utils
from .config_utils import _reload_local_config
from .project_properties_manager import ProjectPropertiesManager

# Expose configuration helpers and global state
_CONFIG_PATH = config_utils._CONFIG_PATH
GATE_NODE_TYPES = config_utils.GATE_NODE_TYPES
_PATTERN_PATH = config_utils._PATTERN_PATH
_REPORT_TEMPLATE_PATH = config_utils._REPORT_TEMPLATE_PATH
unique_node_id_counter = config_utils.unique_node_id_counter
AutoML_Helper = config_utils.AutoML_Helper
import uuid

##########################################
# Edit Dialog 
##########################################
from gui.dialogs.edit_node_dialog import EditNodeDialog, DecompositionDialog
from gui.dialogs.fmea_row_dialog import FMEARowDialog
from gui.dialogs.req_dialog import ReqDialog
from gui.dialogs.select_base_event_dialog import SelectBaseEventDialog
from .safety_ui import SafetyUIMixin

##########################################
# Main Application (Parent Diagram)
##########################################
class AutoMLApp(
    StyleSetupMixin,
    ServiceInitMixin,
    IconSetupMixin,
    SafetyUIMixin,
    UISetupMixin,
    EventHandlersMixin,
    PersistenceWrappersMixin,
    AnalysisUtilsMixin,
):
    """Main application window for AutoML Analyzer."""

    _instance: Optional["AutoMLApp"] = None

    #: Maximum number of characters displayed for a notebook tab title. Longer
    #: titles are truncated with an ellipsis to avoid giant tabs that overflow
    #: the working area.
    MAX_TAB_TEXT_LENGTH = 20
    GATE_NODE_TYPES = GATE_NODE_TYPES

    @property
    def fmedas(self):
        return self.safety_analysis.fmedas

    @fmedas.setter
    def fmedas(self, value):
        self.safety_analysis.fmedas = value

    #: Maximum characters shown for tool notebook tab titles. Tool tabs use
    #: a fixed width so they remain readable but long names are capped at this
    #: length and truncated with an ellipsis.
    MAX_TOOL_TAB_TEXT_LENGTH = 20

    #: Maximum number of tabs displayed at once in the tools and document
    #: notebooks. Additional tabs can be accessed via the navigation buttons.
    MAX_VISIBLE_TABS = 4

    WORK_PRODUCT_INFO = BASE_WORK_PRODUCT_INFO.copy()
    for _wp in REQUIREMENT_WORK_PRODUCTS:
        WORK_PRODUCT_INFO.setdefault(
            _wp,
            (
                "System Design (Item Definition)",
                "Requirements Editor",
                "show_requirements_editor",
            ),
        )

    # Mapping of work products to their parent menu categories.  When a
    # child work product is enabled its parent menu must also become
    # active so the submenu is reachable.
    WORK_PRODUCT_PARENTS = BASE_WORK_PRODUCT_PARENTS.copy()

    # Ensure all requirement work products activate the top-level Requirements
    # menu.  Each specific requirement specification (e.g. vehicle, functional
    # safety) is treated as a child of the generic "Requirements" category so
    # that declaring any of them on a governance diagram enables the
    # corresponding menu items.
    for _wp in REQUIREMENT_WORK_PRODUCTS:
        WORK_PRODUCT_PARENTS.setdefault(_wp, "Requirements")

    @property
    def window_controllers(self) -> WindowControllers:
        if not hasattr(self, "_window_controllers"):
            self._window_controllers = WindowControllers(self)
        return self._window_controllers

    @property
    def top_event_workflows(self) -> Top_Event_Workflows:
        if not hasattr(self, "_top_event_workflows"):
            self._top_event_workflows = Top_Event_Workflows(self)
        return self._top_event_workflows

    def __getattr__(self, name):  # pragma: no cover - simple delegation
        """Delegate missing attributes to the lifecycle UI helper.

        ``AppLifecycleUI`` now hosts a number of UI-centric helpers that were
        previously methods on :class:`AutoMLApp`.  Existing code (and tests)
        still expect these helpers to be accessible directly from the main
        application instance.  This ``__getattr__`` implementation forwards
        such attribute lookups to ``self.lifecycle_ui`` when the attribute
        exists there, preserving backwards compatibility without replicating
        numerous wrapper methods.
        """

        ui = self.__dict__.get("lifecycle_ui")
        if ui and (
            name in ui.__dict__
            or any(name in cls.__dict__ for cls in ui.__class__.mro())
        ):
            return getattr(ui, name)
        raise AttributeError(f"{type(self).__name__!r} object has no attribute {name!r}")

    def __init__(self, root):
        AutoMLApp._instance = self
        self.root = root
        self.setup_style(root)
        self.lifecycle_ui = AppLifecycleUI(self, root)
        self.labels_styling = Editing_Labels_Styling(self)
        self.top_events = []
        self.cta_events = []
        self.paa_events = []
        self.fta_root_node = None
        self.cta_root_node = None
        self.paa_root_node = None
        self.analysis_tabs = {}
        self.shared_product_goals = {}
        self.product_goal_manager = ProductGoalManager()
        self.selected_node = None
        self.clone_offset_counter = {}
        self.node_clone_service = NodeCloneService()
        self.view_updater = ViewUpdater(self)
        self._loaded_model_paths = []
        self.root.title("AutoML-Analyzer")
        self.messagebox = messagebox
        self.version = VERSION
        self.zoom = 1.0
        self.rc_dragged = False
        self.diagram_font = tkFont.Font(family="Arial", size=int(8 * self.zoom))
        self.lifecycle_ui._init_nav_button_style()
        self.setup_services()
        self.setup_icons()
        AppInitializer(self).initialize()

        menubar = tk.Menu(root)
        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label="New AutoML Model", command=self.project_manager.new_model, accelerator="Ctrl+N")
        file_menu.add_command(label="Save AutoML Model", command=self.project_manager.save_model, accelerator="Ctrl+S")
        file_menu.add_command(label="Load AutoML Model", command=self.project_manager.load_model, accelerator="Ctrl+O")
        file_menu.add_command(label="Project Properties", command=self.edit_project_properties)
        file_menu.add_command(label="Save PDF Report", command=self.generate_pdf_report)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.confirm_close)

        edit_menu = tk.Menu(menubar, tearoff=0)
        edit_menu.add_command(label="Undo", command=self.undo, accelerator="Ctrl+Z")
        edit_menu.add_command(label="Redo", command=self.redo, accelerator="Ctrl+Y")
        edit_menu.add_separator()
        edit_menu.add_command(label="Edit Selected", command=self.edit_selected)
        edit_menu.add_command(label="Remove Connection", command=lambda: self.remove_connection(self.selected_node) if self.selected_node else None)
        edit_menu.add_command(label="Delete Node", command=lambda: self.delete_node_and_subtree(self.selected_node) if self.selected_node else None)
        edit_menu.add_command(label="Remove Node", command=self.remove_node)
        edit_menu.add_separator()
        edit_menu.add_command(label="Copy", command=self.copy_node, accelerator="Ctrl+C")
        edit_menu.add_command(label="Cut", command=self.cut_node, accelerator="Ctrl+X")
        edit_menu.add_command(label="Paste", command=self.paste_node, accelerator="Ctrl+V")
        edit_menu.add_separator()
        edit_menu.add_command(label="Edit User Name", command=self.user_manager.edit_user_name, accelerator="Ctrl+U")
        edit_menu.add_command(label="Edit Description", command=self.edit_description, accelerator="Ctrl+D")
        edit_menu.add_command(label="Edit Rationale", command=self.edit_rationale, accelerator="Ctrl+L")
        edit_menu.add_command(label="Edit Value", command=self.edit_value)
        edit_menu.add_command(label="Edit Gate Type", command=self.edit_gate_type, accelerator="Ctrl+G")
        edit_menu.add_command(label="Edit Severity", command=self.edit_severity, accelerator="Ctrl+E")
        edit_menu.add_command(label="Edit Controllability", command=self.edit_controllability)
        edit_menu.add_command(label="Edit Page Flag", command=self.edit_page_flag)
        search_menu = tk.Menu(menubar, tearoff=0)
        search_menu.add_command(
            label="Find...", command=self.open_search_toolbox, accelerator="Ctrl+F"
        )
        process_menu = tk.Menu(menubar, tearoff=0)
        process_menu.add_command(label="Calc Prototype Assurance Level (PAL)", command=self.calculate_overall, accelerator="Ctrl+R")
        process_menu.add_command(label="Calc PMHF", command=self.calculate_pmfh, accelerator="Ctrl+M")
        view_menu = tk.Menu(menubar, tearoff=0)
        view_menu.add_command(label="Zoom In", command=self.zoom_in, accelerator="Ctrl++")
        view_menu.add_command(label="Zoom Out", command=self.zoom_out, accelerator="Ctrl+-")
        view_menu.add_command(label="Style Editor", command=self.open_style_editor)
        view_menu.add_command(
            label="Light Mode",
            command=lambda: self.apply_style('pastel.xml'),
        )
        view_menu.add_command(label="Metrics", command=self.lifecycle_ui.open_metrics_tab)

        requirements_menu = tk.Menu(menubar, tearoff=0)
        requirements_menu.add_command(
            label="Requirements Matrix",
            command=self.show_requirements_matrix,
            state=tk.DISABLED,
        )
        matrix_idx = requirements_menu.index("end")
        requirements_menu.add_command(
            label="Requirements Editor",
            command=self.show_requirements_editor,
            state=tk.DISABLED,
        )
        editor_idx = requirements_menu.index("end")
        requirements_menu.add_command(
            label="Requirements Explorer",
            command=self.show_requirements_explorer,
            state=tk.DISABLED,
        )
        explorer_idx = requirements_menu.index("end")
        for wp in REQUIREMENT_WORK_PRODUCTS:
            self.work_product_menus.setdefault(wp, []).extend(
                [
                    (requirements_menu, matrix_idx),
                    (requirements_menu, editor_idx),
                    (requirements_menu, explorer_idx),
                ]
            )
        requirements_menu.add_command(
            label="Product Goals Matrix", command=self.show_safety_goals_matrix
        )
        requirements_menu.add_command(
            label="Product Goals Editor",
            command=self.show_product_goals_editor,
            state=tk.DISABLED,
        )
        self.work_product_menus.setdefault("Product Goal Specification", []).append(
            (requirements_menu, requirements_menu.index("end"))
        )
        requirements_menu.add_command(
            label="Safety Performance Indicators",
            command=self.show_safety_performance_indicators,
        )
        self.lifecycle_ui._add_lifecycle_requirements_menu(requirements_menu)
        self.phase_req_menu = tk.Menu(requirements_menu, tearoff=0)
        requirements_menu.add_cascade(
            label="Phase Requirements", menu=self.phase_req_menu
        )
        self._refresh_phase_requirements_menu()
        requirements_menu.add_command(
            label="Export Product Goal Requirements",
            command=self.export_product_goal_requirements,
        )
        review_menu = tk.Menu(menubar, tearoff=0)
        review_menu.add_command(label="Start Peer Review", command=self.start_peer_review)
        review_menu.add_command(label="Start Joint Review", command=self.start_joint_review)
        review_menu.add_command(label="Open Review Toolbox", command=self.open_review_toolbox)
        review_menu.add_command(label="Set Current User", command=self.user_manager.set_current_user)
        review_menu.add_command(label="Merge Review Comments", command=self.merge_review_comments)
        review_menu.add_command(label="Compare Versions", command=self.compare_versions)
        architecture_menu = tk.Menu(menubar, tearoff=0)
        architecture_menu.add_command(label="Use Case Diagram", command=self.window_controllers.open_use_case_diagram)
        architecture_menu.add_command(label="Activity Diagram", command=self.window_controllers.open_activity_diagram)
        architecture_menu.add_command(label="Block Diagram", command=self.window_controllers.open_block_diagram)
        architecture_menu.add_command(label="Internal Block Diagram", command=self.window_controllers.open_internal_block_diagram)
        architecture_menu.add_command(label="Control Flow Diagram", command=self.window_controllers.open_control_flow_diagram)
        architecture_menu.add_separator()
        architecture_menu.add_command(
            label="AutoML Explorer",
            command=self.manage_architecture,
            state=tk.DISABLED,
        )
        self.work_product_menus.setdefault("Architecture Diagram", []).append(
            (architecture_menu, architecture_menu.index("end"))
        )
        # --- Risk Assessment Menu ---
        risk_menu = tk.Menu(menubar, tearoff=0)
        risk_menu.add_command(
            label="HAZOP Analysis",
            command=self.open_hazop_window,
            state=tk.DISABLED,
        )
        self.work_product_menus.setdefault("HAZOP", []).append(
            (risk_menu, risk_menu.index("end"))
        )
        risk_menu.add_command(
            label="Risk Assessment",
            command=self.open_risk_assessment_window,
            state=tk.DISABLED,
        )
        self.work_product_menus.setdefault("Risk Assessment", []).append(
            (risk_menu, risk_menu.index("end"))
        )
        risk_menu.add_command(
            label="STPA Analysis",
            command=self.open_stpa_window,
            state=tk.DISABLED,
        )
        self.work_product_menus.setdefault("STPA", []).append(
            (risk_menu, risk_menu.index("end"))
        )
        risk_menu.add_command(
            label="Threat Analysis",
            command=self.open_threat_window,
            state=tk.DISABLED,
        )
        self.work_product_menus.setdefault("Threat Analysis", []).append(
            (risk_menu, risk_menu.index("end"))
        )
        risk_menu.add_command(label="Hazard Explorer", command=self.show_hazard_explorer)
        risk_menu.add_command(label="Hazards Editor", command=self.show_hazard_editor)
        risk_menu.add_command(label="Malfunctions Editor", command=self.show_malfunction_editor)
        risk_menu.add_command(label="Triggering Conditions", command=self.show_triggering_condition_list)
        risk_menu.add_command(label="Functional Insufficiencies", command=self.show_functional_insufficiency_list)
        risk_menu.add_separator()
        risk_menu.add_command(
            label="FI2TC Analysis",
            command=self.open_fi2tc_window,
            state=tk.DISABLED,
        )
        self.work_product_menus.setdefault("FI2TC", []).append(
            (risk_menu, risk_menu.index("end"))
        )
        risk_menu.add_command(
            label="TC2FI Analysis",
            command=self.open_tc2fi_window,
            state=tk.DISABLED,
        )
        self.work_product_menus.setdefault("TC2FI", []).append(
            (risk_menu, risk_menu.index("end"))
        )
                
        # --- Qualitative Analysis Menu ---
        qualitative_menu = tk.Menu(menubar, tearoff=0)
        qualitative_menu.add_command(
            label="FMEA Manager",
            command=self.fmea_service.show_fmea_list,
            state=tk.DISABLED,
        )
        self.work_product_menus.setdefault("FMEA", []).append(
            (qualitative_menu, qualitative_menu.index("end"))
        )

        cta_menu = tk.Menu(qualitative_menu, tearoff=0)
        cta_menu.add_command(label="Add Top Level Event", command=self.cta_manager.create_diagram)
        cta_menu.add_separator()
        cta_menu.add_command(label="Add Triggering Condition", command=lambda: self.add_node_of_type("Triggering Condition"))
        cta_indices = {"add_trigger": cta_menu.index("end")}
        cta_menu.add_command(label="Add Functional Insufficiency", command=lambda: self.add_node_of_type("Functional Insufficiency"))
        cta_indices["add_functional_insufficiency"] = cta_menu.index("end")
        qualitative_menu.add_cascade(label="CTA", menu=cta_menu, state=tk.DISABLED)
        self.work_product_menus.setdefault("CTA", []).append(
            (qualitative_menu, qualitative_menu.index("end"))
        )
        self.cta_manager.register_menu(cta_menu, cta_indices)
        qualitative_menu.add_command(
            label="Fault Prioritization",
            command=self.open_fault_prioritization_window,
        )

        paa_menu = tk.Menu(qualitative_menu, tearoff=0)
        paa_menu.add_command(
            label="Add Top Level Event",
            command=self.paa_manager.create_paa_diagram,
        )
        paa_menu.add_separator()
        paa_menu.add_command(
            label="Add Confidence",
            command=lambda: self.add_node_of_type("Confidence Level"),
            accelerator="Ctrl+Shift+C",
        )
        self._paa_menu_indices = {"add_confidence": paa_menu.index("end")}
        paa_menu.add_command(
            label="Add Robustness",
            command=lambda: self.add_node_of_type("Robustness Score"),
            accelerator="Ctrl+Shift+R",
        )
        self._paa_menu_indices["add_robustness"] = paa_menu.index("end")
        qualitative_menu.add_cascade(
            label="Prototype Assurance Analysis",
            menu=paa_menu,
            state=tk.DISABLED,
        )
        self.work_product_menus.setdefault("Prototype Assurance Analysis", []).append(
            (qualitative_menu, qualitative_menu.index("end"))
        )
        self.paa_menu = paa_menu
        
        # --- Quantitative Analysis Menu ---
        quantitative_menu = tk.Menu(menubar, tearoff=0)
        quantitative_menu.add_command(
            label="Mission Profiles",
            command=self.manage_mission_profiles,
            state=tk.DISABLED,
        )
        self.work_product_menus.setdefault("Mission Profile", []).append(
            (quantitative_menu, quantitative_menu.index("end"))
        )
                
        quantitative_menu.add_separator()
        quantitative_menu.add_command(label="Faults Editor", command=self.show_fault_editor)
        quantitative_menu.add_command(label="Failures Editor", command=self.show_failure_editor)
        quantitative_menu.add_separator()
        
        quantitative_menu.add_command(
            label="Mechanism Libraries", command=self.manage_mechanism_libraries
        )
        quantitative_menu.add_command(
            label="Reliability Analysis",
            command=self.open_reliability_window,
            state=tk.DISABLED,
        )
        self.work_product_menus.setdefault("Reliability Analysis", []).append(
            (quantitative_menu, quantitative_menu.index("end"))
        )
        quantitative_menu.add_command(
            label="Causal Bayesian Network",
            command=self.window_controllers.open_causal_bayesian_network_window,
            state=tk.DISABLED,
        )
        self.work_product_menus.setdefault("Causal Bayesian Network Analysis", []).append(
            (quantitative_menu, quantitative_menu.index("end"))
        )
        quantitative_menu.add_command(
            label="FMEDA Analysis",
            command=self.open_fmeda_window,
            state=tk.DISABLED,
        )
        self.work_product_menus.setdefault("FMEDA", []).append(
            (quantitative_menu, quantitative_menu.index("end"))
        )
        quantitative_menu.add_command(
            label="FMEDA Manager",
            command=self.show_fmeda_list,
            state=tk.DISABLED,
        )
        fta_menu = tk.Menu(quantitative_menu, tearoff=0)
        fta_menu.add_command(label="Add Top Level Event", command=self.create_fta_diagram)
        fta_menu.add_separator()
        fta_menu.add_command(label="Add Gate", command=lambda: self.add_node_of_type("GATE"), accelerator="Ctrl+Shift+G")
        self._fta_menu_indices = {"add_gate": fta_menu.index("end")}
        fta_menu.add_command(label="Add Basic Event", command=lambda: self.add_node_of_type("Basic Event"), accelerator="Ctrl+Shift+B")
        self._fta_menu_indices["add_basic_event"] = fta_menu.index("end")
        fta_menu.add_command(label="Add FMEA/FMEDA Event", command=self.add_basic_event_from_fmea)
        fta_menu.add_command(label="Add Gate from Failure Mode", command=self.add_gate_from_failure_mode)
        self._fta_menu_indices["add_gate_from_failure_mode"] = fta_menu.index("end")
        fta_menu.add_command(label="Add Fault Event", command=self.add_fault_event)
        self._fta_menu_indices["add_fault_event"] = fta_menu.index("end")
        fta_menu.add_separator()
        fta_menu.add_command(label="FTA-FMEA Traceability", command=self.show_traceability_matrix)
        fta_menu.add_command(
            label="FTA Cut Sets",
            command=self.show_cut_sets,
            state=tk.DISABLED,
        )
        fta_menu.add_command(label="Common Cause Toolbox", command=self.show_common_cause_view)
        fta_menu.add_command(label="Cause & Effect Chain", command=self.show_cause_effect_chain)
        self.fta_menu = fta_menu
        quantitative_menu.add_cascade(label="FTA", menu=fta_menu, state=tk.DISABLED)
        self.work_product_menus.setdefault("FTA", []).append(
            (quantitative_menu, quantitative_menu.index("end"))
        )
        
        libs_menu = tk.Menu(menubar, tearoff=0)
        libs_menu.add_command(
            label="Scenario Libraries",
            command=self.manage_scenario_libraries,
            state=tk.DISABLED,
        )
        self.work_product_menus.setdefault("Scenario Library", []).append(
            (libs_menu, libs_menu.index("end"))
        )
        libs_menu.add_command(
            label="ODD Libraries",
            command=self.manage_odd_libraries,
            state=tk.DISABLED,
        )
        self.work_product_menus.setdefault("ODD", []).append(
            (libs_menu, libs_menu.index("end"))
        )

        gsn_menu = tk.Menu(menubar, tearoff=0)
        gsn_menu.add_command(label="GSN Explorer", command=self.gsn_manager.manage_gsn)
        self.work_product_menus.setdefault("GSN Argumentation", []).append(
            (gsn_menu, gsn_menu.index("end"))
        )
        gsn_menu.add_command(
            label="Safety & Security Case Explorer", command=self.manage_safety_cases
        )
        self.work_product_menus.setdefault("Safety & Security Case", []).append(
            (gsn_menu, gsn_menu.index("end"))
        )

        # Add menus to the bar in the desired order
        menubar.add_cascade(label="File", menu=file_menu)
        menubar.add_cascade(label="Edit", menu=edit_menu)
        menubar.add_cascade(label="Search", menu=search_menu)
        menubar.add_cascade(label="View", menu=view_menu)
        menubar.add_cascade(label="Requirements", menu=requirements_menu)
        idx = menubar.index("end")
        self.work_product_menus.setdefault("Requirements", []).append((menubar, idx))
        menubar.entryconfig(idx, state=tk.DISABLED)
        menubar.add_cascade(label="Architecture", menu=architecture_menu)
        idx = menubar.index("end")
        self.work_product_menus.setdefault("Architecture Diagram", []).append((menubar, idx))
        menubar.entryconfig(idx, state=tk.DISABLED)
        menubar.add_cascade(label="Scenario", menu=libs_menu)
        idx = menubar.index("end")
        self.work_product_menus.setdefault("Scenario Library", []).append((menubar, idx))
        self.work_product_menus.setdefault("ODD", []).append((menubar, idx))
        menubar.entryconfig(idx, state=tk.DISABLED)
        menubar.add_cascade(label="Risk Assessment", menu=risk_menu)
        idx = menubar.index("end")
        self.work_product_menus.setdefault("Risk Assessment", []).append((menubar, idx))
        menubar.entryconfig(idx, state=tk.DISABLED)
        menubar.add_cascade(label="Qualitative Analysis", menu=qualitative_menu)
        idx = menubar.index("end")
        self.work_product_menus.setdefault("Qualitative Analysis", []).append((menubar, idx))
        menubar.entryconfig(idx, state=tk.DISABLED)
        menubar.add_cascade(label="Quantitative Analysis", menu=quantitative_menu)
        idx = menubar.index("end")
        self.work_product_menus.setdefault("Quantitative Analysis", []).append((menubar, idx))
        menubar.entryconfig(idx, state=tk.DISABLED)
        menubar.add_cascade(label="GSN", menu=gsn_menu)
        idx = menubar.index("end")
        self.work_product_menus.setdefault("GSN", []).append((menubar, idx))
        menubar.entryconfig(idx, state=tk.DISABLED)
        menubar.add_cascade(label="Process", menu=process_menu)
        idx = menubar.index("end")
        self.work_product_menus.setdefault("Process", []).append((menubar, idx))
        menubar.entryconfig(idx, state=tk.DISABLED)
        menubar.add_cascade(label="Review", menu=review_menu)
        help_menu = tk.Menu(menubar, tearoff=0)
        help_menu.add_command(label="About", command=self.lifecycle_ui.show_about)
        menubar.add_cascade(label="Help", menu=help_menu)

        root.config(menu=menubar)

        # Container to hold the auto-hiding explorer tab and main pane
        self.top_frame = tk.Frame(root)
        self.top_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True)

        self.main_pane = tk.PanedWindow(self.top_frame, orient=tk.HORIZONTAL)
        self.main_pane.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        # Initialise the log window but keep it hidden by default.
        self.log_frame = logger.init_log_window(root, height=7)
        # Status bar showing lifecycle phase and object metadata
        self.status_frame = ttk.Frame(root)
        self.status_frame.pack(side=tk.BOTTOM, fill=tk.X)
        self.toggle_log_button = ttk.Button(
            root, text="Show Logs", command=self.lifecycle_ui.toggle_logs
        )
        self.toggle_log_button.pack(side=tk.BOTTOM, fill=tk.X)
        logger.set_toggle_button(self.toggle_log_button)
        self.style.configure(
            "Phase.TLabel",
            background="#4a6ea9",
            foreground="white",
            font=("Arial", 10, "bold"),
        )
        self.active_phase_lbl = ttk.Label(
            self.status_frame, text="Active phase: None", style="Phase.TLabel"
        )
        self.active_phase_lbl.pack(side=tk.LEFT, padx=5)
        self.status_meta_vars = {
            "Name": tk.StringVar(value=""),
            "Type": tk.StringVar(value=""),
            "Author": tk.StringVar(value=""),
        }
        for key, var in self.status_meta_vars.items():
            ttk.Label(self.status_frame, text=f"{key}:").pack(
                side=tk.LEFT, padx=(10, 0)
            )
            ttk.Label(self.status_frame, textvariable=var).pack(side=tk.LEFT)

        # Explorer pane with notebook and pin button (hidden by default)
        self.explorer_pane = ttk.Frame(self.main_pane)
        self.explorer_nb = ttk.Notebook(self.explorer_pane)
        self.explorer_nb.pack(fill=tk.BOTH, expand=True)
        self._explorer_width = 300
        self._explorer_auto_hide_id = None
        self._explorer_pinned = False
        self._explorer_pin_btn = ttk.Button(
            self.explorer_pane, text="Pin", command=self.lifecycle_ui.toggle_explorer_pin
        )
        self._explorer_pin_btn.pack(anchor="ne")
        self._explorer_tab = ttk.Label(
            self.top_frame,
            text="F\ni\nl\ne\ns",
            relief="raised",
            cursor="hand2",
        )
        self._explorer_tab.pack(side=tk.LEFT, fill=tk.Y)

        self.analysis_tab = ttk.Frame(self.explorer_nb)
        self.explorer_nb.add(self.analysis_tab, text="File Explorer")

        # --- Analyses Group ---
        self.analysis_group = ttk.LabelFrame(
            self.analysis_tab, text="Analyses & Architecture", style="Toolbox.TLabelframe"
        )
        self.analysis_group.pack(fill=tk.BOTH, expand=True)

        tree_frame = ttk.Frame(self.analysis_group)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        self.analysis_tree = ttk.Treeview(tree_frame)
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.analysis_tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.analysis_tree.xview)
        self.analysis_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.analysis_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tree_frame.rowconfigure(0, weight=1)
        tree_frame.columnconfigure(0, weight=1)
        # Maintain backwards compatibility with older code referencing
        # ``self.treeview`` for the main explorer tree.
        self.treeview = self.analysis_tree

        # --- Tools Section ---
        self.tools_group = ttk.LabelFrame(
            self.analysis_tab, text="Tools", style="Toolbox.TLabelframe"
        )
        self.tools_group.pack(fill=tk.BOTH, expand=False, pady=5)
        top = ttk.Frame(self.tools_group)
        top.pack(side=tk.TOP, fill=tk.X)
        ttk.Label(top, text="Lifecycle Phase:").pack(side=tk.LEFT)
        self.lifecycle_var = tk.StringVar(value="")
        self.lifecycle_cb = ttk.Combobox(
            top, textvariable=self.lifecycle_var, state="readonly"
        )
        self.lifecycle_cb.pack(side=tk.LEFT, fill=tk.X, expand=True)

        # Container holding navigation buttons and the tools notebook
        nb_container = ttk.Frame(self.tools_group)
        nb_container.pack(fill=tk.BOTH, expand=True)
        style = ttk.Style()
        apply_translucid_button_style(style)
        # Create a custom notebook style so that a layout is available.  Without a
        # ``TNotebook`` suffix in the style name, ttk cannot find the default
        # layout which led to ``_tkinter.TclError: Layout ToolsNotebook not
        # found`` when instantiating the notebook widget.  The following styles
        # derive from the standard ``TNotebook``/``TNotebook.Tab`` styles and
        # merely customise the tab appearance.
        style.configure(
            "ToolsNotebook.TNotebook",
            padding=0,
            background="#c0d4eb",
            lightcolor="#eaf2fb",
            darkcolor="#5a6d84",
            borderwidth=2,
            relief="raised",
        )
        style.configure(
            "ToolsNotebook.TNotebook.Tab",
            font=("Arial", 10),
            padding=(10, 5),
            width=20,
            background="#b5bdc9",
            foreground="#555555",
            borderwidth=1,
            relief="raised",
        )
        style.map(
            "ToolsNotebook.TNotebook.Tab",
            background=[("selected", "#4a6ea9"), ("!selected", "#b5bdc9")],
            foreground=[("selected", "white"), ("!selected", "#555555")],
        )
        self.tools_left_btn = ttk.Button(
            nb_container, text="<", width=2, command=self.lifecycle_ui._select_prev_tool_tab
        )
        self.tools_right_btn = ttk.Button(
            nb_container, text=">", width=2, command=self.lifecycle_ui._select_next_tool_tab
        )
        self.tools_left_btn.pack(side=tk.LEFT, fill=tk.Y)
        self.tools_right_btn.pack(side=tk.RIGHT, fill=tk.Y)
        self.tools_nb = ttk.Notebook(nb_container, style="ToolsNotebook.TNotebook")
        self.tools_nb.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Track all tool tabs and which range is currently visible
        self._tool_all_tabs: list[str] = []
        self._tool_tab_offset = 0

        # Properties tab for displaying metadata
        self.prop_frame = ttk.Frame(self.tools_nb)
        self.prop_view = ttk.Treeview(
            self.prop_frame, columns=("field", "value"), show="headings"
        )
        self.prop_view.heading("field", text="Field")
        self.prop_view.heading("value", text="Value")
        # ------------------------------------------------------------------
        # NEVER DELETE OR TOUCH THIS:
        # Keep the field column fixed so the value column can expand to occupy
        # the remaining tab width, making long property values easier to read.
        # ------------------------------------------------------------------
        self.prop_view.column("field", width=120, anchor="w", stretch=False)
        self.prop_view.column("value", width=200, anchor="w", stretch=True)
        add_treeview_scrollbars(self.prop_view, self.prop_frame)
        # Bind resize handlers on the treeview, its container, and the notebook
        # via :class:`EventDispatcher` so the value column always fills the tab
        # width even before any manual resize. DO NOT REMOVE.
        self.root.after(0, self._resize_prop_columns)
        self.tools_nb.add(self.prop_frame, text="Properties")
        tab_id = self.tools_nb.tabs()[-1]
        self._tool_all_tabs.append(tab_id)
        self.lifecycle_ui._update_tool_tab_visibility()
        self._resize_prop_columns()

        # Tooltip helper for tabs (text may be clipped)
        self._tools_tip = ToolTip(self.tools_nb, "", automatic=False)

        self.tool_actions = {
            "Safety & Security Management": self.open_safety_management_toolbox,
            "Safety & Security Management Explorer": self.manage_safety_management,
            "Safety & Security Case Explorer": self.manage_safety_cases,
            "Safety Performance Indicators": self.show_safety_performance_indicators,
            "Fault Prioritization": self.open_fault_prioritization_window,
            "Cause & Effect Diagram": self.show_cause_effect_chain,
            "Diagram Rule Editor": self.open_diagram_rules_toolbox,
            "Requirement Pattern Editor": self.open_requirement_patterns_toolbox,
            "Report Template Manager": self.open_report_template_manager,
        }

        self.tool_categories: dict[str, list[str]] = {
            "Safety & Security Management": [
                "Safety & Security Management",
                "Safety & Security Management Explorer",
                "Safety & Security Case Explorer",
                "Safety Performance Indicators",
            ],
            "Safety Analysis": [
                "Fault Prioritization",
                "Cause & Effect Diagram",
                "Prototype Assurance Analysis",
            ],
            "Configuration": [
                "Diagram Rule Editor",
                "Requirement Pattern Editor",
                "Report Template Manager",
            ],
        }
        self.tool_to_work_product = {}
        for name, info in self.WORK_PRODUCT_INFO.items():
            tool_name = info[1]
            if tool_name:
                self.tool_to_work_product.setdefault(tool_name, set()).add(name)
        self.tool_to_work_product.setdefault(
            "Cause & Effect Diagram", set()
        ).add("FTA")
        self.tool_listboxes: dict[str, tk.Listbox] = {}
        self._tool_tab_titles: dict[str, str] = {}
        for cat, names in self.tool_categories.items():
            self.lifecycle_ui._add_tool_category(cat, names)

        self.pmhf_var = tk.StringVar(value="")
        self.pmhf_label = ttk.Label(self.analysis_tab, textvariable=self.pmhf_var, foreground="blue")
        self.pmhf_label.pack(side=tk.BOTTOM, fill=tk.X, pady=2)

        # Notebook for diagrams and analyses with navigation buttons
        self.doc_frame = ttk.Frame(self.main_pane)
        self.doc_nb = ClosableNotebook(self.doc_frame)
        # Mapping of tab identifiers to their full, untruncated titles.  The
        # displayed text may be shortened to keep tabs a reasonable size but we
        # keep the originals here for features like duplicate detection.
        self._tab_titles: dict[str, str] = {}
        self._doc_all_tabs: list[str] = []
        self._doc_tab_offset = 0
        _orig_select = self.doc_nb.select

        def _wrapped_select(tab_id=None):
            if tab_id is not None:
                self.lifecycle_ui._make_doc_tab_visible(tab_id)
            return _orig_select(tab_id)

        self.doc_nb.select = _wrapped_select
        self._tab_left_btn = ttk.Button(
            self.doc_frame,
            text="<",
            width=2,
            command=self.lifecycle_ui._select_prev_tab,
            style="Nav.TButton",
        )
        self._tab_right_btn = ttk.Button(
            self.doc_frame,
            text=">",
            width=2,
            command=self.lifecycle_ui._select_next_tab,
            style="Nav.TButton",
        )
        self._tab_left_btn.pack(side=tk.LEFT, fill=tk.Y)
        self._tab_right_btn.pack(side=tk.RIGHT, fill=tk.Y)
        self.doc_nb.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.lifecycle_ui._update_doc_tab_visibility()
        self.main_pane.add(self.doc_frame, stretch="always")
        # Tooltip helper for document tabs
        self._doc_tip = ToolTip(self.doc_nb, "", automatic=False)

        # Centralised event binding
        self.event_dispatcher = EventDispatcher(self)
        self.event_dispatcher.register_keyboard_shortcuts()
        self.event_dispatcher.register_tab_events()

        # Do not open the FTA tab by default so the application starts with no
        # documents visible. The tab and the initial top event will be created
        # on demand when the user opens an FTA related view or adds a top level
        # event.  This avoids the spurious "Node 1" appearing at startup.
        # Initialize the canvas related attributes so tab-close callbacks work
        # before the FTA tab has ever been created.
        self.analysis_tabs = {}
        self.canvas_tab = None
        self.canvas_frame = None
        self.canvas = None
        self.hbar = None
        self.vbar = None
        self.page_diagram = None
        self.root_node = None
        self.top_events = []
        self.fmea_entries = []
        self.fmea_service = self.safety_analysis
        self.selected_node = None
        self.dragging_node = None
        self.drag_offset_x = 0
        self.drag_offset_y = 0
        self.grid_size = 20
        self.update_views()
        # Track the last saved state so we can prompt on exit
        self.last_saved_state = json.dumps(self.export_model_data(), sort_keys=True)
        root.protocol("WM_DELETE_WINDOW", self.confirm_close)
        self.use_case_windows = []
        self.activity_windows = []
        self.block_windows = []
        self.ibd_windows = []

    # ------------------------------------------------------------------
    # UI lifecycle helper wrappers
    # ------------------------------------------------------------------
    def show_properties(self, *args, **kwargs):
        return self.lifecycle_ui.show_properties(*args, **kwargs)

    def _add_tool_category(self, *args, **kwargs):
        return self.lifecycle_ui._add_tool_category(*args, **kwargs)

    def _add_lifecycle_requirements_menu(self, *args, **kwargs):
        return self.lifecycle_ui._add_lifecycle_requirements_menu(*args, **kwargs)

    def _init_nav_button_style(self, *args, **kwargs):
        return self.lifecycle_ui._init_nav_button_style(*args, **kwargs)

    def _limit_explorer_size(self, *args, **kwargs):
        return self.lifecycle_ui._limit_explorer_size(*args, **kwargs)

    def _animate_explorer_show(self, *args, **kwargs):
        return self.lifecycle_ui._animate_explorer_show(*args, **kwargs)

    def _animate_explorer_hide(self, *args, **kwargs):
        return self.lifecycle_ui._animate_explorer_hide(*args, **kwargs)

    def _schedule_explorer_hide(self, *args, **kwargs):
        return self.lifecycle_ui._schedule_explorer_hide(*args, **kwargs)

    def _cancel_explorer_hide(self, *args, **kwargs):
        return self.lifecycle_ui._cancel_explorer_hide(*args, **kwargs)

    def show_explorer(self, *args, **kwargs):
        return self.lifecycle_ui.show_explorer(*args, **kwargs)

    def hide_explorer(self, *args, **kwargs):
        return self.lifecycle_ui.hide_explorer(*args, **kwargs)

    def toggle_explorer_pin(self, *args, **kwargs):
        return self.lifecycle_ui.toggle_explorer_pin(*args, **kwargs)

    def toggle_logs(self, *args, **kwargs):
        return self.lifecycle_ui.toggle_logs(*args, **kwargs)

    def open_metrics_tab(self, *args, **kwargs):
        return self.lifecycle_ui.open_metrics_tab(*args, **kwargs)

    def open_management_window(self, *args, **kwargs):
        return self.open_windows_features.open_management_window(*args, **kwargs)

    def _register_close(self, *args, **kwargs):
        return self.lifecycle_ui._register_close(*args, **kwargs)

    def _reregister_document(self, *args, **kwargs):
        return self.lifecycle_ui._reregister_document(*args, **kwargs)

    def touch_doc(self, *args, **kwargs):
        return self.lifecycle_ui.touch_doc(*args, **kwargs)

    def show_about(self, *args, **kwargs):
        return self.lifecycle_ui.show_about(*args, **kwargs)

    def _window_has_focus(self, *args, **kwargs):
        return self.lifecycle_ui._window_has_focus(*args, **kwargs)

    def _window_in_selected_tab(self, *args, **kwargs):
        return self.lifecycle_ui._window_in_selected_tab(*args, **kwargs)

    def _on_tab_change(self, *args, **kwargs):
        return self.lifecycle_ui._on_tab_change(*args, **kwargs)

    def _on_tab_close(self, *args, **kwargs):
        return self.lifecycle_ui._on_tab_close(*args, **kwargs)

    def _on_doc_tab_motion(self, *args, **kwargs):
        return self.lifecycle_ui._on_doc_tab_motion(*args, **kwargs)

    def _on_tool_tab_motion(self, *args, **kwargs):
        return self.lifecycle_ui._on_tool_tab_motion(*args, **kwargs)

    def _make_doc_tab_visible(self, *args, **kwargs):
        return self.lifecycle_ui._make_doc_tab_visible(*args, **kwargs)

    def _update_doc_tab_visibility(self, *args, **kwargs):
        return self.lifecycle_ui._update_doc_tab_visibility(*args, **kwargs)

    def _update_tool_tab_visibility(self, *args, **kwargs):
        return self.lifecycle_ui._update_tool_tab_visibility(*args, **kwargs)

    def _truncate_tab_title(self, *args, **kwargs):
        return self.lifecycle_ui._truncate_tab_title(*args, **kwargs)

    def _select_next_tab(self, *args, **kwargs):
        return self.lifecycle_ui._select_next_tab(*args, **kwargs)

    def _select_prev_tab(self, *args, **kwargs):
        return self.lifecycle_ui._select_prev_tab(*args, **kwargs)

    def _select_next_tool_tab(self, *args, **kwargs):
        return self.lifecycle_ui._select_next_tool_tab(*args, **kwargs)

    def _select_prev_tool_tab(self, *args, **kwargs):
        return self.lifecycle_ui._select_prev_tool_tab(*args, **kwargs)

    def _new_tab(self, *args, **kwargs):
        return self.lifecycle_ui._new_tab(*args, **kwargs)

    # ------------------------------------------------------------------
    # Label editing and styling helper wrappers
    # ------------------------------------------------------------------
    def edit_controllability(self, *args, **kwargs):
        return self.labels_styling.edit_controllability(*args, **kwargs)

    def edit_description(self, *args, **kwargs):
        return self.labels_styling.edit_description(*args, **kwargs)

    def edit_gate_type(self, *args, **kwargs):
        return self.labels_styling.edit_gate_type(*args, **kwargs)

    def edit_page_flag(self, *args, **kwargs):
        return self.labels_styling.edit_page_flag(*args, **kwargs)

    def edit_rationale(self, *args, **kwargs):
        return self.labels_styling.edit_rationale(*args, **kwargs)

    def edit_selected(self, *args, **kwargs):
        return self.labels_styling.edit_selected(*args, **kwargs)

    def edit_user_name(self, *args, **kwargs):
        return self.labels_styling.edit_user_name(*args, **kwargs)

    def edit_value(self, *args, **kwargs):
        return self.labels_styling.edit_value(*args, **kwargs)

    def rename_failure(self, *args, **kwargs):
        return self.labels_styling.rename_failure(*args, **kwargs)

    def rename_fault(self, *args, **kwargs):
        return self.labels_styling.rename_fault(*args, **kwargs)

    def rename_functional_insufficiency(self, *args, **kwargs):
        return self.labels_styling.rename_functional_insufficiency(*args, **kwargs)

    def rename_hazard(self, *args, **kwargs):
        return self.labels_styling.rename_hazard(*args, **kwargs)

    def rename_malfunction(self, *args, **kwargs):
        return self.labels_styling.rename_malfunction(*args, **kwargs)

    def rename_selected_tree_item(self, *args, **kwargs):
        return self.labels_styling.rename_selected_tree_item(*args, **kwargs)

    def rename_triggering_condition(self, *args, **kwargs):
        return self.labels_styling.rename_triggering_condition(*args, **kwargs)

    def apply_style(self, *args, **kwargs):
        return self.labels_styling.apply_style(*args, **kwargs)

    def format_failure_mode_label(self, *args, **kwargs):
        return self.labels_styling.format_failure_mode_label(*args, **kwargs)

    def _resize_prop_columns(self, *args, **kwargs):
        return self.labels_styling._resize_prop_columns(*args, **kwargs)

    def _spi_label(self, *args, **kwargs):
        return self.labels_styling._spi_label(*args, **kwargs)

    def _product_goal_name(self, *args, **kwargs):
        return self.labels_styling._product_goal_name(*args, **kwargs)

    def _create_icon(self, *args, **kwargs):
        return self.labels_styling._create_icon(*args, **kwargs)

    @property
    def fmeas(self):
        return self.safety_analysis.fmeas

    @fmeas.setter
    def fmeas(self, value):
        self.safety_analysis.fmeas = value

    def show_fmea_list(self):
        """Delegate to the safety analysis facade to display FMEA manager."""
        self.safety_analysis.show_fmea_list()

    def show_fmea_table(self, fmea=None, fmeda=False):
        """Delegate to safety analysis for rendering FMEA/FMeda tables."""
        return self.safety_analysis.show_fmea_table(fmea, fmeda)

    def show_fmeda_list(self):
        """Open the FMEDA document manager via the facade."""
        self.safety_analysis.show_fmeda_list()

        # --- Requirement Traceability Helpers used by reviews and matrix view ---
    def get_requirement_allocation_names(self, req_id):
        return self.requirements_manager.get_requirement_allocation_names(req_id)

    def get_requirement_goal_names(self, req_id):
        return self.requirements_manager.get_requirement_goal_names(req_id)

    def format_requirement(self, req, include_id=True):
        """Return a formatted requirement string without empty ASIL/CAL fields."""
        return self.requirements_manager.format_requirement(req, include_id)

    def format_requirement_with_trace(self, req):
        return self.requirements_manager.format_requirement_with_trace(req)

    def build_requirement_diff_html(self, review):
        return self.reporting_export.build_requirement_diff_html(review)

    def generate_recommendations_for_top_event(self, node):
        return self.safety_analysis.generate_recommendations_for_top_event(node)

    def back_all_pages(self):
        return self.nav_input.back_all_pages()

    def move_top_event_up(self):
        return self.safety_analysis.move_top_event_up()

    def move_top_event_down(self):
        return self.safety_analysis.move_top_event_down()

    def get_top_level_nodes(self):
        return self.data_access_queries.get_top_level_nodes()

    def get_all_nodes_no_filter(self, node):
        return self.structure_tree_operations.get_all_nodes_no_filter(node)

    def derive_requirements_for_event(self, event):
        return self.safety_analysis.derive_requirements_for_event(event)

    def get_combined_safety_requirements(self, node):
        return self.fta_app.get_combined_safety_requirements(self, node)

    def get_top_event(self, node):
        return self.safety_analysis.get_top_event(node)

    def aggregate_safety_requirements(self, node, all_nodes):
        return self.fta_app.aggregate_safety_requirements(self, node, all_nodes)

    def generate_top_event_summary(self, top_event):
        return self.safety_analysis.generate_top_event_summary(top_event)

    def get_all_nodes(self, node=None):
        return self.structure_tree_operations.get_all_nodes(node)

    def get_all_nodes_table(self, root_node):
        return self.structure_tree_operations.get_all_nodes_table(root_node)

    def get_all_nodes_in_model(self):
        return self.structure_tree_operations.get_all_nodes_in_model()

    def get_all_basic_events(self):
        return self.safety_analysis.get_all_basic_events()

    def get_all_gates(self):
        return self.safety_analysis.get_all_gates()

    def metric_to_text(self, metric_type, value):
        return self.probability_reliability.metric_to_text(metric_type, value)

    def assurance_level_text(self, level):
        return self.probability_reliability.assurance_level_text(level)

    def calculate_cut_sets(self, node):
        return self.safety_analysis.calculate_cut_sets(node)

    def build_hierarchical_argumentation(self, node, indent=0):
        return self.fta_app.build_hierarchical_argumentation(self, node, indent)

    def build_hierarchical_argumentation_common(self, node, indent=0, described=None):
        return self.fta_app.build_hierarchical_argumentation_common(self, node, indent, described)

    def build_page_argumentation(self, page_node):
        return self.fta_app.build_page_argumentation(self, page_node)

    def build_unified_recommendation_table(self):
        return self.reporting_export.build_unified_recommendation_table()

    def build_dynamic_recommendations_table(self):
        return self.reporting_export.build_dynamic_recommendations_table()

    def build_base_events_table_html(self):
        return self.reporting_export.build_base_events_table_html()

    def get_extra_recommendations_list(self, description, level):
        return self.data_access_queries.get_extra_recommendations_list(description, level)

    def get_extra_recommendations_from_level(self, description, level):
        return self.data_access_queries.get_extra_recommendations_from_level(description, level)

    def get_recommendation_from_description(self, description, level):
        return self.data_access_queries.get_recommendation_from_description(description, level)

    def build_argumentation(self, node):
        return self.fta_app.build_argumentation(self, node)

    def auto_create_argumentation(self, node, suppress_top_event_recommendations=False):
        return self.fta_app.auto_create_argumentation(self, node, suppress_top_event_recommendations)

    def analyze_common_causes(self, node):
        return self.probability_reliability.analyze_common_causes(node)

    def build_text_report(self, node, indent=0):
        return self.reporting_export.build_text_report(node, indent)

    def all_children_are_base_events(self, node):
        return self.safety_analysis.all_children_are_base_events(node)

    def build_simplified_fta_model(self, top_event):
        return self.safety_analysis.build_simplified_fta_model(top_event)

    @staticmethod
    def auto_generate_fta_diagram(fta_model, output_path):
        return FTASubApp.auto_generate_fta_diagram(fta_model, output_path)
        
    def find_node_by_id_all(self, unique_id):
        return self.structure_tree_operations.find_node_by_id_all(unique_id)

    def get_hazop_by_name(self, name):
        return self.safety_analysis.get_hazop_by_name(name)

    def get_hara_by_name(self, name):
        return self.safety_analysis.get_hara_by_name(name)

    def update_hara_statuses(self):
        return self.safety_analysis.update_hara_statuses()

    def update_fta_statuses(self):
        return self.safety_analysis.update_fta_statuses()

    def get_safety_goal_asil(self, sg_name):
        return self.safety_analysis.get_safety_goal_asil(sg_name)

    def get_hara_goal_asil(self, sg_name):
        return self.safety_analysis.get_hara_goal_asil(sg_name)

    def get_cyber_goal_cal(self, goal_id):
        return self.safety_analysis.get_cyber_goal_cal(goal_id)

    def get_top_event_safety_goals(self, node):
        return self.safety_analysis.get_top_event_safety_goals(node)

    def get_safety_goals_for_malfunctions(self, malfunctions: list[str]) -> list[str]:
        return self.safety_analysis.get_safety_goals_for_malfunctions(malfunctions)

    def is_malfunction_used(self, name: str) -> bool:
        return self.safety_analysis.is_malfunction_used(name)

    def add_malfunction(self, name: str) -> None:
        return self.safety_analysis.add_malfunction(name)

    def add_fault(self, name: str) -> None:
        return self.safety_analysis.add_fault(name)

    def add_failure(self, name: str) -> None:
        return self.safety_analysis.add_failure(name)

    def add_hazard(self, name: str, severity: int | str = 1) -> None:
        return self.safety_analysis.add_hazard(name, severity)

    def add_triggering_condition(self, name: str) -> None:
        return self.safety_analysis.add_triggering_condition(name)

    def delete_triggering_condition(self, name: str) -> None:
        return self.safety_analysis.delete_triggering_condition(name)

    def add_functional_insufficiency(self, name: str) -> None:
        return self.safety_analysis.add_functional_insufficiency(name)

    def delete_functional_insufficiency(self, name: str) -> None:
        return self.safety_analysis.delete_functional_insufficiency(name)


    def _update_shared_product_goals(self):
        events = self.top_events + getattr(self, "cta_events", []) + getattr(self, "paa_events", [])
        self.shared_product_goals = self.product_goal_manager.update_shared_product_goals(
            events, getattr(self, "shared_product_goals", {})
        )


    def update_hazard_severity(self, hazard: str, severity: int | str) -> None:
        return self.safety_analysis.update_hazard_severity(hazard, severity)


    def calculate_fmeda_metrics(self, events):
        """Delegate FMEDA metric calculation to the safety analysis facade."""
        return self.safety_analysis.calculate_fmeda_metrics(events)

    def compute_fmeda_metrics(self, events):
        """Delegate detailed FMEDA metric computation to the facade."""
        return self.safety_analysis.compute_fmeda_metrics(events)

    def sync_hara_to_safety_goals(self):
        """Synchronise HARA results with safety goals via the risk sub-app."""
        return self.risk_app.sync_hara_to_safety_goals(self)

    def sync_cyber_risk_to_goals(self):
        return self.probability_reliability.sync_cyber_risk_to_goals()

    def add_top_level_event(self):
        return self.safety_analysis.add_top_level_event()

    def _build_probability_frame(self, parent, title: str, levels: range, values: dict, row: int, dialog_font):
        return self.probability_reliability._build_probability_frame(parent, title, levels, values, row, dialog_font)

    def edit_project_properties(self) -> None:
        """Open the project properties dialog."""
        ProjectPropertiesDialog(self).show()

    def create_diagram_image(self):  # pragma: no cover - delegation
        return self.diagram_renderer.create_diagram_image()

    def get_page_nodes(self, node):
        return self.data_access_queries.get_page_nodes(node)

    def capture_page_diagram(self, page_node):  # pragma: no cover - delegation
        return self.diagram_renderer.capture_page_diagram(page_node)

    def capture_event_diagram(self, *args, **kwargs):  # pragma: no cover - delegation
        return self.diagram_renderer.capture_event_diagram(*args, **kwargs)

    def capture_gsn_diagram(self, diagram):
        return self.diagram_renderer.capture_gsn_diagram(diagram)

    def capture_sysml_diagram(self, diagram):
        return self.diagram_renderer.capture_sysml_diagram(diagram)

    def capture_cbn_diagram(self, doc):
        return self.diagram_renderer.capture_cbn_diagram(doc)
    
    def draw_subtree_with_filter(self, canvas, root_event, visible_nodes):
        return self.diagram_renderer.draw_subtree_with_filter(canvas, root_event, visible_nodes)

    def draw_subtree(self, canvas, root_event):
        return self.diagram_renderer.draw_subtree(canvas, root_event)

    def draw_connections_subtree(self, canvas, node, drawn_ids):
        return self.diagram_renderer.draw_connections_subtree(canvas, node, drawn_ids)

    def draw_node_on_canvas_pdf(self, canvas, node):
        return self.diagram_renderer.draw_node_on_canvas_pdf(canvas, node)

    def rename_selected_tree_item(self):
        self.tree_app.rename_selected_tree_item(self)

    def save_diagram_png(self):  # pragma: no cover - delegation
        return self.diagram_renderer.save_diagram_png()

    def on_treeview_click(self, event):
        return self.nav_input.on_treeview_click(event)

    def on_analysis_tree_double_click(self, event):
        return self.nav_input.on_analysis_tree_double_click(event)

    def on_analysis_tree_right_click(self, event):
        return self.nav_input.on_analysis_tree_right_click(event)

    def on_analysis_tree_select(self, _event):
        return self.nav_input.on_analysis_tree_select(_event)

    def on_tool_list_double_click(self, event):
        return self.nav_input.on_tool_list_double_click(event)

    def _on_toolbox_change(self) -> None:
        self.governance_manager._on_toolbox_change()

    def apply_governance_rules(self) -> None:
        """Apply governance rules and refresh related UI elements."""
        self.governance_manager.apply_governance_rules()

    def refresh_tool_enablement(self) -> None:
        self.governance_manager.refresh_tool_enablement()

    def update_lifecycle_cb(self) -> None:
        self.governance_manager.update_lifecycle_cb()

    def _export_toolbox_dict(self) -> dict:
        toolbox = getattr(self, "safety_mgmt_toolbox", None)
        if toolbox is None:
            toolbox = SafetyManagementToolbox()
            self.governance_manager.attach_toolbox(toolbox)
            self.safety_mgmt_toolbox = toolbox
        return toolbox.to_dict()

    def on_lifecycle_selected(self, _event=None) -> None:
        phase = self.lifecycle_var.get()
        gm = getattr(self, "governance_manager", None)
        if gm is None:
            gm = GovernanceManager(self)
            self.governance_manager = gm
            gm.attach_toolbox(getattr(self, "safety_mgmt_toolbox", None))
        gm.on_lifecycle_selected(phase)


    def enable_process_area(self, area: str) -> None:  # pragma: no cover - delegation
        return self.validation_consistency.enable_process_area(area)

    def enable_work_product(self, name: str, *, refresh: bool = True) -> None:  # pragma: no cover - delegation
        return self.validation_consistency.enable_work_product(name, refresh=refresh)

    def can_remove_work_product(self, name: str) -> bool:  # pragma: no cover - delegation
        return self.validation_consistency.can_remove_work_product(name)

    def disable_work_product(self, name: str, *, force: bool = False, refresh: bool = True) -> bool:  # pragma: no cover - delegation
        return self.validation_consistency.disable_work_product(name, force=force, refresh=refresh)

    def open_work_product(self, name: str) -> None:
        """Open a diagram or analysis work product within the application."""
        wp = next(
            (wp for wp, info in self.WORK_PRODUCT_INFO.items() if info[1] == name or wp == name),
            None,
        )
        global_enabled = getattr(self, "enabled_work_products", set())
        smt = getattr(self, "safety_mgmt_toolbox", None)
        if smt and getattr(smt, "work_products", None):
            phase_enabled = smt.enabled_products()
        else:
            phase_enabled = global_enabled
        if wp and wp not in (global_enabled & phase_enabled):
            return
        action = self.tool_actions.get(name)
        if callable(action):
            action()
            return

        for diag in self.arch_diagrams:
            if getattr(diag, "name", "") == name or getattr(diag, "diag_id", "") == name:
                self.window_controllers.open_arch_window(diag.diag_id)
                return

        for idx, diag in enumerate(self.management_diagrams):
            if getattr(diag, "name", "") == name or getattr(diag, "diag_id", "") == name:
                self.open_windows_features.open_management_window(idx)
                return

        for diag in getattr(self, "all_gsn_diagrams", []):
            if getattr(diag.root, "user_name", "") == name or getattr(diag, "diag_id", "") == name:
                self.window_controllers.open_gsn_diagram(diag)
                return


    # ----------------------------------------------------------------------
    # NEVER DELETE OR TOUCH THIS: helper keeps the value column synced with
    # the Properties tab width. Removing this breaks property display.
    # ----------------------------------------------------------------------


    def on_ctrl_mousewheel(self, event):
        return self.nav_input.on_ctrl_mousewheel(event)

    def new_model(self):
        self.project_manager.new_model()

    def compute_occurrence_counts(self):
        counts = {}
        visited = set()

        if self.root_node is None:
            return counts

        def rec(node):
            if node.unique_id in visited:
                counts[node.unique_id] += 1
            else:
                visited.add(node.unique_id)
                counts[node.unique_id] = 1
            for child in node.children:
                rec(child)

        rec(self.root_node)
        return counts

    def get_node_fill_color(self, node, mode: str | None = None):
        """Return the fill color for *node* based on analysis mode.

        Parameters
        ----------
        node: FaultTreeNode | None
            Unused but kept for API compatibility.
        mode: str | None
            Explicit diagram mode to use.  Falls back to the currently
            focused canvas' ``diagram_mode`` when ``None``.
        """

        diagram_mode = mode or getattr(getattr(self, "canvas", None), "diagram_mode", "FTA")
        if diagram_mode == "CTA":
            return "#EE82EE"
        if diagram_mode == "PAA":
            return "#40E0D0"
        return "#FAD7A0"

    def on_right_mouse_press(self, event):
        return self.nav_input.on_right_mouse_press(event)

    def on_right_mouse_drag(self, event):
        return self.nav_input.on_right_mouse_drag(event)

    def on_right_mouse_release(self, event):
        return self.nav_input.on_right_mouse_release(event)

    def show_context_menu(self, event):
        return self.nav_input.show_context_menu(event)

    def on_canvas_click(self, event):
        return self.nav_input.on_canvas_click(event)

    def on_canvas_double_click(self, event):
        return self.nav_input.on_canvas_double_click(event)

    def on_canvas_drag(self, event):
        return self.nav_input.on_canvas_drag(event)

    def on_canvas_release(self, event):
        return self.nav_input.on_canvas_release(event)

    def move_subtree(self, node, dx, dy):
        return self.structure_tree_operations.move_subtree(node, dx, dy)

    def zoom_in(self):  # pragma: no cover - delegation
        return self.diagram_renderer.zoom_in()

    def zoom_out(self):  # pragma: no cover - delegation
        return self.diagram_renderer.zoom_out()


    # ------------------------------------------------------------------
    # Explorer panel show/hide helpers

    def auto_arrange(self):
        if self.root_node is None:
            return
        horizontal_gap = 150
        vertical_gap = 100
        next_y = [100]
        def layout(node, depth):
            node.x = depth * horizontal_gap + 100
            if not node.children:
                node.y = next_y[0]
                next_y[0] += vertical_gap
            else:
                for child in node.children:
                    layout(child, depth+1)
                node.y = (node.children[0].y + node.children[-1].y) / 2
        layout(self.root_node, 0)
        # --- Center the layout horizontally on the canvas ---
        all_nodes = self.get_all_nodes(self.root_node)
        if all_nodes:
            min_x = min(n.x for n in all_nodes)
            max_x = max(n.x for n in all_nodes)
            canvas_width = self.canvas.winfo_width()
            if canvas_width < 10:
                canvas_width = 800
            diagram_width = max_x - min_x
            offset = (canvas_width / self.zoom - diagram_width) / 2 - min_x
            for n in all_nodes:
                n.x += offset
        self.update_views()


    def get_all_triggering_conditions(self):
        return self.data_access_queries.get_all_triggering_conditions()

    def get_all_functional_insufficiencies(self):
        return self.data_access_queries.get_all_functional_insufficiencies()

    def get_all_scenario_names(self):
        return self.data_access_queries.get_all_scenario_names()

    def get_validation_targets_for_odd(self, element_name):
        """Return product goals linked to scenarios using ``element_name``.

        The search traverses scenario libraries, HAZOP documents and risk
        assessment entries to locate safety goals whose top events contain
        validation targets. The returned list contains the matching top event
        nodes so their validation data can be displayed.
        """
        scenarios = set()
        for lib in self.scenario_libraries:
            for sc in lib.get("scenarios", []):
                if isinstance(sc, dict):
                    name = sc.get("name", "")
                    scenery = sc.get("scenery", "")
                    desc = sc.get("description", "")
                else:
                    name = sc
                    scenery = ""
                    desc = ""
                elems = {e.strip() for e in str(scenery).split(",") if e}
                if desc:
                    elems.update(re.findall(r"\[\[(.+?)\]\]", str(desc)))
                if element_name and name and element_name in elems:
                    scenarios.add(name)

        if not scenarios:
            return []

        hazop_scenarios = set()
        for doc in self.hazop_docs:
            for entry in doc.entries:
                if getattr(entry, "scenario", "") in scenarios:
                    hazop_scenarios.add(entry.scenario)

        if not hazop_scenarios:
            return []

        goals = []
        seen = set()
        for doc in self.hara_docs:
            for entry in doc.entries:
                if getattr(entry, "scenario", "") in hazop_scenarios:
                    sg_name = getattr(entry, "safety_goal", "")
                    for te in self.top_events:
                        name = te.safety_goal_description or (
                            te.user_name or f"SG {te.unique_id}"
                        )
                        if name == sg_name and sg_name not in seen:
                            goals.append(te)
                            seen.add(sg_name)
        return goals

    def get_scenario_exposure(self, name: str) -> int:
        return self.data_access_queries.get_scenario_exposure(name)

    def get_all_scenery_names(self):
        return self.data_access_queries.get_all_scenery_names()


    def get_all_function_names(self):
        return self.data_access_queries.get_all_function_names()

    def get_all_action_names(self):
        return self.data_access_queries.get_all_action_names()

    def get_all_action_labels(self) -> list[str]:
        return self.data_access_queries.get_all_action_labels()

    def get_use_case_for_function(self, func: str) -> str:
        return self.data_access_queries.get_use_case_for_function(func)

    def get_all_component_names(self):
        return self.data_access_queries.get_all_component_names()

    def get_all_part_names(self) -> list[str]:
        return self.data_access_queries.get_all_part_names()

    def get_all_malfunction_names(self):
        return self.data_access_queries.get_all_malfunction_names()

    def get_hazards_for_malfunction(self, malfunction: str, hazop_names=None) -> list[str]:
        """Return hazards linked to the malfunction in the given HAZOPs."""
        hazards: list[str] = []
        names = hazop_names or [d.name for d in self.hazop_docs]
        for hz_name in names:
            doc = self.get_hazop_by_name(hz_name)
            if not doc:
                continue
            for entry in doc.entries:
                if entry.malfunction == malfunction:
                    h = getattr(entry, "hazard", "").strip()
                    if h and h not in hazards:
                        hazards.append(h)
        return hazards

    def update_odd_elements(self):
        return self.safety_analysis.update_odd_elements()

    def update_hazard_list(self):
        return self.safety_analysis.update_hazard_list()

    def update_failure_list(self):
        return self.safety_analysis.update_failure_list()

    def update_triggering_condition_list(self):
        return self.safety_analysis.update_triggering_condition_list()

    def update_functional_insufficiency_list(self):
        return self.safety_analysis.update_functional_insufficiency_list()

    def get_entry_field(self, entry, field, default=""):
        return self.data_access_queries.get_entry_field(entry, field, default)

    def get_all_failure_modes(self):
        return self.safety_analysis.get_all_failure_modes()

    def get_all_fmea_entries(self):
        return self.safety_analysis.get_all_fmea_entries()

    def get_non_basic_failure_modes(self):
        return self.safety_analysis.get_non_basic_failure_modes()

    def get_available_failure_modes_for_gates(self, current_gate=None):
        return self.safety_analysis.get_available_failure_modes_for_gates(current_gate)

    def get_failure_mode_node(self, node):
        return self.safety_analysis.get_failure_mode_node(node)

    def get_component_name_for_node(self, node):
        return self.safety_analysis.get_component_name_for_node(node)

    def get_failure_modes_for_malfunction(self, malfunction: str) -> list[str]:
        return self.safety_analysis.get_failure_modes_for_malfunction(malfunction)

    def get_faults_for_failure_mode(self, failure_mode_node) -> list[str]:
        return self.safety_analysis.get_faults_for_failure_mode(failure_mode_node)

    def get_fit_for_fault(self, fault_name: str) -> float:
        return self.safety_analysis.get_fit_for_fault(fault_name)

    def update_views(self):
        """Refresh project views via the dedicated :class:`ViewUpdater`."""
        self.view_updater.update_views()

    def update_basic_event_probabilities(self):
        return self.safety_analysis.update_basic_event_probabilities()

    def validate_float(self, value):
        """Return ``True`` if ``value`` resembles a float.

        This validator is tolerant of scientific-notation inputs that are
        entered incrementally (e.g. ``"1e"`` or ``"1e-"``) to keep the entry
        widget from rejecting keystrokes during editing.
        """

        if value in ("", "-", "+", ".", "-.", "+.", "e", "E", "e-", "e+", "E-", "E+"):
            return True
        try:
            float(value)
            return True
        except ValueError:
            lower = value.lower()
            if lower.endswith("e"):
                try:
                    float(lower[:-1])
                    return True
                except ValueError:
                    return False
            if lower.endswith(("e-", "e+")):
                try:
                    float(lower[:-2])
                    return True
                except ValueError:
                    return False
            return False

    def compute_failure_prob(self, node, failure_mode_ref=None, formula=None):
        return self.probability_reliability.compute_failure_prob(node, failure_mode_ref=failure_mode_ref, formula=formula)


    def propagate_failure_mode_attributes(self, fm_node):
        """Update basic events referencing ``fm_node`` and recompute probability."""
        for be in self.get_all_basic_events():
            if getattr(be, "failure_mode_ref", None) == fm_node.unique_id:
                be.fmeda_fit = fm_node.fmeda_fit
                be.fmeda_diag_cov = fm_node.fmeda_diag_cov
                # Always propagate the formula so edits take effect
                be.prob_formula = fm_node.prob_formula
                be.failure_prob = self.compute_failure_prob(be)


    def refresh_model(self):
        return self.safety_analysis.refresh_model()

    def refresh_all(self):
        return self.safety_analysis.refresh_all()

    def insert_node_in_tree(self, parent_item, node):
        return self.structure_tree_operations.insert_node_in_tree(parent_item, node)

    def redraw_canvas(self):
        return self.diagram_renderer.redraw_canvas()

    def create_diagram_image_without_grid(self):
        return self.diagram_renderer.create_diagram_image_without_grid()

    def draw_connections(self, node, drawn_ids=set()):
        return self.diagram_renderer.draw_connections(node, drawn_ids)

    def draw_node(self, node):
        return self.diagram_renderer.draw_node(node)

    def find_node_by_id(self, node, unique_id, visited=None):
        return self.structure_tree_operations.find_node_by_id(node, unique_id, visited)

    def is_descendant(self, node, possible_ancestor):
        return self.structure_tree_operations.is_descendant(node, possible_ancestor)

    def add_node_of_type(self, event_type):
        self.push_undo_state()
        diag_mode = getattr(self, "diagram_mode", "FTA")
        if diag_mode == "PAA":
            allowed = {"CONFIDENCE LEVEL", "ROBUSTNESS SCORE"}
            if event_type.upper() not in allowed:
                messagebox.showwarning(
                    "Invalid",
                    "Only Confidence and Robustness nodes are allowed in Prototype Assurance Analysis.",
                )
                return
        else:
            if diag_mode == "CTA":
                allowed = {"TRIGGERING CONDITION", "FUNCTIONAL INSUFFICIENCY"}
            else:
                allowed = {"GATE", "BASIC EVENT"}
            if event_type.upper() not in allowed:
                messagebox.showwarning(
                    "Invalid",
                    f"Node type '{event_type}' is not allowed in {diag_mode} diagrams.",
                )
                return
        # If a node is selected, ensure it is a primary instance.
        if self.selected_node:
            if not self.selected_node.is_primary_instance:
                messagebox.showwarning("Invalid Operation",
                    "Cannot add new elements to a clone node.\nPlease select the original node instead.")
                return
            parent_node = self.selected_node
        else:
            sel = self.analysis_tree.selection()
            if sel:
                try:
                    node_id = int(self.analysis_tree.item(sel[0], "tags")[0])
                except (IndexError, ValueError):
                    messagebox.showwarning("No selection", "Select a parent node from the tree.")
                    return
                parent_node = self.find_node_by_id_all(node_id)
            else:
                messagebox.showwarning("No selection", "Select a parent node to paste into.")
                return

        # Prevent adding to base events.
        if parent_node.node_type.upper() in ["CONFIDENCE LEVEL", "ROBUSTNESS SCORE", "BASIC EVENT"]:
            messagebox.showwarning("Invalid", "Base events cannot have children.")
            return

        # Now create the new node.
        if event_type.upper() == "CONFIDENCE LEVEL":
            new_node = FaultTreeNode("", "Confidence Level", parent=parent_node)
            new_node.quant_value = 1
        elif event_type.upper() == "ROBUSTNESS SCORE":
            new_node = FaultTreeNode("", "Robustness Score", parent=parent_node)
            new_node.quant_value = 1
        elif event_type.upper() == "GATE":
            new_node = FaultTreeNode("", "GATE", parent=parent_node)
            new_node.gate_type = "AND"
        elif event_type.upper() == "BASIC EVENT":
            new_node = FaultTreeNode("", "Basic Event", parent=parent_node)
            new_node.failure_prob = 0.0
        elif event_type.upper() == "TRIGGERING CONDITION":
            new_node = FaultTreeNode("", "Triggering Condition", parent=parent_node)
        elif event_type.upper() == "FUNCTIONAL INSUFFICIENCY":
            new_node = FaultTreeNode("", "Functional Insufficiency", parent=parent_node)
            new_node.gate_type = "AND"
        else:
            new_node = FaultTreeNode("", event_type, parent=parent_node)
        new_node.x = parent_node.x + 100
        new_node.y = parent_node.y + 100
        parent_node.children.append(new_node)
        new_node.parents.append(parent_node)
        self.update_views()
        # Capture the post-addition state so future moves can be undone back
        # to this initial location.
        self.push_undo_state()

    def add_basic_event_from_fmea(self):
        self.push_undo_state()
        events = list(self.fmea_entries)
        for doc in self.fmeas:
            events.extend(doc.get("entries", []))
        for doc in self.fmedas:
            events.extend(doc.get("entries", []))
        if not events:
            messagebox.showinfo("No Failure Modes", "No FMEA or FMEDA failure modes available.")
            return
        dialog = SelectBaseEventDialog(self.root, events)
        selected = dialog.selected
        if not selected:
            return
        if self.selected_node:
            parent_node = self.selected_node
            if not parent_node.is_primary_instance:
                messagebox.showwarning("Invalid Operation", "Cannot add to a clone node. Select the original.")
                return
        else:
            sel = self.analysis_tree.selection()
            if not sel:
                messagebox.showwarning("No selection", "Select a parent node to paste into.")
                return
            try:
                node_id = int(self.analysis_tree.item(sel[0], "tags")[0])
            except (IndexError, ValueError):
                messagebox.showwarning("No selection", "Select a parent node from the tree.")
                return
            parent_node = self.find_node_by_id_all(node_id)
        if parent_node.node_type.upper() in ["CONFIDENCE LEVEL", "ROBUSTNESS SCORE", "BASIC EVENT"]:
            messagebox.showwarning("Invalid", "Base events cannot have children.")
            return
        data = selected.to_dict()
        data.pop("unique_id", None)
        data["children"] = []
        new_node = FaultTreeNode.from_dict(data, parent_node)
        if hasattr(selected, "unique_id"):
            new_node.failure_mode_ref = selected.unique_id
        parent_node.children.append(new_node)
        new_node.parents.append(parent_node)
        self.update_views()

    def add_basic_event_from_fmea(self):
        self.push_undo_state()
        events = list(self.fmea_entries)
        for doc in self.fmeas:
            events.extend(doc.get("entries", []))
        for doc in self.fmedas:
            events.extend(doc.get("entries", []))
        if not events:
            messagebox.showinfo("No Failure Modes", "No FMEA or FMEDA failure modes available.")
            return
        dialog = SelectBaseEventDialog(self.root, events)
        selected = dialog.selected
        if not selected:
            return
        if self.selected_node:
            parent_node = self.selected_node
            if not parent_node.is_primary_instance:
                messagebox.showwarning("Invalid Operation", "Cannot add to a clone node. Select the original.")
                return
        else:
            sel = self.analysis_tree.selection()
            if not sel:
                messagebox.showwarning("No selection", "Select a parent node to paste into.")
                return
            try:
                node_id = int(self.analysis_tree.item(sel[0], "tags")[0])
            except (IndexError, ValueError):
                messagebox.showwarning("No selection", "Select a parent node from the tree.")
                return
            parent_node = self.find_node_by_id_all(node_id)
        if parent_node.node_type.upper() in ["CONFIDENCE LEVEL", "ROBUSTNESS SCORE", "BASIC EVENT"]:
            messagebox.showwarning("Invalid", "Base events cannot have children.")
            return
        data = selected.to_dict()
        data.pop("unique_id", None)
        data["children"] = []
        new_node = FaultTreeNode.from_dict(data, parent_node)
        if hasattr(selected, "unique_id"):
            new_node.failure_mode_ref = selected.unique_id
        parent_node.children.append(new_node)
        new_node.parents.append(parent_node)
        self.update_views()

    def add_basic_event_from_fmea(self):
        self.push_undo_state()
        events = list(self.fmea_entries)
        for doc in self.fmeas:
            events.extend(doc.get("entries", []))
        for doc in self.fmedas:
            events.extend(doc.get("entries", []))
        if not events:
            messagebox.showinfo("No Failure Modes", "No FMEA or FMEDA failure modes available.")
            return
        dialog = SelectBaseEventDialog(self.root, events)
        selected = dialog.selected
        if not selected:
            return
        if self.selected_node:
            parent_node = self.selected_node
            if not parent_node.is_primary_instance:
                messagebox.showwarning("Invalid Operation", "Cannot add to a clone node. Select the original.")
                return
        else:
            sel = self.analysis_tree.selection()
            if not sel:
                messagebox.showwarning("No selection", "Select a parent node to paste into.")
                return
            try:
                node_id = int(self.analysis_tree.item(sel[0], "tags")[0])
            except (IndexError, ValueError):
                messagebox.showwarning("No selection", "Select a parent node from the tree.")
                return
            parent_node = self.find_node_by_id_all(node_id)
        if parent_node.node_type.upper() in ["CONFIDENCE LEVEL", "ROBUSTNESS SCORE", "BASIC EVENT"]:
            messagebox.showwarning("Invalid", "Base events cannot have children.")
            return
        data = selected.to_dict()
        data.pop("unique_id", None)
        data["children"] = []
        new_node = FaultTreeNode.from_dict(data, parent_node)
        if hasattr(selected, "unique_id"):
            new_node.failure_mode_ref = selected.unique_id
        parent_node.children.append(new_node)
        new_node.parents.append(parent_node)
        self.update_views()


    def remove_node(self):
        return self.structure_tree_operations.remove_node()

    def remove_connection(self, node):
        return self.structure_tree_operations.remove_connection(node)

    def delete_node_and_subtree(self, node):
        return self.structure_tree_operations.delete_node_and_subtree(node)

    # ------------------------------------------------------------------
    # Helpers for malfunctions and failure modes
    # ------------------------------------------------------------------
    def create_top_event_for_malfunction(self, name: str) -> None:
        return self.top_event_workflows.create_top_event_for_malfunction(name)

    def delete_top_events_for_malfunction(self, name: str) -> None:
        return self.safety_analysis.delete_top_events_for_malfunction(name)

    def add_gate_from_failure_mode(self):
        return self.top_event_workflows.add_gate_from_failure_mode()

    def add_fault_event(self):
        return self.safety_analysis.add_fault_event()

    def calculate_overall(self):
        return self.probability_reliability.calculate_overall()

    def calculate_pmfh(self):
        return self.probability_reliability.calculate_pmfh()

    def show_requirements_matrix(self):
        """Display a matrix table of requirements vs. basic events."""
        self.update_requirement_statuses()
        basic_events = [n for n in self.get_all_nodes(self.root_node)
                        if n.node_type.upper() == "BASIC EVENT"]
        reqs = list(global_requirements.values())
        reqs.sort(key=lambda r: r.get("req_type", ""))

        win = tk.Toplevel(self.root)
        win.title("Requirements Matrix")

        columns = [
            "Req ID",
            "ASIL",
            "CAL",
            "Type",
            "Status",
            "Parent",
            "Text",
        ] + [be.user_name or f"BE {be.unique_id}" for be in basic_events]
        tree = ttk.Treeview(win, columns=columns, show="headings")
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=120 if col not in ["Text"] else 300, anchor="center")
        tree.pack(fill=tk.BOTH, expand=True)


        for req in reqs:
            row = [
                req.get("id", ""),
                req.get("asil", ""),
                req.get("cal", ""),
                req.get("req_type", ""),
                req.get("status", "draft"),
                req.get("parent_id", ""),
                req.get("text", ""),
            ]
            for be in basic_events:
                linked = any(r.get("id") == req.get("id") for r in getattr(be, "safety_requirements", []))
                row.append("X" if linked else "")
            tree.insert("", "end", values=row)

        # Show allocation and safety goal traceability below the table
        frame = tk.Frame(win)
        frame.pack(fill=tk.BOTH, expand=True)
        vbar = ttk.Scrollbar(frame, orient="vertical")
        text = tk.Text(frame, wrap="word", yscrollcommand=vbar.set, height=8)
        text.tag_configure("added", foreground="blue")
        text.tag_configure("removed", foreground="red")
        vbar.config(command=text.yview)
        text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vbar.pack(side=tk.RIGHT, fill=tk.Y)

        base_data = self.versions[-1]["data"] if self.versions else None

        def alloc_from_data(req_id):
            if not base_data:
                return ""
            names = []
            def traverse(d):
                if any(r.get("id") == req_id for r in d.get("safety_requirements", [])):
                    names.append(d.get("user_name") or f"Node {d.get('unique_id')}")
                for ch in d.get("children", []):
                    traverse(ch)
            for t in base_data.get("top_events", []):
                traverse(t)
            for fmea in base_data.get("fmeas", []):
                for e in fmea.get("entries", []):
                    if any(r.get("id") == req_id for r in e.get("safety_requirements", [])):
                        name = e.get("description") or e.get("user_name", f"BE {e.get('unique_id','')}")
                        names.append(f"{fmea['name']}:{name}")
            return ", ".join(names)

        def goals_from_data(req_id):
            if not base_data:
                return ""
            nodes = []
            def gather(n):
                nodes.append(n)
                for ch in n.get("children", []):
                    gather(ch)
            for t in base_data.get("top_events", []):
                gather(t)
            id_map = {n["unique_id"]: n for n in nodes}
            def collect_goal_names(nd, acc):
                if nd.get("node_type", "").upper() == "TOP EVENT":
                    acc.add(nd.get("safety_goal_description") or nd.get("user_name") or f"SG {nd.get('unique_id')}")
                for p in nd.get("parents", []):
                    pid = p.get("unique_id")
                    if pid and pid in id_map:
                        collect_goal_names(id_map[pid], acc)
            goals = set()
            for n in nodes:
                if any(r.get("id") == req_id for r in n.get("safety_requirements", [])):
                    collect_goal_names(n, goals)
            for fmea in base_data.get("fmeas", []):
                for e in fmea.get("entries", []):
                    if any(r.get("id") == req_id for r in e.get("safety_requirements", [])):
                        parents = e.get("parents", [])
                        if parents:
                            pid = parents[0].get("unique_id")
                            if pid and pid in id_map:
                                collect_goal_names(id_map[pid], goals)
            return ", ".join(sorted(goals))

        import difflib

        def insert_diff(widget, old, new):
            matcher = difflib.SequenceMatcher(None, old, new)
            for tag, i1, i2, j1, j2 in matcher.get_opcodes():
                if tag == "equal":
                    widget.insert(tk.END, new[j1:j2])
                elif tag == "delete":
                    widget.insert(tk.END, old[i1:i2], "removed")
                elif tag == "insert":
                    widget.insert(tk.END, new[j1:j2], "added")
                elif tag == "replace":
                    widget.insert(tk.END, old[i1:i2], "removed")
                    widget.insert(tk.END, new[j1:j2], "added")

        def insert_list_diff(widget, old, new):
            old_items = [s.strip() for s in old.split(',') if s.strip()]
            new_items = [s.strip() for s in new.split(',') if s.strip()]
            old_set = set(old_items)
            new_set = set(new_items)
            first = True
            for item in new_items:
                if not first:
                    widget.insert(tk.END, ", ")
                first = False
                if item not in old_set:
                    widget.insert(tk.END, item, "added")
                else:
                    widget.insert(tk.END, item)
            for item in old_items:
                if item not in new_set:
                    if not first:
                        widget.insert(tk.END, ", ")
                    first = False
                    widget.insert(tk.END, item, "removed")

        for req in reqs:
            rid = req.get("id")
            alloc = ", ".join(self.get_requirement_allocation_names(rid))
            goals = ", ".join(self.get_requirement_goal_names(rid))
            text.insert(tk.END, f"[{rid}] {req.get('text','')}\n")
            text.insert(tk.END, "  Allocated to: ")
            if base_data:
                insert_list_diff(text, alloc_from_data(rid), alloc)
            else:
                text.insert(tk.END, alloc)
            text.insert(tk.END, "\n  Safety Goals: ")
            if base_data:
                insert_diff(text, goals_from_data(rid), goals)
            else:
                text.insert(tk.END, goals)
            text.insert(tk.END, "\n\n")

        tk.Button(win, text="Open Requirements Editor", command=self.show_requirements_editor).pack(pady=5)

    def show_item_definition_editor(self):
        """Open editor for item description and assumptions."""
        if hasattr(self, "_item_def_tab") and self._item_def_tab.winfo_exists():
            self.doc_nb.select(self._item_def_tab)
            return
        self._item_def_tab = self.lifecycle_ui._new_tab("Item Definition")
        win = self._item_def_tab
        ttk.Label(win, text="Item Description:").pack(anchor="w")
        self._item_desc_text = tk.Text(win, height=8, wrap="word")
        self._item_desc_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        ttk.Label(win, text="Assumptions:").pack(anchor="w")
        self._item_assum_text = tk.Text(win, height=8, wrap="word")
        self._item_assum_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self._item_desc_text.insert("1.0", self.item_definition.get("description", ""))
        self._item_assum_text.insert("1.0", self.item_definition.get("assumptions", ""))

        def save():
            self.item_definition["description"] = self._item_desc_text.get("1.0", "end").strip()
            self.item_definition["assumptions"] = self._item_assum_text.get("1.0", "end").strip()

        ttk.Button(win, text="Save", command=save).pack(anchor="e", padx=5, pady=5)

    def show_safety_concept_editor(self):
        """Open editor for safety & security concept descriptions and assumptions."""
        if hasattr(self, "_safety_concept_tab") and self._safety_concept_tab.winfo_exists():
            self.doc_nb.select(self._safety_concept_tab)
            return
        self._safety_concept_tab = self.lifecycle_ui._new_tab("Safety & Security Concept")
        win = self._safety_concept_tab
        ttk.Label(
            win,
            text="Functional & Cybersecurity Concept Description and Assumptions:",
        ).pack(anchor="w")
        f_frame = ttk.Frame(win)
        f_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self._fsc_text = tk.Text(f_frame, height=8, wrap="word")
        f_scroll = ttk.Scrollbar(f_frame, orient=tk.VERTICAL, command=self._fsc_text.yview)
        self._fsc_text.configure(yscrollcommand=f_scroll.set)
        self._fsc_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        f_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        ttk.Label(
            win,
            text="Technical & Cybersecurity Concept Description & Assumptions:",
        ).pack(anchor="w")
        t_frame = ttk.Frame(win)
        t_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self._tsc_text = tk.Text(t_frame, height=8, wrap="word")
        t_scroll = ttk.Scrollbar(t_frame, orient=tk.VERTICAL, command=self._tsc_text.yview)
        self._tsc_text.configure(yscrollcommand=t_scroll.set)
        self._tsc_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        t_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        c_frame = ttk.Frame(win)
        c_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self._csc_text = tk.Text(c_frame, height=8, wrap="word")
        c_scroll = ttk.Scrollbar(c_frame, orient=tk.VERTICAL, command=self._csc_text.yview)
        self._csc_text.configure(yscrollcommand=c_scroll.set)
        self._csc_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        c_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        self._fsc_text.insert("1.0", self.safety_concept.get("functional", ""))
        self._tsc_text.insert("1.0", self.safety_concept.get("technical", ""))
        self._csc_text.insert("1.0", self.safety_concept.get("cybersecurity", ""))

        def save():
            self.safety_concept["functional"] = self._fsc_text.get("1.0", "end").strip()
            self.safety_concept["technical"] = self._tsc_text.get("1.0", "end").strip()
            self.safety_concept["cybersecurity"] = self._csc_text.get("1.0", "end").strip()

        ttk.Button(win, text="Save", command=save).pack(anchor="e", padx=5, pady=5)

    def show_requirements_editor(self):
        """Open an editor to manage global requirements."""
        import textwrap

        self.update_requirement_statuses()
        if hasattr(self, "_req_tab") and self._req_tab.winfo_exists():
            self.doc_nb.select(self._req_tab)
            return
        self._req_tab = self.lifecycle_ui._new_tab("Requirements")
        win = self._req_tab

        columns = ["ID", "ASIL", "CAL", "Type", "Status", "Parent", "Trace", "Links", "Text"]
        tree_frame = ttk.Frame(win)
        style = ttk.Style(tree_frame)
        style.configure("ReqEditor.Treeview", rowheight=20)
        tree = ttk.Treeview(
            tree_frame,
            columns=columns,
            show="headings",
            selectmode="browse",
            style="ReqEditor.Treeview",
        )
        tree.configure(height=10)
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        for col in columns:
            tree.heading(col, text=col)
            if col == "Text":
                width = 300
            elif col in ("Trace", "Links"):
                width = 200
            else:
                width = 120
            tree.column(col, width=width, anchor="center")
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        def _get_requirement_allocations(rid: str) -> list[str]:
            repo = SysMLRepository.get_instance()
            names: list[str] = []
            for diag in repo.diagrams.values():
                dname = diag.name or diag.diag_id
                for obj in getattr(diag, "objects", []):
                    for r in obj.get("requirements", []):
                        if r.get("id") == rid:
                            oname = obj.get("properties", {}).get("name", obj.get("obj_type"))
                            names.append(f"{dname}:{oname}")
            return sorted(set(names))

        def refresh_tree():
            tree.delete(*tree.get_children())
            max_lines = 1
            for req in global_requirements.values():
                rid = req.get("id", "")
                trace = ", ".join(_get_requirement_allocations(rid))
                links = ", ".join(
                    f"{r.get('type')} {r.get('id')}" for r in req.get("relations", [])
                )
                text = textwrap.fill(req.get("text", ""), width=40)
                max_lines = max(max_lines, text.count("\n") + 1)
                tree.insert(
                    "",
                    "end",
                    iid=rid,
                    values=[
                        rid,
                        req.get("asil", ""),
                        req.get("cal", ""),
                        req.get("req_type", ""),
                        req.get("status", "draft"),
                        req.get("parent_id", ""),
                        trace,
                        links,
                        text,
                    ],
                )
            style.configure("ReqEditor.Treeview", rowheight=20 * max_lines)

        def add_req():
            dlg = ReqDialog(win, "Add Requirement")
            if dlg.result:
                global_requirements[dlg.result["id"]] = dlg.result
                refresh_tree()

        def edit_req():
            sel = tree.selection()
            if not sel:
                return
            rid = sel[0]
            dlg = ReqDialog(win, "Edit Requirement", global_requirements.get(rid))
            if dlg.result:
                global_requirements[rid].update(dlg.result)
                refresh_tree()

        def del_req():
            sel = tree.selection()
            if not sel:
                return
            rid = sel[0]
            if messagebox.askyesno("Delete", "Delete requirement?"):
                del global_requirements[rid]
                for n in self.get_all_nodes(self.root_node):
                    reqs = getattr(n, "safety_requirements", [])
                    n.safety_requirements = [r for r in reqs if r.get("id") != rid]
                for fmea in self.fmeas:
                    for e in fmea.get("entries", []):
                        reqs = e.get("safety_requirements", [])
                        e["safety_requirements"] = [r for r in reqs if r.get("id") != rid]
                refresh_tree()

        def link_to_diagram():
            sel = tree.selection()
            if not sel:
                return
            rid = sel[0]
            req = global_requirements.get(rid)
            if not req:
                return
            repo = SysMLRepository.get_instance()
            toolbox = getattr(self, "safety_mgmt_toolbox", None)
            can_trace = toolbox.can_trace if toolbox else (lambda a, b: True)
            req_wp = (
                toolbox.requirement_work_product(req.get("req_type", ""))
                if toolbox
                else ""
            )

            # Determine currently allocated diagram objects
            existing: set[tuple[str, int]] = set()
            for diag in repo.diagrams.values():
                for obj in getattr(diag, "objects", []):
                    if any(r.get("id") == rid for r in obj.get("requirements", [])):
                        existing.add((diag.diag_id, obj.get("obj_id")))

            dlg = DiagramElementDialog(win, repo, req_wp, can_trace, selected=list(existing))
            targets = set(getattr(dlg, "selection", []))
            if not targets and not existing:
                return

            # Add newly selected links
            for diag_id, obj_id in targets - existing:
                diag = repo.diagrams.get(diag_id)
                if not diag:
                    continue
                obj = next(
                    (o for o in getattr(diag, "objects", []) if o.get("obj_id") == obj_id),
                    None,
                )
                if not obj:
                    continue
                link_requirement_to_object(obj, rid, diag_id)
                repo.touch_diagram(diag_id)
                elem_id = obj.get("element_id")
                if elem_id:
                    repo.touch_element(elem_id)

            # Remove deselected links
            for diag_id, obj_id in existing - targets:
                diag = repo.diagrams.get(diag_id)
                if not diag:
                    continue
                obj = next(
                    (o for o in getattr(diag, "objects", []) if o.get("obj_id") == obj_id),
                    None,
                )
                if not obj:
                    continue
                unlink_requirement_from_object(obj, rid, diag_id)
                repo.touch_diagram(diag_id)
                elem_id = obj.get("element_id")
                if elem_id:
                    repo.touch_element(elem_id)

            refresh_tree()

        def link_requirement():
            sel = tree.selection()
            if not sel:
                return
            rid = sel[0]
            req = global_requirements.get(rid)
            if not req:
                return
            toolbox = getattr(self, "safety_mgmt_toolbox", None)
            dlg = _RequirementRelationDialog(win, req, toolbox)
            if not dlg.result:
                return
            relation, targets = dlg.result
            selected = set(targets)
            existing = {
                r.get("id")
                for r in req.get("relations", [])
                if r.get("type") == relation
            }
            for tid in selected - existing:
                link_requirements(rid, relation, tid)
            for tid in existing - selected:
                unlink_requirements(rid, relation, tid)
            refresh_tree()

        def save_csv():
            path = filedialog.asksaveasfilename(
                defaultextension=".csv", filetypes=[("CSV", "*.csv")]
            )
            if not path:
                return
            try:
                with open(path, "w", newline="") as fh:
                    writer = csv.writer(fh)
                    writer.writerow(columns)
                    for req in global_requirements.values():
                        rid = req.get("id", "")
                        trace = ", ".join(_get_requirement_allocations(rid))
                        links = ", ".join(
                            f"{r.get('type')} {r.get('id')}" for r in req.get("relations", [])
                        )
                        writer.writerow(
                            [
                                rid,
                                req.get("asil", ""),
                                req.get("cal", ""),
                                req.get("req_type", ""),
                                req.get("status", "draft"),
                                req.get("parent_id", ""),
                                trace,
                                links,
                                req.get("text", ""),
                            ]
                        )
                messagebox.showinfo(
                    "Requirements", f"Saved {len(global_requirements)} requirements to {path}"
                )
            except Exception as exc:
                messagebox.showerror("Requirements", f"Failed to save CSV:\n{exc}")

        if hasattr(tree, "bind"):
            try:
                menu = tk.Menu(tree, tearoff=False)
            except Exception:
                menu = None
            if menu:
                menu.add_command(label="Add", command=add_req)
                menu.add_command(label="Edit", command=edit_req)
                menu.add_command(label="Delete", command=del_req)
                menu.add_command(label="Link to Diagram...", command=link_to_diagram)
                menu.add_command(label="Link Requirement...", command=link_requirement)
                menu.add_command(label="Save CSV", command=save_csv)

                def _popup(event: tk.Event) -> None:
                    row = tree.identify_row(event.y)
                    if row:
                        tree.selection_set(row)
                        tree.focus(row)
                    try:
                        menu.tk_popup(event.x_root, event.y_root)
                    finally:
                        menu.grab_release()

                def _on_double(event: tk.Event) -> None:
                    row = tree.identify_row(event.y)
                    if row:
                        tree.selection_set(row)
                        tree.focus(row)
                        edit_req()

                tree.bind("<Button-3>", _popup)
                tree.bind("<Button-2>", _popup)
                tree.bind("<Control-Button-1>", _popup)
                tree.bind("<Double-1>", _on_double)
                tree.context_menu = menu

        btn = tk.Frame(win)
        btn.pack(fill=tk.X)
        tk.Button(btn, text="Add", command=add_req).pack(side=tk.LEFT)
        tk.Button(btn, text="Edit", command=edit_req).pack(side=tk.LEFT)
        tk.Button(btn, text="Delete", command=del_req).pack(side=tk.LEFT)
        tk.Button(btn, text="Save CSV", command=save_csv).pack(side=tk.LEFT)
        tk.Button(btn, text="Link to Diagram...", command=link_to_diagram).pack(side=tk.LEFT)
        tk.Button(btn, text="Link Requirement...", command=link_requirement).pack(side=tk.LEFT)
        tree_frame.pack(fill=tk.BOTH, expand=True)

        refresh_tree()

    def show_hazard_list(self):
        return self.safety_analysis.show_hazard_list()

    class SelectFailureModeDialog(simpledialog.Dialog):
        def __init__(self, parent, app, modes):
            self.app = app
            self.modes = modes
            self.selected = None
            super().__init__(parent, title="Select Failure Mode")

        def body(self, master):
            self.listbox = tk.Listbox(master, height=10, width=50)
            for m in self.modes:
                label = self.app.format_failure_mode_label(m)
                self.listbox.insert(tk.END, label)
            self.listbox.grid(row=0, column=0, padx=5, pady=5)
            return self.listbox

        def apply(self):
            sel = self.listbox.curselection()
            if sel:
                self.selected = self.modes[sel[0]]

    class SelectFaultDialog(simpledialog.Dialog):
        def __init__(self, parent, faults, allow_new=False):
            self.faults = faults
            self.allow_new = allow_new
            self.selected = None
            super().__init__(parent, title="Select Fault")

        def body(self, master):
            self.listbox = tk.Listbox(master, height=10, width=40)
            for f in self.faults:
                self.listbox.insert(tk.END, f)
            if self.allow_new:
                self.listbox.insert(tk.END, "<Create New Fault>")
            self.listbox.grid(row=0, column=0, padx=5, pady=5)
            return self.listbox

        def apply(self):
            sel = self.listbox.curselection()
            if sel:
                idx = sel[0]
                if self.allow_new and idx == len(self.faults):
                    self.selected = "NEW"
                else:
                    self.selected = self.faults[idx]

    class SelectSafetyGoalsDialog(simpledialog.Dialog):
        def __init__(self, parent, goals, initial=None):
            self.goals = goals
            self.initial = initial or []
            self.result = []
            super().__init__(parent, title="Select Safety Goals")

        def body(self, master):
            ttk.Label(master, text="Select violated safety goals:").pack(padx=5, pady=5)
            self.vars = {}
            for sg in self.goals:
                name = sg.user_name or sg.safety_goal_description or f"SG {sg.unique_id}"
                var = tk.BooleanVar(value=name in self.initial)
                self.vars[sg] = var
                ttk.Checkbutton(master, text=name, variable=var).pack(anchor="w", padx=5, pady=2)
            return master

        def apply(self):
            self.result = [sg for sg, var in self.vars.items() if var.get()]

    def _show_fmea_table_impl(self, fmea=None, fmeda=False):
        """Internal implementation for rendering FMEA/FMeda tables."""
        # Use failure modes defined on gates or within FMEA/FMEDA documents.
        # Do not include FTA base events as selectable failure modes.
        basic_events = self.get_non_basic_failure_modes()
        entries = self.fmea_entries if fmea is None else fmea['entries']
        title = f"FMEA Table - {fmea['name']}" if fmea else "FMEA Table"
        win = self.lifecycle_ui._new_tab(title)

        # give the table a nicer look similar to professional FMEA tools
        style = ttk.Style(self.root)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass
        apply_translucid_button_style(style)
        style.configure(
            "FMEA.Treeview",
            font=("Segoe UI", 10),
            rowheight=60,
            background="#ffffff",
            fieldbackground="#ffffff",
            foreground="black",
        )
        style.configure(
            "FMEA.Treeview.Heading",
            font=("Segoe UI", 10, "bold"),
            background="#b5bdc9",
            foreground="black",
            relief="raised",
        )
        style.map(
            "FMEA.Treeview.Heading",
            background=[("active", "#4a6ea9"), ("!active", "#b5bdc9")],
            foreground=[("active", "white"), ("!active", "black")],
        )

        columns = [
            "Component",
            "Parent",
            "Failure Mode",
            "Failure Effect",
            "Cause",
            "S",
            "O",
            "D",
            "RPN",
            "Requirements",
            "Malfunction",
        ]
        if fmeda:
            columns.extend([
                "Safety Goal",
                "FaultType",
                "Fraction",
                "FIT",
                "DiagCov",
                "Mechanism",
            ])
        columns.extend(["Created", "Author", "Modified", "ModifiedBy"])
        btn_frame = ttk.Frame(win)
        btn_frame.pack(side=tk.TOP, pady=2)
        add_btn = ttk.Button(btn_frame, text="Add Failure Mode")
        add_btn.pack(side=tk.LEFT, padx=2)
        remove_btn = ttk.Button(btn_frame, text="Remove from FMEA")
        remove_btn.pack(side=tk.LEFT, padx=2)
        del_btn = ttk.Button(btn_frame, text="Delete Selected")
        del_btn.pack(side=tk.LEFT, padx=2)
        comment_btn = ttk.Button(btn_frame, text="Comment")
        comment_btn.pack(side=tk.LEFT, padx=2)
        toolbox = getattr(self, "safety_mgmt_toolbox", None)
        if fmea and toolbox and toolbox.document_read_only("FMEA", fmea["name"]):
            for b in (add_btn, remove_btn, del_btn, comment_btn):
                b.state(["disabled"])
        if fmeda:
            def calculate_fmeda():
                if bom_var.get():
                    load_bom()
                else:
                    refresh_tree()
                metrics = self.compute_fmeda_metrics(entries)
                asil = metrics["asil"]
                dc = metrics["dc"]
                spfm_m = metrics["spfm_metric"]
                lpfm_m = metrics["lpfm_metric"]
                thresh = ASIL_TARGETS.get(asil, ASIL_TARGETS["QM"])
                ok_dc = dc >= thresh["dc"]
                ok_spf = spfm_m >= thresh["spfm"]
                ok_lpf = lpfm_m >= thresh["lpfm"]
                msg = (
                    f"Total FIT: {self.reliability_total_fit:.2f}\n"
                    f"DC: {dc:.2f} {'PASS' if ok_dc else 'FAIL'}\n"
                    f"SPFM: {spfm_m:.2f} {'PASS' if ok_spf else 'FAIL'}\n"
                    f"LPFM: {lpfm_m:.2f} {'PASS' if ok_lpf else 'FAIL'}\n"
                    f"ASIL {asil}"
                )
                messagebox.showinfo("FMEDA", msg)

            calc_btn = ttk.Button(btn_frame, text="Calculate FMEDA", command=calculate_fmeda)
            calc_btn.pack(side=tk.LEFT, padx=2)
            ttk.Label(btn_frame, text="BOM:").pack(side=tk.LEFT, padx=2)
            bom_var = tk.StringVar(value=fmea.get('bom', ''))
            bom_combo = ttk.Combobox(
                btn_frame,
                textvariable=bom_var,
                values=[ra.name for ra in self.reliability_analyses],
                state="readonly",
                width=20,
            )
            bom_combo.pack(side=tk.LEFT, padx=2)

            def add_component():
                dlg = tk.Toplevel(win)
                dlg.title("New Component")
                ttk.Label(dlg, text="Name").grid(row=0, column=0, padx=5, pady=5, sticky="e")
                name_var = tk.StringVar()
                ttk.Entry(dlg, textvariable=name_var).grid(row=0, column=1, padx=5, pady=5)
                ttk.Label(dlg, text="Type").grid(row=1, column=0, padx=5, pady=5, sticky="e")
                type_var = tk.StringVar(value="capacitor")
                type_cb = ttk.Combobox(
                    dlg,
                    textvariable=type_var,
                    values=list(COMPONENT_ATTR_TEMPLATES.keys()),
                    state="readonly",
                )
                type_cb.grid(row=1, column=1, padx=5, pady=5)
                ttk.Label(dlg, text="Quantity").grid(row=2, column=0, padx=5, pady=5, sticky="e")
                qty_var = tk.IntVar(value=1)
                ttk.Entry(dlg, textvariable=qty_var).grid(row=2, column=1, padx=5, pady=5)
                ttk.Label(dlg, text="Qualification").grid(row=3, column=0, padx=5, pady=5, sticky="e")
                qual_var = tk.StringVar(value="None")
                ttk.Combobox(dlg, textvariable=qual_var, values=QUALIFICATIONS, state="readonly").grid(row=3, column=1, padx=5, pady=5)
                passive_var = tk.BooleanVar(value=False)
                ttk.Checkbutton(dlg, text="Passive", variable=passive_var).grid(row=4, column=0, columnspan=2, pady=5)

                attr_frame = ttk.Frame(dlg)
                attr_frame.grid(row=5, column=0, columnspan=2)
                attr_vars = {}

                def refresh_attr_fields(*_):
                    for child in attr_frame.winfo_children():
                        child.destroy()
                    attr_vars.clear()
                    template = COMPONENT_ATTR_TEMPLATES.get(type_var.get(), {})
                    for i, (k, v) in enumerate(template.items()):
                        ttk.Label(attr_frame, text=k).grid(row=i, column=0, padx=5, pady=5, sticky="e")
                        if isinstance(v, list):
                            var = tk.StringVar(value=v[0])
                            ttk.Combobox(attr_frame, textvariable=var, values=v, state="readonly").grid(row=i, column=1, padx=5, pady=5)
                        else:
                            var = tk.StringVar(value=str(v))
                            ttk.Entry(attr_frame, textvariable=var).grid(row=i, column=1, padx=5, pady=5)
                        attr_vars[k] = var

                type_cb.bind("<<ComboboxSelected>>", refresh_attr_fields)
                refresh_attr_fields()

                def ok():
                    comp = ReliabilityComponent(
                        name_var.get(),
                        type_var.get(),
                        qty_var.get(),
                        {},
                        qual_var.get(),
                        is_passive=passive_var.get(),
                    )
                    for k, var in attr_vars.items():
                        comp.attributes[k] = var.get()
                    self.reliability_components.append(comp)
                    dlg.destroy()
                    refresh_tree()

                ttk.Button(dlg, text="Add", command=ok).grid(row=6, column=0, columnspan=2, pady=5)
                dlg.grab_set()
                dlg.wait_window()

            ttk.Button(btn_frame, text="Add Component", command=add_component).pack(side=tk.LEFT, padx=2)

            selected_libs = self.selected_mechanism_libraries

            def choose_libs():
                dlg = tk.Toplevel(win)
                dlg.title("Select Libraries")
                vars = {}
                for i, lib in enumerate(self.mechanism_libraries):
                    var = tk.BooleanVar(value=lib in selected_libs)
                    tk.Checkbutton(dlg, text=lib.name, variable=var).pack(anchor="w")
                    vars[i] = (var, lib)

                def ok():
                    selected_libs.clear()
                    for _, (v, lib) in vars.items():
                        if v.get():
                            selected_libs.append(lib)
                    dlg.destroy()

                ttk.Button(dlg, text="OK", command=ok).pack(pady=5)
                dlg.grab_set()
                dlg.wait_window()

            ttk.Button(btn_frame, text="Libraries", command=choose_libs).pack(side=tk.LEFT, padx=2)

            def load_bom(*_):
                name = bom_var.get()
                ra = next((r for r in self.reliability_analyses if r.name == name), None)
                if ra:
                    self.reliability_components = copy.deepcopy(ra.components)
                    self.reliability_total_fit = ra.total_fit
                    self.spfm = ra.spfm
                    self.lpfm = ra.lpfm
                    if fmea is not None:
                        fmea['bom'] = name
                    refresh_tree()

            bom_combo.bind("<<ComboboxSelected>>", load_bom)

        tree_frame = ttk.Frame(win)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        tree = ttk.Treeview(
            tree_frame,
            columns=columns,
            show="tree headings",
            style="FMEA.Treeview",
        )
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        for col in columns:
            tree.heading(col, text=col)
            width = 120
            if col in ["Requirements", "Failure Effect", "Cause", "Safety Goal", "Malfunction"]:
                width = 200
            elif col == "Parent":
                width = 150
            elif col in ["FaultType", "Fraction", "FIT", "DiagCov", "Mechanism"]:
                width = 80
            elif col in ["Created", "Modified"]:
                width = 130
            elif col in ["Author", "ModifiedBy"]:
                width = 100
            tree.column(col, width=width, anchor="center")
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tree_frame.grid_columnconfigure(0, weight=1)
        tree_frame.grid_rowconfigure(0, weight=1)

        metrics_lbl = tk.Label(win, text="", anchor="w")
        metrics_lbl.pack(anchor="w", padx=5, pady=2)

        # alternating row colours and high RPN highlight
        tree.tag_configure("component", background="#e2e2e2", font=("Segoe UI", 10, "bold"))
        tree.tag_configure("evenrow", background="#ffffff")
        tree.tag_configure("oddrow", background="#f5f5f5")
        tree.tag_configure("highrpn", background="#ffe6e6")

        node_map = {}
        comp_items = {}
        # expose the current FMEA tree and node mapping for external tools
        self._fmea_tree = tree
        self._fmea_node_map = node_map

        def refresh_tree():
            tree.delete(*tree.get_children())
            node_map.clear()
            comp_items.clear()
            # remove any duplicate nodes based on unique_id
            unique = {}
            for be in entries:
                unique[be.unique_id] = be
            entries[:] = list(unique.values())
            events = entries

            comp_fit = component_fit_map(self.reliability_components)
            frac_totals = {}
            for be in events:
                src = self.get_failure_mode_node(be)
                comp_name = self.get_component_name_for_node(src)
                fit = comp_fit.get(comp_name)
                frac = src.fmeda_fault_fraction
                if frac > 1.0:
                    frac /= 100.0
                if fit is not None:
                    value = fit * frac
                else:
                    value = getattr(src, "fmeda_fit", 0.0)

                be.fmeda_fit = value
                src.fmeda_fit = value

                if src.fmeda_fault_type == "permanent":
                    spfm = value * (1 - src.fmeda_diag_cov)
                    lpfm = 0.0
                else:
                    lpfm = value * (1 - src.fmeda_diag_cov)
                    spfm = 0.0

                be.fmeda_spfm = spfm
                be.fmeda_lpfm = lpfm
                src.fmeda_spfm = spfm
                src.fmeda_lpfm = lpfm
                frac_totals[comp_name] = frac_totals.get(comp_name, 0.0) + frac

            warnings = [f"{name} fractions={val:.2f}" for name, val in frac_totals.items() if abs(val - 1.0) > 0.01]
            if warnings:
                messagebox.showwarning("Distribution", "Fault fraction sum != 1:\n" + "\n".join(warnings))

            for idx, be in enumerate(events):
                src = self.get_failure_mode_node(be)
                comp = self.get_component_name_for_node(src) or "N/A"
                parent = src.parents[0] if src.parents else None
                parent_name = parent.user_name if parent and getattr(parent, "node_type", "").upper() not in GATE_NODE_TYPES else ""
                if comp not in comp_items:
                    comp_items[comp] = tree.insert(
                        "",
                        "end",
                        text=comp,
                        values=[comp] + [""] * (len(columns) - 1),
                        tags=("component",),
                    )
                comp_iid = comp_items[comp]
                req_ids = "; ".join(
                    [f"{req['req_type']}:{req['text']}" for req in getattr(src, "safety_requirements", [])]
                )
                rpn = src.fmea_severity * src.fmea_occurrence * src.fmea_detection
                failure_mode = src.description or (src.user_name or f"BE {src.unique_id}")
                vals = [
                    "",
                    parent_name,
                    failure_mode,
                    src.fmea_effect,
                    src.fmea_cause,
                    src.fmea_severity,
                    src.fmea_occurrence,
                    src.fmea_detection,
                    rpn,
                    req_ids,
                    src.fmeda_malfunction,
                ]
                if fmeda:
                    sg_value = src.fmeda_safety_goal
                    goals = self.get_top_event_safety_goals(src)
                    if goals:
                        sg_value = ", ".join(goals)
                    vals.extend([
                        sg_value,
                        src.fmeda_fault_type,
                        f"{src.fmeda_fault_fraction:.2f}",
                        f"{src.fmeda_fit:.2f}",
                        f"{src.fmeda_diag_cov:.2f}",
                        getattr(src, "fmeda_mechanism", ""),
                    ])
                vals.extend([
                    getattr(src, "created", ""),
                    getattr(src, "author", ""),
                    getattr(src, "modified", ""),
                    getattr(src, "modified_by", ""),
                ])
                tags = ["evenrow" if idx % 2 == 0 else "oddrow"]
                if rpn >= 100:
                    tags.append("highrpn")
                iid = tree.insert(comp_iid, "end", text="", values=vals, tags=tags)
                node_map[iid] = be
            for iid in comp_items.values():
                tree.item(iid, open=True)

            if fmeda:
                metrics = self.compute_fmeda_metrics(events)
                asil = metrics["asil"]
                dc = metrics["dc"]
                spfm_metric = metrics["spfm_metric"]
                lpfm_metric = metrics["lpfm_metric"]
                thresh = ASIL_TARGETS.get(asil, ASIL_TARGETS["QM"])
                ok_dc = dc >= thresh["dc"]
                ok_spf = spfm_metric >= thresh["spfm"]
                ok_lpf = lpfm_metric >= thresh["lpfm"]
                text = (
                    f"Total FIT: {self.reliability_total_fit:.2f}  DC: {dc:.2f}{CHECK_MARK if ok_dc else CROSS_MARK}"
                    f"  SPFM: {spfm_metric:.2f}{CHECK_MARK if ok_spf else CROSS_MARK}"
                    f"  LPFM: {lpfm_metric:.2f}{CHECK_MARK if ok_lpf else CROSS_MARK}"
                    f"  (ASIL {asil})"
                )
                if metrics.get("goal_metrics"):
                    parts = []
                    for sg, gm in metrics["goal_metrics"].items():
                        ok = gm["ok_dc"] and gm["ok_spfm"] and gm["ok_lpfm"]
                        symbol = CHECK_MARK if ok else CROSS_MARK
                        parts.append(f"{sg}:{symbol}")
                    text += " [" + "; ".join(parts) + "]"
                overall_ok = ok_dc and ok_spf and ok_lpf
                if metrics.get("goal_metrics"):
                    overall_ok = overall_ok and all(
                        gm["ok_dc"] and gm["ok_spfm"] and gm["ok_lpfm"]
                        for gm in metrics["goal_metrics"].values()
                    )
                color = "#c8ffc8" if overall_ok else "#ffc8c8"
                metrics_lbl.config(text=text, bg=color)

        if fmeda and bom_var.get():
            load_bom()
        else:
            refresh_tree()

        def on_double(event):
            sel = tree.focus()
            node = node_map.get(sel)
            if node:
                mechs = []
                for lib in selected_libs:
                    mechs.extend(lib.mechanisms)
                comp_name = self.get_component_name_for_node(node)
                is_passive = any(c.name == comp_name and c.is_passive for c in self.reliability_components)
                FMEARowDialog(win, node, self, entries, mechanisms=mechs, hide_diagnostics=is_passive, is_fmeda=fmeda)
                refresh_tree()

        tree.bind("<Double-1>", on_double)

        def add_failure_mode():
            dialog = SelectBaseEventDialog(win, basic_events, allow_new=True)
            node = dialog.selected
            if node == "NEW":
                node = FaultTreeNode("", "Basic Event")
                entries.append(node)
                mechs = []
                for lib in selected_libs:
                    mechs.extend(lib.mechanisms)
                comp_name = getattr(node, "fmea_component", "")
                is_passive = any(c.name == comp_name and c.is_passive for c in self.reliability_components)
                FMEARowDialog(win, node, self, entries, mechanisms=mechs, hide_diagnostics=is_passive, is_fmeda=fmeda)
            elif node:
                # gather all failure modes under the same component/parent
                if node.parents:
                    parent_id = node.parents[0].unique_id
                    related = [
                        be
                        for be in basic_events
                        if be.parents and be.parents[0].unique_id == parent_id
                    ]
                else:
                    comp = getattr(node, "fmea_component", "")
                    related = [
                        be
                        for be in basic_events
                        if not be.parents and getattr(be, "fmea_component", "") == comp
                    ]
                if node not in related:
                    related.append(node)
                existing_ids = {be.unique_id for be in entries}
                for be in related:
                    if be.unique_id not in existing_ids:
                        entries.append(be)
                        existing_ids.add(be.unique_id)
                    mechs = []
                    for lib in selected_libs:
                        mechs.extend(lib.mechanisms)
                    comp_name = self.get_component_name_for_node(be)
                is_passive = any(c.name == comp_name and c.is_passive for c in self.reliability_components)
                FMEARowDialog(win, be, self, entries, mechanisms=mechs, hide_diagnostics=is_passive, is_fmeda=fmeda)
            refresh_tree()
            if fmea is not None:
                self.lifecycle_ui.touch_doc(fmea)

        add_btn.config(command=add_failure_mode)

        def remove_from_fmea():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Remove Entry", "Select a row to remove.")
                return
            for iid in sel:
                node = node_map.get(iid)
                if node in entries:
                    entries.remove(node)
            refresh_tree()
            if fmea is not None:
                self.lifecycle_ui.touch_doc(fmea)

        remove_btn.config(command=remove_from_fmea)

        def delete_failure_mode():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Delete Failure Mode", "Select a row to delete.")
                return
            if not messagebox.askyesno("Delete Failure Mode", "Remove selected failure modes from the FMEA?"):
                return
            for iid in sel:
                node = node_map.get(iid)
                if node in entries:
                    entries.remove(node)
            refresh_tree()
            if fmea is not None:
                self.lifecycle_ui.touch_doc(fmea)

        del_btn.config(command=delete_failure_mode)

        def comment_fmea_entry():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Comment", "Select a row to comment.")
                return
            node = node_map.get(sel[0])
            if not node:
                return
            self.selected_node = node
            self.comment_target = ("fmea", node.unique_id)
            self.open_review_toolbox()

        comment_btn.config(command=comment_fmea_entry)

        def on_close():
            if fmea is not None:
                self.lifecycle_ui.touch_doc(fmea)
                if fmeda:
                    self.export_fmeda_to_csv(fmea, fmea['file'])
                else:
                    self.export_fmea_to_csv(fmea, fmea['file'])
                if fmeda:
                    fmea['bom'] = bom_var.get()
            win.destroy()

        if hasattr(win, "protocol"):
            win.protocol("WM_DELETE_WINDOW", on_close)
        else:
            win.bind("<Destroy>", lambda e: on_close() if e.widget is win else None)

    def export_fmea_to_csv(self, fmea, path):
        return self.safety_analysis.export_fmea_to_csv(fmea, path)

    def export_fmeda_to_csv(self, fmeda, path):
        return self.safety_analysis.export_fmeda_to_csv(fmeda, path)


    def show_traceability_matrix(self):
        """Display a traceability matrix linking FTA basic events to FMEA components."""
        basic_events = [n for n in self.get_all_nodes(self.root_node)
                        if n.node_type.upper() == "BASIC EVENT"]
        win = tk.Toplevel(self.root)
        win.title("FTA-FMEA Traceability")
        columns = ["Basic Event", "Component"]
        tree = ttk.Treeview(win, columns=columns, show="headings")
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=200, anchor="center")
        tree.pack(fill=tk.BOTH, expand=True)

        for be in basic_events:
            comp = self.get_component_name_for_node(be) or "N/A"
            tree.insert(
                "",
                "end",
                values=[be.user_name or f"BE {be.unique_id}", comp],
            )

    def collect_requirements_recursive(self, node):
        return self.safety_analysis.collect_requirements_recursive(node)

    def show_safety_goals_matrix(self):
        """Display product goals and derived requirements in a tree view."""
        if hasattr(self, "_sg_matrix_tab") and self._sg_matrix_tab.winfo_exists():
            self.doc_nb.select(self._sg_matrix_tab)
            return
        self._sg_matrix_tab = self.lifecycle_ui._new_tab("Product Goals Matrix")
        win = self._sg_matrix_tab
        columns = [
            "ID",
            "ASIL",
            "Target PMHF",
            "CAL",
            "SafeState",
            "FTTI",
            "Acc Rate",
            "On Hours",
            "Val Target",
            "Profile",
            "Val Desc",
            "Acceptance",
            "Description",
            "Text",
        ]
        tree = ttk.Treeview(win, columns=columns, show="tree headings")
        tree.heading("ID", text="Requirement ID")
        tree.heading("ASIL", text="ASIL")
        tree.heading("Target PMHF", text="Target PMHF (1/h)")
        tree.heading("CAL", text="CAL")
        tree.heading("SafeState", text="Safe State")
        tree.heading("FTTI", text="FTTI")
        tree.heading("Acc Rate", text="Acc Rate (1/h)")
        tree.heading("On Hours", text="On Hours")
        tree.heading("Val Target", text="Val Target")
        tree.heading("Profile", text="Profile")
        tree.heading("Val Desc", text="Val Desc")
        tree.heading("Acceptance", text="Acceptance")
        tree.heading("Description", text="Description")
        tree.heading("Text", text="Text")
        tree.column("ID", width=120)
        tree.column("ASIL", width=60)
        tree.column("Target PMHF", width=120)
        tree.column("CAL", width=60)
        tree.column("SafeState", width=100)
        tree.column("FTTI", width=80)
        tree.column("Acc Rate", width=100)
        tree.column("On Hours", width=100)
        tree.column("Val Target", width=120)
        tree.column("Profile", width=120)
        tree.column("Val Desc", width=200)
        tree.column("Acceptance", width=200)
        tree.column("Description", width=200)
        tree.column("Text", width=300)

        vsb = ttk.Scrollbar(win, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(win, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        win.columnconfigure(0, weight=1)
        win.rowconfigure(0, weight=1)

        for te in self.top_events:
            sg_text = te.safety_goal_description or (te.user_name or f"SG {te.unique_id}")
            sg_id = te.user_name or f"SG {te.unique_id}"
            cal = self.get_cyber_goal_cal(sg_id)
            asil = te.safety_goal_asil or "QM"
            target = PMHF_TARGETS.get(asil, 1.0)
            parent_iid = tree.insert(
                "",
                "end",
                text=sg_text,
                values=[
                    sg_id,
                    te.safety_goal_asil,
                    f"{target:.1e}",
                    cal,
                    te.safe_state,
                    getattr(te, "ftti", ""),
                    str(getattr(te, "acceptance_rate", "")),
                    getattr(te, "operational_hours_on", ""),
                    getattr(te, "validation_target", ""),
                    getattr(te, "mission_profile", ""),
                    getattr(te, "validation_desc", ""),
                    getattr(te, "acceptance_criteria", ""),
                    sg_text,
                    "",
                ],
            )
            reqs = self.collect_requirements_recursive(te)
            seen_ids = set()
            for req in reqs:
                req_id = req.get("id")
                if req_id in seen_ids:
                    continue
                seen_ids.add(req_id)
                tree.insert(
                    parent_iid,
                    "end",
                    text="",
                    values=[
                        req_id,
                        req.get("asil", ""),
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        req.get("text", ""),
                    ],
                )

    def show_product_goals_editor(self):
        """Allow editing of top-level product goals."""
        if hasattr(self, "_sg_tab") and self._sg_tab.winfo_exists():
            self.doc_nb.select(self._sg_tab)
            return
        self._sg_tab = self.lifecycle_ui._new_tab("Product Goals")
        win = self._sg_tab

        columns = [
            "ID",
            "ASIL",
            "Target PMHF",
            "Safe State",
            "FTTI",
            "Acc Rate",
            "On Hours",
            "Val Target",
            "Profile",
            "Val Desc",
            "Acceptance",
            "Description",
        ]
        tree = ttk.Treeview(win, columns=columns, show="headings", selectmode="browse")
        for c in columns:
            heading = "Target PMHF (1/h)" if c == "Target PMHF" else c
            tree.heading(c, text=heading)
            tree.column(c, width=120 if c not in ("Description", "Val Desc", "Acceptance") else 300, anchor="center")
        tree.pack(fill=tk.BOTH, expand=True)

        def refresh_tree():
            tree.delete(*tree.get_children())
            for sg in self.top_events:
                name = sg.safety_goal_description or (sg.user_name or f"SG {sg.unique_id}")
                sg.safety_goal_asil = self.get_hara_goal_asil(name)
                pmhf_target = PMHF_TARGETS.get(sg.safety_goal_asil, 1.0)
                tree.insert(
                    "",
                    "end",
                    iid=sg.unique_id,
                    values=[
                        sg.user_name or f"SG {sg.unique_id}",
                        sg.safety_goal_asil,
                        f"{pmhf_target:.2e}",
                        sg.safe_state,
                        getattr(sg, "ftti", ""),
                        str(getattr(sg, "acceptance_rate", "")),
                        getattr(sg, "operational_hours_on", ""),
                        getattr(sg, "validation_target", ""),
                        getattr(sg, "mission_profile", ""),
                        getattr(sg, "validation_desc", ""),
                        getattr(sg, "acceptance_criteria", ""),
                        sg.safety_goal_description,
                    ],
                )

        class SGDialog(simpledialog.Dialog):
            def __init__(self, parent, app, title, initial=None):
                self.app = app
                self.initial = initial
                super().__init__(parent, title=title)

            def body(self, master):
                nb = ttk.Notebook(master)
                nb.pack(fill=tk.BOTH, expand=True)

                fs_tab = ttk.Frame(nb)
                sotif_tab = ttk.Frame(nb)
                cyber_tab = ttk.Frame(nb)
                nb.add(fs_tab, text="Functional Safety")
                nb.add(sotif_tab, text="SOTIF")
                nb.add(cyber_tab, text="Cybersecurity")

                name = getattr(self.initial, "safety_goal_description", "") or getattr(self.initial, "user_name", "")

                # --- Functional Safety fields ---
                ttk.Label(fs_tab, text="ID:").grid(row=0, column=0, sticky="e")
                self.id_var = tk.StringVar(value=getattr(self.initial, "user_name", ""))
                self.id_entry = tk.Entry(fs_tab, textvariable=self.id_var)
                self.id_entry.grid(row=0, column=1, padx=5, pady=5)

                ttk.Label(fs_tab, text="ASIL:").grid(row=1, column=0, sticky="e")
                self.asil_var = tk.StringVar(value=self.app.get_hara_goal_asil(name))
                ttk.Label(fs_tab, textvariable=self.asil_var).grid(row=1, column=1, padx=5, pady=5, sticky="w")

                ttk.Label(fs_tab, text="Target PMHF (1/h):").grid(row=2, column=0, sticky="e")
                pmhf = PMHF_TARGETS.get(self.asil_var.get(), 1.0)
                self.pmhf_var = tk.StringVar(value=f"{pmhf:.2e}")
                tk.Entry(fs_tab, textvariable=self.pmhf_var, state="readonly").grid(row=2, column=1, padx=5, pady=5, sticky="w")

                ttk.Label(fs_tab, text="Safe State:").grid(row=3, column=0, sticky="e")
                self.state_var = tk.StringVar(value=getattr(self.initial, "safe_state", ""))
                tk.Entry(fs_tab, textvariable=self.state_var).grid(row=3, column=1, padx=5, pady=5)

                ttk.Label(fs_tab, text="FTTI:").grid(row=4, column=0, sticky="e")
                self.ftti_var = tk.StringVar(value=getattr(self.initial, "ftti", ""))
                tk.Entry(
                    fs_tab,
                    textvariable=self.ftti_var,
                    validate="key",
                    validatecommand=(
                        master.register(self.app.validation_consistency.validate_float),
                        "%P",
                    ),
                ).grid(row=4, column=1, padx=5, pady=5)

                ttk.Label(fs_tab, text="Description:").grid(row=5, column=0, sticky="ne")
                self.desc_text = tk.Text(fs_tab, width=30, height=3, wrap="word")
                self.desc_text.insert("1.0", getattr(self.initial, "safety_goal_description", ""))
                self.desc_text.grid(row=5, column=1, padx=5, pady=5)

                # --- SOTIF fields ---
                self.app.sotif_manager.build_goal_dialog(self, sotif_tab, self.initial)

                # --- Cybersecurity fields ---
                self.cal_var = self.app.cyber_manager.add_goal_dialog_fields(cyber_tab, name)
                return self.id_entry

            def apply(self):
                desc = self.desc_text.get("1.0", "end-1c").strip()
                sg_name = desc or self.id_var.get().strip()
                asil = self.app.get_hara_goal_asil(sg_name)
                self.result = {
                    "id": self.id_var.get().strip(),
                    "asil": asil,
                    "state": self.state_var.get().strip(),
                    "ftti": self.ftti_var.get().strip(),
                    "desc": desc,
                }
                self.result.update(self.app.sotif_manager.collect_goal_data(self))

        def add_sg():
            dlg = SGDialog(win, self, "Add Product Goal")
            if dlg.result:
                node = FaultTreeNode(dlg.result["id"], "TOP EVENT")
                node.safety_goal_asil = dlg.result["asil"]
                node.safe_state = dlg.result["state"]
                node.ftti = dlg.result["ftti"]
                node.acceptance_rate = float(dlg.result.get("accept_rate", 0.0) or 0.0)
                node.operational_hours_on = float(dlg.result.get("op_hours", 0.0) or 0.0)
                node.update_validation_target()
                node.mission_profile = dlg.result.get("profile", "")
                node.validation_desc = dlg.result["val_desc"]
                node.acceptance_criteria = dlg.result["accept"]
                node.safety_goal_description = dlg.result["desc"]
                self.top_events.append(node)
                refresh_tree()
                self.update_views()

        def edit_sg():
            sel = tree.selection()
            if not sel:
                return
            uid = int(sel[0])
            sg = self.find_node_by_id_all(uid)
            dlg = SGDialog(win, self, "Edit Product Goal", sg)
            if dlg.result:
                sg.user_name = dlg.result["id"]
                sg.safety_goal_asil = dlg.result["asil"]
                sg.safe_state = dlg.result["state"]
                sg.ftti = dlg.result["ftti"]
                sg.acceptance_rate = float(dlg.result.get("accept_rate", 0.0) or 0.0)
                sg.operational_hours_on = float(dlg.result.get("op_hours", 0.0) or 0.0)
                sg.update_validation_target()
                sg.mission_profile = dlg.result.get("profile", "")
                sg.validation_desc = dlg.result["val_desc"]
                sg.acceptance_criteria = dlg.result["accept"]
                sg.safety_goal_description = dlg.result["desc"]
                refresh_tree()
                self.update_views()

        def del_sg():
            sel = tree.selection()
            if not sel:
                return
            uid = int(sel[0])
            sg = self.find_node_by_id_all(uid)
            if sg and messagebox.askyesno("Delete", "Delete product goal?"):
                self.top_events = [t for t in self.top_events if t.unique_id != uid]
                refresh_tree()
                self.update_views()

        tree.bind("<Double-1>", lambda e: edit_sg())

        btn = ttk.Frame(win)
        btn.pack(fill=tk.X)
        ttk.Button(btn, text="Add", command=add_sg).pack(side=tk.LEFT)
        ttk.Button(btn, text="Edit", command=edit_sg).pack(side=tk.LEFT)
        ttk.Button(btn, text="Delete", command=del_sg).pack(side=tk.LEFT)

        refresh_tree()



    def _parse_spi_target(self, target: str) -> tuple[str, str]:
        """Split ``target`` into product goal name and SPI type."""
        if target.endswith(")") and "(" in target:
            name, typ = target.rsplit(" (", 1)
            return name, typ[:-1]
        return target, ""

    def get_spi_targets(self) -> list[str]:
        """Return sorted list of SPI options formatted as 'Product Goal (Type)'."""
        targets: set[str] = set()
        for sg in getattr(self, "top_events", []):
            pg_name = self._product_goal_name(sg)
            targets.update(self.sotif_manager.get_spi_targets_for_goal(sg, pg_name))
            asil = getattr(sg, "safety_goal_asil", "")
            if asil in PMHF_TARGETS:
                targets.add(f"{pg_name} (FUSA)")
        return sorted(targets)

    def show_safety_performance_indicators(self):
        """Display Safety Performance Indicators."""
        if hasattr(self, "_spi_tab") and self._spi_tab.winfo_exists():
            self.doc_nb.select(self._spi_tab)
            self.refresh_safety_performance_indicators()
            return
        self._spi_tab = self.lifecycle_ui._new_tab("Safety Performance Indicators")
        win = self._spi_tab

        columns = [
            "Product Goal",
            "Validation Target",
            "Achieved Probability",
            "SPI",
            "Target Description",
            "Acceptance Criteria",
        ]
        tree = ttk.Treeview(win, columns=columns, show="headings", selectmode="browse")
        for c in columns:
            tree.heading(c, text=c)
            width = 120
            if c in ("Target Description", "Acceptance Criteria"):
                width = 300
            tree.column(c, width=width, anchor="center")
        tree.pack(fill=tk.BOTH, expand=True)
        self._spi_tree = tree
        self._spi_lookup = {}

        def edit_selected():
            sel = tree.selection()
            if not sel:
                return
            iid = sel[0]
            sg_info = self._spi_lookup.get(iid)
            if not sg_info:
                return
            sg, spi_type = sg_info
            if spi_type != "SOTIF":
                return
            new_val = simpledialog.askfloat(
                "Achieved Probability",
                "Enter achieved probability:",
                initialvalue=getattr(sg, "spi_probability", 0.0),
            )
            if new_val is not None:
                self.push_undo_state()
                sg.spi_probability = float(new_val)
                self.refresh_safety_case_table()
                self.refresh_safety_performance_indicators()
                self.update_views()

        btn = ttk.Button(win, text="Edit", command=edit_selected)
        btn.pack(pady=4)
        self._edit_spi_item = edit_selected

        self.refresh_safety_performance_indicators()

    def refresh_safety_performance_indicators(self):
        """Populate the SPI explorer table."""
        tree = getattr(self, "_spi_tree", None)
        if not tree or not getattr(tree, "winfo_exists", lambda: True)():
            return
        for iid in list(tree.get_children("")):
            tree.delete(iid)
        self._spi_lookup = {}

        manager = getattr(self, "sotif_manager", None)
        if manager is None:
            manager = SOTIFManager(self)
            self.sotif_manager = manager
        for sg, values in manager.iter_spi_rows():
            iid = tree.insert("", "end", values=values)
            self._spi_lookup[iid] = (sg, "SOTIF")

        for sg in getattr(self, "top_events", []):
            asil = getattr(sg, "safety_goal_asil", "")
            if asil in PMHF_TARGETS:
                target = PMHF_TARGETS[asil]
                v_str = f"{target:.2e}"
                fusa_prob = getattr(sg, "probability", "")
                p_str = f"{fusa_prob:.2e}" if fusa_prob not in ("", None) else ""
                spi_val = ""
                try:
                    if fusa_prob not in ("", None):
                        p_val = float(fusa_prob)
                        if target > 0 and p_val > 0:
                            spi_val = f"{math.log10(target / p_val):.2f}"
                except Exception:
                    spi_val = ""
                iid = tree.insert(
                    "",
                    "end",
                    values=[
                        sg.user_name or f"SG {sg.unique_id}",
                        v_str,
                        p_str,
                        spi_val,
                        "Target PMHF",
                        getattr(sg, "acceptance_criteria", ""),
                    ],
                )
                self._spi_lookup[iid] = (sg, "FUSA")

    def refresh_safety_case_table(self):
        """Populate the Safety & Security Case table with solution nodes."""
        tree = getattr(self, "_safety_case_tree", None)
        if not tree or not getattr(tree, "winfo_exists", lambda: True)():
            return
        for iid in list(tree.get_children("")):
            tree.delete(iid)
        self._solution_lookup = {}
        for diag in getattr(self, "all_gsn_diagrams", []):
            for node in getattr(diag, "nodes", []):
                if (
                    getattr(node, "node_type", "").lower() == "solution"
                    and getattr(node, "is_primary_instance", True)
                ):
                    self._solution_lookup[node.unique_id] = (node, diag)
                    prob = ""
                    v_target = ""
                    spi_val = ""
                    p_val = None
                    vt_val = None
                    target = getattr(node, "spi_target", "")
                    if target:
                        pg_name, spi_type = self._parse_spi_target(target)
                        te = None
                        for candidate in getattr(self, "top_events", []):
                            if self._product_goal_name(candidate) == pg_name:
                                te = candidate
                                break
                        if te:
                            if spi_type == "FUSA":
                                p = getattr(te, "probability", "")
                                vt = PMHF_TARGETS.get(getattr(te, "safety_goal_asil", ""), "")
                            else:
                                p = getattr(te, "spi_probability", "")
                                vt = getattr(te, "validation_target", "")
                            if p not in ("", None):
                                try:
                                    p_val = float(p)
                                    prob = f"{p_val:.2e}"
                                except Exception:
                                    prob = ""
                                    p_val = None
                            if vt not in ("", None):
                                try:
                                    vt_val = float(vt)
                                    v_target = f"{vt_val:.2e}"
                                except Exception:
                                    v_target = ""
                                    vt_val = None
                            try:
                                if vt_val not in (None, 0) and p_val not in (None, 0):
                                    spi_val = f"{math.log10(vt_val / p_val):.2f}"
                            except Exception:
                                spi_val = ""
                    tree.insert(
                        "",
                        "end",
                        values=[
                            node.user_name,
                            node.description,
                            node.work_product,
                            node.evidence_link,
                            v_target,
                            prob,
                            spi_val,
                            CHECK_MARK if getattr(node, "evidence_sufficient", False) else "",
                            getattr(node, "manager_notes", ""),
                        ],
                        tags=(node.unique_id,),
                    )

    def show_safety_case(self):
        """Display table of all solution nodes from GSN diagrams."""
        if hasattr(self, "_safety_case_tab") and self._safety_case_tab.winfo_exists():
            self.doc_nb.select(self._safety_case_tab)
            self.refresh_safety_case_table()
            return
        self._safety_case_tab = self.lifecycle_ui._new_tab("Safety & Security Case")
        win = self._safety_case_tab

        columns = [
            "Solution",
            "Description",
            "Work Product",
            "Evidence Link",
            "Validation Target",
            "Achieved Probability",
            "SPI",
            "Evidence OK",
            "Notes",
        ]
        if hasattr(win, "tk"):
            tree_frame = ttk.Frame(win)
            tree_frame.pack(fill=tk.BOTH, expand=True)
            tree = ttk.Treeview(
                tree_frame, columns=columns, show="headings", selectmode="browse"
            )
            for c in columns:
                tree.heading(c, text=c)
                width = 120
                if c in ("Description", "Evidence Link", "Notes"):
                    width = 200
                tree.column(c, width=width, anchor="center")
            vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
            hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=tree.xview)
            tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
            tree.grid(row=0, column=0, sticky="nsew")
            vsb.grid(row=0, column=1, sticky="ns")
            hsb.grid(row=1, column=0, sticky="ew")
            tree_frame.rowconfigure(0, weight=1)
            tree_frame.columnconfigure(0, weight=1)
        else:
            tree = ttk.Treeview(win, columns=columns, show="headings", selectmode="browse")
            for c in columns:
                tree.heading(c, text=c)
                width = 120
                if c in ("Description", "Evidence Link", "Notes"):
                    width = 200
                tree.column(c, width=width, anchor="center")
            tree.pack(fill=tk.BOTH, expand=True)
        self._safety_case_tree = tree
        self._solution_lookup = {}

        def on_double_click(event):
            row = tree.identify_row(event.y)
            col = tree.identify_column(event.x)
            if not row or not col:
                return
            idx = int(col[1:]) - 1
            col_name = columns[idx]
            tags = tree.item(row, "tags")
            if not tags:
                return
            uid = tags[0]
            node_diag = self._solution_lookup.get(uid)
            if not node_diag:
                return
            node = node_diag[0]
            if col_name == "Evidence OK":
                current = tree.set(row, "Evidence OK")
                new_val = "" if current == CHECK_MARK else CHECK_MARK
                if messagebox.askokcancel("Evidence", "Are you sure?"):
                    self.push_undo_state()
                    tree.set(row, "Evidence OK", new_val)
                    node.evidence_sufficient = new_val == CHECK_MARK
            elif col_name == "Achieved Probability":
                target = getattr(node, "spi_target", "")
                pg_name, spi_type = self._parse_spi_target(target)
                te = None
                for sg in getattr(self, "top_events", []):
                    if self._product_goal_name(sg) == pg_name:
                        te = sg
                        break
                if te:
                    attr = "probability" if spi_type == "FUSA" else "spi_probability"
                    new_val = simpledialog.askfloat(
                        "Achieved Probability",
                        "Enter achieved probability:",
                        initialvalue=getattr(te, attr, 0.0),
                    )
                    if new_val is not None:
                        self.push_undo_state()
                        setattr(te, attr, float(new_val))
                        self.refresh_safety_case_table()
                        self.refresh_safety_performance_indicators()
                        self.update_views()
            elif col_name == "Notes":
                current = tree.set(row, "Notes")
                new_val = simpledialog.askstring(
                    "Notes", "Enter notes:", initialvalue=current
                )
                if new_val is not None:
                    self.push_undo_state()
                    tree.set(row, "Notes", new_val)
                    node.manager_notes = new_val

        for seq in ("<Double-Button-1>", "<Double-1>"):
            tree.bind(seq, on_double_click)

        def edit_selected(row=None):
            if row is None:
                sel = tree.selection()
                if not sel:
                    return
                row = sel[0]
            tags = tree.item(row, "tags")
            if not tags:
                return
            uid = tags[0]
            node_diag = self._solution_lookup.get(uid)
            if not node_diag:
                return
            node, diag = node_diag
            self.push_undo_state()
            GSNElementConfig(win, node, diag)
            self.refresh_safety_case_table()

        self._edit_safety_case_item = edit_selected

        def export_csv():
            path = filedialog.asksaveasfilename(
                defaultextension=".csv", filetypes=[("CSV", "*.csv")]
            )
            if not path:
                return
            with open(path, "w", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(columns)
                for iid in tree.get_children():
                    writer.writerow(tree.item(iid, "values"))
            messagebox.showinfo("Export", "Safety & Security Case exported")

        self.export_safety_case_csv = export_csv

        btn = ttk.Button(win, text="Edit", command=edit_selected)
        btn.pack(pady=4)
        ttk.Button(win, text="Export CSV", command=export_csv).pack(pady=4)

        menu = tk.Menu(win, tearoff=0)
        menu.add_command(label="Edit", command=edit_selected)
        menu.add_command(label="Export CSV", command=export_csv)

        def on_right_click(event):
            row = tree.identify_row(event.y)
            if row:
                tree.selection_set(row)
                menu.post(event.x_root, event.y_root)

        tree.bind("<Button-3>", on_right_click)

        self.refresh_safety_case_table()

    def export_product_goal_requirements(self):
        return self.reporting_export.export_product_goal_requirements()
    def generate_phase_requirements(self, phase: str) -> None:
        """Generate requirements for all governance diagrams in a phase."""
        self.open_safety_management_toolbox(show_diagrams=False)
        win = getattr(self, "safety_mgmt_window", None)
        if win:
            win.generate_phase_requirements(phase)

    def generate_lifecycle_requirements(self) -> None:
        """Generate requirements for all governance diagrams outside phases."""
        self.open_safety_management_toolbox(show_diagrams=False)
        win = getattr(self, "safety_mgmt_window", None)
        if win:
            win.generate_lifecycle_requirements()


    def _refresh_phase_requirements_menu(self) -> None:
        if not hasattr(self, "phase_req_menu"):
            return
        self.phase_req_menu.delete(0, tk.END)
        toolbox = getattr(self, "safety_mgmt_toolbox", None)
        if not toolbox:
            return
        phases = sorted(toolbox.list_modules())
        for phase in phases:
            # Use ``functools.partial`` to bind the phase name at creation time
            # so each menu entry triggers generation for its own phase.
            self.phase_req_menu.add_command(
                label=phase,
                command=partial(self.generate_phase_requirements, phase),
            )
        if phases:
            self.phase_req_menu.add_separator()
        self.phase_req_menu.add_command(
            label="Lifecycle",
            command=self.generate_lifecycle_requirements,
        )

    def export_cybersecurity_goal_requirements(self):
        return self.reporting_export.export_cybersecurity_goal_requirements()

    def build_cause_effect_data(self):
        return self.probability_reliability.build_cause_effect_data()

    def _build_cause_effect_graph(self, row):
        return self.probability_reliability._build_cause_effect_graph(row)

    def render_cause_effect_diagram(self, row):
        """Render *row* as a PIL image matching the on-screen diagram."""
        try:
            from PIL import Image, ImageDraw, ImageFont
        except Exception:
            return None
        import textwrap, math

        nodes, edges, pos = self._build_cause_effect_graph(row)
        color_map = {
            "hazard": "#F08080",
            "malfunction": "#ADD8E6",
            "failure_mode": "#FFA500",
            "fault": "#D3D3D3",
            "fi": "#FFFFE0",
            "tc": "#90EE90",
            "attack_path": "#E0FFFF",
            "threat": "#FFB6C1",
        }

        scale = 80
        x_off = 50
        y_off = 50
        box_w = 80
        box_h = 40

        max_x = max(x for x, _ in pos.values())
        max_y = max(y for _, y in pos.values())
        width = int(x_off * 2 + scale * max_x + box_w)
        height = int(y_off * 2 + scale * max_y + box_h)

        img = Image.new("RGB", (width, height), "white")
        draw = ImageDraw.Draw(img)
        font = ImageFont.load_default()

        def to_canvas(x: float, y: float) -> tuple[float, float]:
            return x_off + scale * x, y_off + scale * y

        for u, v in edges:
            x1, y1 = to_canvas(*pos[u])
            x2, y2 = to_canvas(*pos[v])
            draw.line((x1, y1, x2, y2), fill="black")
            dx, dy = x2 - x1, y2 - y1
            length = math.hypot(dx, dy) or 1
            ux, uy = dx / length, dy / length
            arrow = 10
            px, py = x2 - arrow * ux, y2 - arrow * uy
            perp = (-uy, ux)
            left = (px + perp[0] * arrow / 2, py + perp[1] * arrow / 2)
            right = (px - perp[0] * arrow / 2, py - perp[1] * arrow / 2)
            draw.polygon([ (x2, y2), left, right ], fill="black")
            if hasattr(draw, "text"):
                draw.text(((x1 + x2) / 2, (y1 + y2) / 2), "caused by", fill="black", font=font, anchor="mm")

        for n, (x, y) in pos.items():
            label, kind = nodes.get(n, (n, ""))
            color = color_map.get(kind, "white")
            cx, cy = to_canvas(x, y)
            rect = [cx - box_w / 2, cy - box_h / 2, cx + box_w / 2, cy + box_h / 2]
            draw.rectangle(
                rect,
                fill=color,
                outline=StyleManager.get_instance().outline_color,
            )
            text = textwrap.fill(str(label), 20)
            bbox = draw.multiline_textbbox((0, 0), text, font=font)
            tw = bbox[2] - bbox[0]
            th = bbox[3] - bbox[1]
            draw.multiline_text((cx - tw / 2, cy - th / 2), text, font=font, align="center")

        return img
    def show_cause_effect_chain(self):
        return self.safety_analysis.show_cause_effect_chain()

    def show_cut_sets(self):
        return self.safety_analysis.show_cut_sets()

    def show_common_cause_view(self):
        return self.safety_analysis.show_common_cause_view()

    def manage_mission_profiles(self):
        if hasattr(self, "_mp_tab") and self._mp_tab.winfo_exists():
            self.doc_nb.select(self._mp_tab)
            return
        self._mp_tab = self.lifecycle_ui._new_tab("Mission Profiles")
        win = self._mp_tab
        listbox = tk.Listbox(win, height=8, width=40)
        listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        btn_frame = ttk.Frame(win)
        btn_frame.pack(side=tk.RIGHT, fill=tk.Y)

        def refresh():
            listbox.delete(0, tk.END)
            for mp in self.mission_profiles:
                if mp is None:
                    continue
                info = (
                    f"{mp.name} (on: {mp.tau_on}h, off: {mp.tau_off}h, "
                    f"board: {mp.board_temp}\u00b0C, ambient: {mp.ambient_temp}\u00b0C)"
                )
                listbox.insert(tk.END, info)

        class MPDialog(simpledialog.Dialog):
            def __init__(self, master, mp=None):
                self.mp = mp
                super().__init__(master)

            def body(self, master):
                self.vars = {}
                fields = [
                    ("Name", "name"),
                    ("TAU On (h)", "tau_on"),
                    ("TAU Off (h)", "tau_off"),
                    ("Board Temp (\u00b0C)", "board_temp"),
                    ("Board Temp Min (\u00b0C)", "board_temp_min"),
                    ("Board Temp Max (\u00b0C)", "board_temp_max"),
                    ("Ambient Temp (\u00b0C)", "ambient_temp"),
                    ("Ambient Temp Min (\u00b0C)", "ambient_temp_min"),
                    ("Ambient Temp Max (\u00b0C)", "ambient_temp_max"),
                    ("Humidity (%)", "humidity"),
                    ("Duty Cycle", "duty_cycle"),
                    ("Notes", "notes"),
                ]
                self.entries = {}
                for row, (label, key) in enumerate(fields):
                    ttk.Label(master, text=label).grid(row=row, column=0, padx=5, pady=5, sticky="e")
                    var = tk.StringVar()
                    if self.mp:
                        var.set(str(getattr(self.mp, key)))
                    state = "readonly" if key == "duty_cycle" else "normal"
                    entry = ttk.Entry(master, textvariable=var, state=state)
                    entry.grid(row=row, column=1, padx=5, pady=5)
                    self.vars[key] = var
                    self.entries[key] = entry

                def update_dc(*_):
                    try:
                        on = float(self.vars["tau_on"].get() or 0)
                        off = float(self.vars["tau_off"].get() or 0)
                        total = on + off
                        dc = on / total if total else 0.0
                    except ValueError:
                        dc = 0.0
                    self.vars["duty_cycle"].set(str(dc))

                self.vars["tau_on"].trace_add("write", update_dc)
                self.vars["tau_off"].trace_add("write", update_dc)
                update_dc()

            def apply(self):
                vals = {k: v.get() for k, v in self.vars.items()}
                tau_on = float(vals.get("tau_on") or 0.0)
                tau_off = float(vals.get("tau_off") or 0.0)
                total = tau_on + tau_off
                dc = tau_on / total if total else 0.0
                if self.mp is None:
                    mp = MissionProfile(
                        vals["name"],
                        tau_on,
                        tau_off,
                        float(vals["board_temp"] or 25.0),
                        float(vals["board_temp_min"] or 25.0),
                        float(vals["board_temp_max"] or 25.0),
                        float(vals["ambient_temp"] or 25.0),
                        float(vals["ambient_temp_min"] or 25.0),
                        float(vals["ambient_temp_max"] or 25.0),
                        float(vals["humidity"] or 50.0),
                        dc,
                        vals["notes"],
                    )
                    self.result = mp
                else:
                    self.mp.name = vals["name"]
                    self.mp.tau_on = tau_on
                    self.mp.tau_off = tau_off
                    self.mp.board_temp = float(vals["board_temp"] or 25.0)
                    self.mp.board_temp_min = float(vals["board_temp_min"] or 25.0)
                    self.mp.board_temp_max = float(vals["board_temp_max"] or 25.0)
                    self.mp.ambient_temp = float(vals["ambient_temp"] or 25.0)
                    self.mp.ambient_temp_min = float(vals["ambient_temp_min"] or 25.0)
                    self.mp.ambient_temp_max = float(vals["ambient_temp_max"] or 25.0)
                    self.mp.humidity = float(vals["humidity"] or 50.0)
                    self.mp.duty_cycle = dc
                    self.mp.notes = vals["notes"]
                    self.result = self.mp

        def add_profile():
            dlg = MPDialog(win)
            if getattr(dlg, "result", None) is not None:
                self.mission_profiles.append(dlg.result)
                refresh()
                if hasattr(self, "_rel_window") and self._rel_window.winfo_exists():
                    self._rel_window.refresh_tree()

        def edit_profile():
            sel = listbox.curselection()
            if not sel:
                return
            mp = self.mission_profiles[sel[0]]
            dlg = MPDialog(win, mp)
            if getattr(dlg, "result", None) is not None:
                refresh()
                if hasattr(self, "_rel_window") and self._rel_window.winfo_exists():
                    self._rel_window.refresh_tree()

        def delete_profile():
            sel = listbox.curselection()
            if not sel:
                return
            del self.mission_profiles[sel[0]]
            refresh()
            if hasattr(self, "_rel_window") and self._rel_window.winfo_exists():
                self._rel_window.refresh_tree()

        ttk.Button(btn_frame, text="Add", command=add_profile).pack(fill=tk.X)
        ttk.Button(btn_frame, text="Edit", command=edit_profile).pack(fill=tk.X)
        ttk.Button(btn_frame, text="Delete", command=delete_profile).pack(fill=tk.X)

        refresh()

    def manage_mechanism_libraries(self):
        return self.open_windows_features.manage_mechanism_libraries()

    def manage_scenario_libraries(self):
        if hasattr(self, "_scen_tab") and self._scen_tab.winfo_exists():
            self.doc_nb.select(self._scen_tab)
            return
        self._scen_tab = self.lifecycle_ui._new_tab("Scenario Libraries")
        win = self._scen_tab
        lib_lb = tk.Listbox(win, height=8, width=25)
        lib_lb.grid(row=0, column=0, rowspan=4, sticky="nsew")
        scen_tree = ttk.Treeview(
            win,
            columns=("cls", "beh", "sce", "tc", "fi", "exp", "desc"),
            show="tree headings",
        )
        scen_tree.heading("#0", text="Name")
        scen_tree.column("#0", width=150)
        scen_tree.heading("cls", text="Class")
        scen_tree.column("cls", width=100)
        scen_tree.heading("beh", text="Other Users")
        scen_tree.column("beh", width=140)
        scen_tree.heading("sce", text="Scenery")
        scen_tree.column("sce", width=140)
        scen_tree.heading("tc", text="TC")
        scen_tree.column("tc", width=80)
        scen_tree.heading("fi", text="FI")
        scen_tree.column("fi", width=80)
        scen_tree.heading("exp", text="Exposure")
        scen_tree.column("exp", width=80)
        scen_tree.heading("desc", text="Description")
        scen_tree.column("desc", width=200)
        scen_tree.grid(row=0, column=1, columnspan=3, sticky="nsew")
        win.grid_rowconfigure(0, weight=1)
        win.grid_columnconfigure(1, weight=1)

        if not hasattr(self, "scenario_icon"):
            self.scenario_icon = self._create_icon("circle", "#1e90ff")

        def refresh_libs():
            lib_lb.delete(0, tk.END)
            for lib in self.scenario_libraries:
                lib_lb.insert(tk.END, lib.get("name", ""))
            refresh_scenarios()

        def refresh_scenarios(*_):
            scen_tree.delete(*scen_tree.get_children())
            sel = lib_lb.curselection()
            if not sel:
                return
            lib = self.scenario_libraries[sel[0]]
            for sc in lib.get("scenarios", []):
                if isinstance(sc, dict):
                    name = sc.get("name", "")
                    cls = sc.get("class", "")
                    beh = sc.get("behavior", "")
                    sce = sc.get("scenery", "")
                    tc = sc.get("tc", "")
                    fi = sc.get("fi", "")
                    exp = sc.get("exposure", "")
                    desc = sc.get("description", "")
                else:
                    name = str(sc)
                    cls = beh = sce = tc = fi = exp = desc = ""
                scen_tree.insert(
                    "",
                    tk.END,
                    text=name,
                    values=(cls, beh, sce, tc, fi, exp, desc),
                    image=self.scenario_icon,
                )

        class LibraryDialog(simpledialog.Dialog):
            def __init__(self, parent, app, data=None):
                self.app = app
                self.data = data or {"name": "", "odds": []}
                super().__init__(parent, title="Edit Library")

            def body(self, master):
                ttk.Label(master, text="Name").grid(row=0, column=0, sticky="e")
                self.name_var = tk.StringVar(value=self.data.get("name", ""))
                ttk.Entry(master, textvariable=self.name_var).grid(row=0, column=1, sticky="ew")
                ttk.Label(master, text="ODD Libraries").grid(row=1, column=0, sticky="ne")
                toolbox = getattr(self.app, "safety_mgmt_toolbox", None) or ACTIVE_TOOLBOX
                self.allowed_inputs = bool(
                    toolbox and "ODD" in toolbox.analysis_inputs("Scenario Library")
                )
                self.lb = tk.Listbox(master, selectmode=tk.MULTIPLE, height=5)
                if self.allowed_inputs:
                    for i, lib in enumerate(self.app.odd_libraries):
                        self.lb.insert(tk.END, lib.get("name", ""))
                        if lib.get("name", "") in self.data.get("odds", []):
                            self.lb.selection_set(i)
                else:
                    self.lb.configure(state=tk.DISABLED)
                self.lb.grid(row=1, column=1, sticky="nsew")
                master.grid_rowconfigure(1, weight=1)
                master.grid_columnconfigure(1, weight=1)

            def apply(self):
                self.data["name"] = self.name_var.get()
                sels = self.lb.curselection() if self.allowed_inputs else []
                self.data["odds"] = [
                    self.app.odd_libraries[i].get("name", "") for i in sels
                ]

        class ScenarioDialog(simpledialog.Dialog):
            def __init__(self, parent, app, lib, data=None):
                self.app = app
                self.lib = lib
                self.data = data or {
                    "name": "",
                    "class": "",
                    "behavior": "",
                    "action": "",
                    "scenery": "",
                    "tc": "",
                    "fi": "",
                    "exposure": 1,
                    "description": "",
                }
                self.tag_counter = 0
                super().__init__(parent, title="Edit Scenario")

            def body(self, master):
                ttk.Label(master, text="Name").grid(row=0, column=0, sticky="e")
                self.name_var = tk.StringVar(value=self.data.get("name", ""))
                ttk.Entry(master, textvariable=self.name_var).grid(row=0, column=1, sticky="ew")

                ttk.Label(master, text="Scenario Class").grid(row=1, column=0, sticky="e")
                self.cls_var = tk.StringVar(value=self.data.get("class", ""))
                cls_opts = ["Frontal", "Side", "Rear", "Free"]
                ttk.Combobox(master, textvariable=self.cls_var, values=cls_opts, state="readonly").grid(row=1, column=1, sticky="ew")

                ttk.Label(master, text="Other Road Users").grid(row=2, column=0, sticky="e")
                self.beh_var = tk.StringVar(value=self.data.get("behavior", ""))
                ttk.Entry(master, textvariable=self.beh_var).grid(row=2, column=1, sticky="ew")

                ttk.Label(master, text="Action of Other Road Users").grid(row=3, column=0, sticky="e")
                self.act_var = tk.StringVar(value=self.data.get("action", ""))
                ttk.Entry(master, textvariable=self.act_var).grid(row=3, column=1, sticky="ew")

                ttk.Label(master, text="Scenery").grid(row=4, column=0, sticky="e")
                self.sce_var = tk.StringVar(value=self.data.get("scenery", ""))
                ttk.Entry(master, textvariable=self.sce_var, state="readonly").grid(
                    row=4, column=1, sticky="ew"
                )

                elems = []
                self.elem_classes = {}
                self.elem_params = {}
                for name in self.lib.get("odds", []):
                    for l in self.app.odd_libraries:
                        if l.get("name") == name:
                            for el in l.get("elements", []):
                                if isinstance(el, dict):
                                    val = el.get("name") or el.get("element") or el.get("id")
                                    cls = el.get("class", "")
                                    params = []
                                    for k, v in el.items():
                                        if k in {"name", "element", "id", "class"}:
                                            continue
                                        if isinstance(v, (list, tuple, set)):
                                            params.extend(str(x) for x in v if x)
                                        elif v:
                                            params.append(str(v))
                                else:
                                    val = str(el)
                                    cls = ""
                                    params = []
                                if val:
                                    elems.append(val)
                                    self.elem_classes[val] = cls
                                    self.elem_params[val] = params

                ttk.Label(master, text="ODD Elements").grid(row=5, column=0, sticky="e")
                self.elem_list = tk.Listbox(
                    master, selectmode=tk.MULTIPLE, height=5, exportselection=False
                )
                for el in elems:
                    self.elem_list.insert(tk.END, el)
                selected = {
                    s.strip()
                    for s in str(self.data.get("scenery", "")).split(",")
                    if s.strip()
                }
                for idx, el in enumerate(elems):
                    if el in selected:
                        self.elem_list.selection_set(idx)
                self.elem_list.grid(row=5, column=1, sticky="nsew")
                self.elem_list.bind("<<ListboxSelect>>", lambda e: self.update_description())
                master.grid_rowconfigure(5, weight=1)

                tc_names = [n.user_name or f"TC {n.unique_id}" for n in self.app.get_all_triggering_conditions()]
                fi_names = [n.user_name or f"FI {n.unique_id}" for n in self.app.get_all_functional_insufficiencies()]
                ttk.Label(master, text="Triggering Condition").grid(row=6, column=0, sticky="e")
                self.tc_var = tk.StringVar(value=self.data.get("tc", ""))
                ttk.Combobox(master, textvariable=self.tc_var, values=tc_names, state="readonly").grid(row=6, column=1, sticky="ew")
                ttk.Label(master, text="Functional Insufficiency").grid(row=7, column=0, sticky="e")
                self.fi_var = tk.StringVar(value=self.data.get("fi", ""))
                ttk.Combobox(master, textvariable=self.fi_var, values=fi_names, state="readonly").grid(row=7, column=1, sticky="ew")

                ttk.Label(master, text="Exposure").grid(row=8, column=0, sticky="e")
                self.exp_var = tk.StringVar(value=str(self.data.get("exposure", 1)))
                ttk.Combobox(
                    master,
                    textvariable=self.exp_var,
                    values=["1", "2", "3", "4"],
                    state="readonly",
                ).grid(row=8, column=1, sticky="ew")

                ttk.Label(master, text="Description").grid(row=9, column=0, sticky="ne")
                self.desc = tk.Text(master, height=4, width=40, wrap="word")
                self.desc.grid(row=9, column=1, columnspan=3, sticky="nsew")
                self.load_desc_links()
                master.grid_columnconfigure(1, weight=1)

                # Automatically update description on parameter changes
                for var in (
                    self.cls_var,
                    self.beh_var,
                    self.act_var,
                    self.tc_var,
                    self.fi_var,
                ):
                    var.trace_add("write", lambda *a: self.update_description())
                self.update_description()

            def update_description(self, *args):
                names = [self.elem_list.get(i) for i in self.elem_list.curselection()]
                self.sce_var.set(", ".join(names))
                odds = [
                    (
                        n,
                        self.elem_classes.get(n, ""),
                        self.elem_params.get(n, []),
                    )
                    for n in names
                ]
                phrases = template_phrases(
                    self.cls_var.get(),
                    self.beh_var.get(),
                    self.act_var.get(),
                    odds,
                    self.tc_var.get(),
                    self.fi_var.get(),
                )
                text = " ".join(phrases)
                self.desc.delete("1.0", "end")
                self.desc.insert("1.0", text)
                for m in re.finditer(r"\[\[(.+?)\]\]", text):
                    name = m.group(1)
                    start = f"1.0+{m.start()}c"
                    end = f"1.0+{m.end()}c"
                    tag = f"link{self.tag_counter}"
                    self.tag_counter += 1
                    self.desc.tag_add(tag, start, end)
                    self.desc.tag_config(tag, foreground="blue", underline=1)
                    self.desc.tag_bind(tag, "<Button-1>", lambda e, n=name: self.show_elem(n))

            def load_desc_links(self):
                desc = self.data.get("description", "")
                self.desc.insert("1.0", desc)
                for m in re.finditer(r"\[\[(.+?)\]\]", desc):
                    name = m.group(1)
                    start = f"1.0+{m.start()}c"
                    end = f"1.0+{m.end()}c"
                    tag = f"link{self.tag_counter}"
                    self.tag_counter += 1
                    self.desc.tag_add(tag, start, end)
                    self.desc.tag_config(tag, foreground="blue", underline=1)
                    self.desc.tag_bind(tag, "<Button-1>", lambda e, n=name: self.show_elem(n))

            def show_elem(self, name):
                for lib_name in self.lib.get("odds", []):
                    for l in self.app.odd_libraries:
                        if l.get("name") == lib_name:
                            for el in l.get("elements", []):
                                val = el.get("name") or el.get("element") or el.get("id")
                                if val == name:
                                    msg = "\n".join(f"{k}: {v}" for k, v in el.items())
                                    messagebox.showinfo("ODD Element", msg)
                                    return
                messagebox.showinfo("ODD Element", f"{name}")

            def apply(self):
                self.data["name"] = self.name_var.get()
                self.data["class"] = self.cls_var.get()
                self.data["behavior"] = self.beh_var.get()
                self.data["action"] = self.act_var.get()
                names = [self.elem_list.get(i) for i in self.elem_list.curselection()]
                self.data["scenery"] = ", ".join(names)
                self.data["tc"] = self.tc_var.get()
                self.data["fi"] = self.fi_var.get()
                try:
                    self.data["exposure"] = int(self.exp_var.get())
                except (TypeError, ValueError):
                    self.data["exposure"] = 1
                self.data["description"] = self.desc.get("1.0", "end-1c")

        def add_lib():
            dlg = LibraryDialog(win, self)
            if dlg.data.get("name"):
                self.scenario_libraries.append({"name": dlg.data["name"], "scenarios": [], "odds": dlg.data["odds"]})
                refresh_libs()

        def edit_lib():
            sel = lib_lb.curselection()
            if not sel:
                return
            lib = self.scenario_libraries[sel[0]]
            dlg = LibraryDialog(win, self, lib)
            lib.update(dlg.data)
            refresh_libs()

        def delete_lib():
            sel = lib_lb.curselection()
            if sel:
                idx = sel[0]
                del self.scenario_libraries[idx]
                refresh_libs()

        def add_scen():
            sel = lib_lb.curselection()
            if not sel:
                return
            lib = self.scenario_libraries[sel[0]]
            dlg = ScenarioDialog(win, self, lib)
            if dlg.data.get("name"):
                lib.setdefault("scenarios", []).append(dlg.data)
                refresh_scenarios()

        def edit_scen():
            sel_lib = lib_lb.curselection()
            sel_sc = scen_tree.selection()
            if not sel_lib or not sel_sc:
                return
            lib = self.scenario_libraries[sel_lib[0]]
            idx = scen_tree.index(sel_sc[0])
            data = lib.get("scenarios", [])[idx]
            dlg = ScenarioDialog(win, self, lib, data)
            lib["scenarios"][idx] = dlg.data
            refresh_scenarios()

        def del_scen():
            sel_lib = lib_lb.curselection()
            sel_sc = scen_tree.selection()
            if not sel_lib or not sel_sc:
                return
            lib = self.scenario_libraries[sel_lib[0]]
            idx = scen_tree.index(sel_sc[0])
            del lib.get("scenarios", [])[idx]
            refresh_scenarios()

        btnf = ttk.Frame(win)
        btnf.grid(row=1, column=1, columnspan=3, sticky="ew")
        ttk.Button(btnf, text="Add Lib", command=add_lib).pack(side=tk.LEFT)
        ttk.Button(btnf, text="Edit Lib", command=edit_lib).pack(side=tk.LEFT)
        ttk.Button(btnf, text="Del Lib", command=delete_lib).pack(side=tk.LEFT)
        ttk.Button(btnf, text="Add Scen", command=add_scen).pack(side=tk.LEFT, padx=5)
        ttk.Button(btnf, text="Edit Scen", command=edit_scen).pack(side=tk.LEFT)
        ttk.Button(btnf, text="Del Scen", command=del_scen).pack(side=tk.LEFT)

        lib_lb.bind("<<ListboxSelect>>", refresh_scenarios)
        refresh_libs()

    def manage_odd_libraries(self):
        if hasattr(self, "_odd_tab") and self._odd_tab.winfo_exists():
            self.doc_nb.select(self._odd_tab)
            return
        self._odd_tab = self.lifecycle_ui._new_tab("ODD Libraries")
        win = self._odd_tab
        lib_lb = tk.Listbox(win, height=8, width=25)
        lib_lb.grid(row=0, column=0, rowspan=4, sticky="nsew")
        elem_tree = ttk.Treeview(win, columns=("cls", "attrs"), show="tree headings")
        elem_tree.heading("#0", text="Name")
        elem_tree.column("#0", width=150)
        elem_tree.heading("cls", text="Class")
        elem_tree.column("cls", width=120)
        elem_tree.heading("attrs", text="Attributes")
        elem_tree.column("attrs", width=200)
        elem_tree.grid(row=0, column=1, columnspan=3, sticky="nsew")
        win.grid_rowconfigure(0, weight=1)
        win.grid_columnconfigure(1, weight=1)

        if not hasattr(self, "odd_elem_icon"):
            self.odd_elem_icon = self._create_icon("rect", "#696969")

        def refresh_libs():
            lib_lb.delete(0, tk.END)
            for lib in self.odd_libraries:
                lib_lb.insert(tk.END, lib.get("name", ""))
            refresh_elems()

        def refresh_elems(*_):
            elem_tree.delete(*elem_tree.get_children())
            sel = lib_lb.curselection()
            if not sel:
                return
            lib = self.odd_libraries[sel[0]]
            for el in lib.get("elements", []):
                name = el.get("name") or el.get("element") or el.get("id")
                cls = el.get("class", "")
                attrs = ", ".join(
                    f"{k}={v}" for k, v in el.items() if k not in {"name", "class"}
                )
                opts = {"values": (cls, attrs), "text": name}
                if self.odd_elem_icon:
                    opts["image"] = self.odd_elem_icon
                elem_tree.insert("", tk.END, **opts)

        class ElementDialog(simpledialog.Dialog):
            def __init__(self, parent, app, data=None):
                self.app = app
                self.data = data or {"name": "", "class": ""}
                super().__init__(parent, title="Edit Element")

            def add_attr_row(self, key="", val=""):
                r = len(self.attr_rows)
                k_var = tk.StringVar(value=key)
                v_var = tk.StringVar(value=str(val))
                k_entry = ttk.Entry(self.attr_frame, textvariable=k_var)
                v_entry = ttk.Entry(self.attr_frame, textvariable=v_var)
                k_entry.grid(row=r, column=0, padx=2, pady=2)
                v_entry.grid(row=r, column=1, padx=2, pady=2)

                row = {}

                def remove_row():
                    k_entry.destroy()
                    v_entry.destroy()
                    del_btn.destroy()
                    self.attr_rows.remove(row)
                    for i, rdata in enumerate(self.attr_rows):
                        rdata["k_entry"].grid_configure(row=i)
                        rdata["v_entry"].grid_configure(row=i)
                        rdata["del_btn"].grid_configure(row=i)

                del_btn = ttk.Button(self.attr_frame, text="Delete", command=remove_row)
                del_btn.grid(row=r, column=2, padx=2, pady=2)

                row.update(
                    {
                        "k_var": k_var,
                        "v_var": v_var,
                        "k_entry": k_entry,
                        "v_entry": v_entry,
                        "del_btn": del_btn,
                    }
                )
                self.attr_rows.append(row)

            def body(self, master):
                ttk.Label(master, text="Name").grid(row=0, column=0, sticky="e")
                self.name_var = tk.StringVar(value=self.data.get("name", ""))
                ttk.Entry(master, textvariable=self.name_var).grid(row=0, column=1, sticky="ew")

                ttk.Label(master, text="Class").grid(row=1, column=0, sticky="e")
                self.class_var = tk.StringVar(value=self.data.get("class", ""))
                cls_opts = [
                    "Road",
                    "Infrastructure",
                    "Temporal",
                    "Movable",
                    "Environment",
                ]
                ttk.Combobox(master, textvariable=self.class_var, values=cls_opts, state="readonly").grid(row=1, column=1, sticky="ew")

                nb = ttk.Notebook(master)
                nb.grid(row=2, column=0, columnspan=2, sticky="nsew")
                master.grid_rowconfigure(2, weight=1)
                master.grid_columnconfigure(1, weight=1)

                # Attributes tab
                self.attr_frame = ttk.Frame(nb)
                nb.add(self.attr_frame, text="Attributes")
                self.attr_rows = []
                for k, v in self.data.items():
                    if k not in {"name", "class", "p", "n", "tp", "fp", "tn", "fn"}:
                        self.add_attr_row(k, v)
                ttk.Button(self.attr_frame, text="Add Attribute", command=self.add_attr_row).grid(row=99, column=0, columnspan=3, pady=5)

                # Confusion matrix tab
                cm_frame = ttk.Frame(nb)
                nb.add(cm_frame, text="Confusion Matrix")
                self.p_var = tk.DoubleVar(value=float(self.data.get("p", 0) or 0))
                self.n_var = tk.DoubleVar(value=float(self.data.get("n", 0) or 0))
                self.tp_var = tk.DoubleVar(value=float(self.data.get("tp", 0) or 0))
                self.fp_var = tk.DoubleVar(value=float(self.data.get("fp", 0) or 0))
                self.tn_var = tk.DoubleVar(value=float(self.data.get("tn", 0) or 0))
                self.fn_var = tk.DoubleVar(value=float(self.data.get("fn", 0) or 0))

                matrix_metrics = ttk.Frame(cm_frame)
                matrix_metrics.grid(row=0, column=0, pady=5, sticky="w")
                matrix = ttk.Frame(matrix_metrics)
                matrix.grid(row=0, column=0, sticky="w")
                ttk.Label(matrix, text="").grid(row=0, column=0)
                ttk.Label(matrix, text="Pred P").grid(row=0, column=1)
                ttk.Label(matrix, text="Pred N").grid(row=0, column=2)
                ttk.Label(matrix, text="Actual P").grid(row=1, column=0)
                ttk.Entry(matrix, textvariable=self.tp_var, width=6).grid(row=1, column=1)
                ttk.Entry(matrix, textvariable=self.fn_var, width=6).grid(row=1, column=2)
                ttk.Label(matrix, text="Actual N").grid(row=2, column=0)
                ttk.Entry(matrix, textvariable=self.fp_var, width=6).grid(row=2, column=1)
                ttk.Entry(matrix, textvariable=self.tn_var, width=6).grid(row=2, column=2)

                metrics_frame = ttk.Frame(matrix_metrics)
                metrics_frame.grid(row=0, column=1, padx=10, sticky="n")
                ttk.Label(metrics_frame, text="Accuracy:").grid(row=0, column=0, sticky="e")
                ttk.Label(metrics_frame, text="Precision:").grid(row=1, column=0, sticky="e")
                ttk.Label(metrics_frame, text="Recall:").grid(row=2, column=0, sticky="e")
                ttk.Label(metrics_frame, text="F1 Score:").grid(row=3, column=0, sticky="e")
                self.acc_var = tk.StringVar()
                self.prec_var = tk.StringVar()
                self.rec_var = tk.StringVar()
                self.f1_var = tk.StringVar()
                ttk.Label(metrics_frame, textvariable=self.acc_var).grid(row=0, column=1, sticky="w")
                ttk.Label(metrics_frame, textvariable=self.prec_var).grid(row=1, column=1, sticky="w")
                ttk.Label(metrics_frame, textvariable=self.rec_var).grid(row=2, column=1, sticky="w")
                ttk.Label(metrics_frame, textvariable=self.f1_var).grid(row=3, column=1, sticky="w")

                def update_metrics(*_):
                    from analysis.confusion_matrix import compute_metrics

                    tp = self.tp_var.get()
                    fp = self.fp_var.get()
                    tn = self.tn_var.get()
                    fn = self.fn_var.get()
                    metrics = compute_metrics(tp, fp, tn, fn)
                    self.acc_var.set(f"{metrics['accuracy']:.3f}")
                    self.prec_var.set(f"{metrics['precision']:.3f}")
                    self.rec_var.set(f"{metrics['recall']:.3f}")
                    self.f1_var.set(f"{metrics['f1']:.3f}")
                    self.p_var.set(tp + fn)
                    self.n_var.set(tn + fp)

                for var in (self.tp_var, self.fp_var, self.tn_var, self.fn_var):
                    var.trace_add("write", update_metrics)
                update_metrics()

                vt_frame = ttk.Frame(cm_frame)
                vt_frame.grid(row=1, column=0, sticky="nsew", pady=5)
                cm_frame.grid_rowconfigure(1, weight=1)
                cm_frame.grid_columnconfigure(0, weight=1)
                columns = [
                    "Product Goal",
                    "Validation Target",
                    "Target Description",
                    "Acceptance Criteria",
                ]
                self.vt_tree = ttk.Treeview(
                    vt_frame, columns=columns, show="headings", height=4
                )
                for c in columns:
                    self.vt_tree.heading(c, text=c)
                    width = 120 if c in ("Product Goal", "Validation Target") else 200
                    self.vt_tree.column(c, width=width, anchor="center")
                self.vt_tree.pack(fill=tk.BOTH, expand=True)
                self.vt_item_to_goal = {}
                self.selected_goal = None
                self.current_tau_on = 0.0

                def on_vt_select(event=None):
                    sel = self.vt_tree.selection()
                    if not sel:
                        self.selected_goal = None
                        self.current_tau_on = 0.0
                    else:
                        sg = self.vt_item_to_goal.get(sel[0])
                        self.selected_goal = sg
                        tau_on = 0.0
                        mp_name = getattr(sg, "mission_profile", "")
                        if mp_name:
                            for mp in self.app.mission_profiles:
                                if mp.name == mp_name:
                                    tau_on = mp.tau_on
                                    break
                        self.current_tau_on = tau_on
                    update_metrics()

                self.vt_tree.bind("<<TreeviewSelect>>", on_vt_select)

                def refresh_vt(*_):
                    self.vt_tree.delete(*self.vt_tree.get_children())
                    self.vt_item_to_goal.clear()
                    name = self.name_var.get().strip()
                    for sg in self.app.get_validation_targets_for_odd(name):
                        iid = self.vt_tree.insert(
                            "",
                            "end",
                            values=[
                                sg.user_name or f"SG {sg.unique_id}",
                                getattr(sg, "validation_target", ""),
                                getattr(sg, "validation_desc", ""),
                                getattr(sg, "acceptance_criteria", ""),
                            ],
                        )
                        self.vt_item_to_goal[iid] = sg
                    items = self.vt_tree.get_children()
                    if items:
                        self.vt_tree.selection_set(items[0])
                        on_vt_select()

                refresh_vt()
                self.name_var.trace_add("write", refresh_vt)

            def apply(self):
                new_data = {"name": self.name_var.get(), "class": self.class_var.get()}
                for row in self.attr_rows:
                    key = row["k_var"].get().strip()
                    if key:
                        new_data[key] = row["v_var"].get()
                tp = float(self.tp_var.get())
                fp = float(self.fp_var.get())
                tn = float(self.tn_var.get())
                fn = float(self.fn_var.get())
                from analysis.confusion_matrix import compute_metrics

                metrics = compute_metrics(tp, fp, tn, fn)
                p = tp + fn
                n = tn + fp
                new_data.update({
                    "tp": tp,
                    "fp": fp,
                    "tn": tn,
                    "fn": fn,
                    "p": p,
                    "n": n,
                })
                new_data.update(metrics)
                self.data = new_data

        def add_lib():
            name = simpledialog.askstring("New Library", "Library name:")
            if not name:
                return
            elems = []
            if messagebox.askyesno("Import", "Import elements from file?"):
                path = filedialog.askopenfilename(filetypes=[("CSV/Excel", "*.csv *.xlsx")])
                if path:
                    if path.lower().endswith(".csv"):
                        with open(path, newline="") as f:
                            elems = list(csv.DictReader(f))
                    elif path.lower().endswith(".xlsx"):
                        try:
                            if load_workbook is None:
                                raise ImportError
                            wb = load_workbook(path, read_only=True)
                            ws = wb.active
                            headers = [c.value for c in next(ws.iter_rows(max_row=1))]
                            for row in ws.iter_rows(min_row=2, values_only=True):
                                elem = {headers[i]: row[i] for i in range(len(headers))}
                                elems.append(elem)
                        except Exception:
                            messagebox.showerror("Import", "Failed to read Excel file. openpyxl required.")
            self.odd_libraries.append({"name": name, "elements": elems})
            refresh_libs()
            self.update_odd_elements()

        def edit_lib():
            sel = lib_lb.curselection()
            if not sel:
                return
            lib = self.odd_libraries[sel[0]]
            name = simpledialog.askstring("Edit Library", "Library name:", initialvalue=lib.get("name", ""))
            if name:
                lib["name"] = name
                refresh_libs()

        def delete_lib():
            sel = lib_lb.curselection()
            if sel:
                idx = sel[0]
                del self.odd_libraries[idx]
                refresh_libs()
                self.update_odd_elements()

        def add_elem():
            sel = lib_lb.curselection()
            if not sel:
                return
            lib = self.odd_libraries[sel[0]]
            dlg = ElementDialog(win, self)
            lib.setdefault("elements", []).append(dlg.data)
            refresh_elems()
            self.update_odd_elements()

        def edit_elem():
            sel_lib = lib_lb.curselection()
            sel_elem = elem_tree.selection()
            if not sel_lib or not sel_elem:
                return
            lib = self.odd_libraries[sel_lib[0]]
            idx = elem_tree.index(sel_elem[0])
            data = lib.get("elements", [])[idx]
            dlg = ElementDialog(win, self, data)
            lib["elements"][idx] = dlg.data
            refresh_elems()
            self.update_odd_elements()

        def del_elem():
            sel_lib = lib_lb.curselection()
            sel_elem = elem_tree.selection()
            if not sel_lib or not sel_elem:
                return
            lib = self.odd_libraries[sel_lib[0]]
            idx = elem_tree.index(sel_elem[0])
            del lib.get("elements", [])[idx]
            refresh_elems()
            self.update_odd_elements()

        btnf = ttk.Frame(win)
        btnf.grid(row=1, column=1, columnspan=3, sticky="ew")
        ttk.Button(btnf, text="Add Lib", command=add_lib).pack(side=tk.LEFT)
        ttk.Button(btnf, text="Edit Lib", command=edit_lib).pack(side=tk.LEFT)
        ttk.Button(btnf, text="Del Lib", command=delete_lib).pack(side=tk.LEFT)
        ttk.Button(btnf, text="Add Elem", command=add_elem).pack(side=tk.LEFT, padx=5)
        ttk.Button(btnf, text="Edit Elem", command=edit_elem).pack(side=tk.LEFT)
        ttk.Button(btnf, text="Del Elem", command=del_elem).pack(side=tk.LEFT)

        lib_lb.bind("<<ListboxSelect>>", refresh_elems)
        refresh_libs()

    def open_reliability_window(self):
        return self.open_windows_features.open_reliability_window()

    def open_fmeda_window(self):
        return self.open_windows_features.open_fmeda_window()

    def open_hazop_window(self):
        return self.open_windows_features.open_hazop_window()

    def open_risk_assessment_window(self):
        return self.open_windows_features.open_risk_assessment_window()

    def open_stpa_window(self):
        return self.open_windows_features.open_stpa_window()

    def open_threat_window(self):
        return self.open_windows_features.open_threat_window()

    def open_fi2tc_window(self):
        return self.open_windows_features.open_fi2tc_window()

    def open_tc2fi_window(self):
        return self.open_windows_features.open_tc2fi_window()

    def open_fault_prioritization_window(self):
        return self.open_windows_features.open_fault_prioritization_window()

    def open_safety_management_toolbox(self, show_diagrams: bool = True):
        return self.open_windows_features.open_safety_management_toolbox(show_diagrams)

    def open_diagram_rules_toolbox(self):
        """Open editor for diagram rule configuration."""
        tab_exists = (
            hasattr(self, "_diagram_rules_tab") and self._diagram_rules_tab.winfo_exists()
        )
        editor_exists = (
            hasattr(self, "diagram_rules_editor")
            and self.diagram_rules_editor.winfo_exists()
        )
        if tab_exists:
            self.doc_nb.select(self._diagram_rules_tab)
            if editor_exists:
                return
            parent = self._diagram_rules_tab
        else:
            parent = self._diagram_rules_tab = self.lifecycle_ui._new_tab("Diagram Rules")

        from gui.diagram_rules_toolbox import DiagramRulesEditor

        self.diagram_rules_editor = DiagramRulesEditor(parent, self, _CONFIG_PATH)
        self.diagram_rules_editor.pack(fill=tk.BOTH, expand=True)

    def open_requirement_patterns_toolbox(self):
        """Open editor for requirement pattern definitions."""
        tab_exists = (
            hasattr(self, "_req_patterns_tab") and self._req_patterns_tab.winfo_exists()
        )
        editor_exists = (
            hasattr(self, "requirement_patterns_editor")
            and self.requirement_patterns_editor.winfo_exists()
        )
        if tab_exists:
            self.doc_nb.select(self._req_patterns_tab)
            if editor_exists:
                return
            parent = self._req_patterns_tab
        else:
            parent = self._req_patterns_tab = self.lifecycle_ui._new_tab("Requirement Patterns")

        from gui.requirement_patterns_toolbox import RequirementPatternsEditor

        self.requirement_patterns_editor = RequirementPatternsEditor(
            parent, self, _PATTERN_PATH
        )
        self.requirement_patterns_editor.pack(fill=tk.BOTH, expand=True)

    def open_report_template_toolbox(self):
        """Open editor for PDF report template configuration."""
        tab_exists = (
            hasattr(self, "_report_template_tab") and self._report_template_tab.winfo_exists()
        )
        editor_exists = (
            hasattr(self, "report_template_editor")
            and self.report_template_editor.winfo_exists()
        )
        if tab_exists:
            self.doc_nb.select(self._report_template_tab)
            if editor_exists:
                return
            parent = self._report_template_tab
        else:
            parent = self._report_template_tab = self.lifecycle_ui._new_tab("Report Template")

        from gui.report_template_toolbox import ReportTemplateEditor

        self.report_template_editor = ReportTemplateEditor(
            parent, self, _REPORT_TEMPLATE_PATH
        )
        self.report_template_editor.pack(fill=tk.BOTH, expand=True)

    def open_report_template_manager(self):
        """Open manager for multiple report templates."""
        tab_exists = (
            hasattr(self, "_report_template_mgr_tab")
            and self._report_template_mgr_tab.winfo_exists()
        )
        manager_exists = (
            hasattr(self, "report_template_manager")
            and self.report_template_manager.winfo_exists()
        )
        if tab_exists:
            self.doc_nb.select(self._report_template_mgr_tab)
            if manager_exists:
                return
            parent = self._report_template_mgr_tab
        else:
            parent = self._report_template_mgr_tab = self.lifecycle_ui._new_tab("Report Templates")

        from gui.report_template_manager import ReportTemplateManager

        self.report_template_manager = ReportTemplateManager(
            parent, self, _REPORT_TEMPLATE_PATH.parent
        )
        self.report_template_manager.pack(fill=tk.BOTH, expand=True)

    def reload_config(self) -> None:
        """Reload diagram rule configuration across all interested modules."""

        import sys

        _reload_local_config()

        for mod in list(sys.modules.values()):
            if hasattr(mod, "reload_config"):
                try:  # pragma: no cover - defensive programming
                    mod.reload_config()
                except Exception:
                    pass

        if hasattr(self, "canvas") and getattr(self.canvas, "winfo_exists", lambda: False)():
            self.redraw_canvas()
        pd = getattr(self, "page_diagram", None)
        if pd and getattr(pd.canvas, "winfo_exists", lambda: False)():
            pd.redraw_canvas()

    def open_search_toolbox(self):
        return self.nav_input.open_search_toolbox()

    def open_style_editor(self):
        """Open the diagram style editor window."""
        StyleEditor(self.root)



    def refresh_styles(self, event=None):
        """Redraw all open diagram windows using current styles."""
        if getattr(self, 'canvas', None):
            self.canvas.config(bg=StyleManager.get_instance().canvas_bg)
        for tab in getattr(self, 'diagram_tabs', {}).values():
            for child in tab.winfo_children():
                if hasattr(child, 'redraw'):
                    child.redraw()

    def show_hazard_explorer(self):
        return self.safety_analysis.show_hazard_explorer()

    def show_requirements_explorer(self):
        if hasattr(self, "_req_exp_tab") and self._req_exp_tab.winfo_exists():
            self.doc_nb.select(self._req_exp_tab)
        else:
            self._req_exp_tab = self.lifecycle_ui._new_tab("Requirements Explorer")
            self._req_exp_window = RequirementsExplorerWindow(self._req_exp_tab, self)
            self._req_exp_window.pack(fill=tk.BOTH, expand=True)


    def _create_fta_tab(self, diagram_mode: str = "FTA"):
        """Create the main FTA tab with canvas and bindings.

        Parameters
        ----------
        diagram_mode: str
            The operational mode of the diagram (``"FTA"`` or ``"CTA"``).
        """
        tabs = getattr(self, "analysis_tabs", {})
        existing = tabs.get(diagram_mode)
        
        if existing and existing["tab"].winfo_exists():
            self.canvas_tab = existing["tab"]
            self.canvas_frame = existing["tab"]
            self.canvas = existing["canvas"]
            self.hbar = existing["hbar"]
            self.vbar = existing["vbar"]
            self.diagram_mode = diagram_mode
            self.doc_nb.select(self.canvas_tab)
            self._update_analysis_menus(diagram_mode)
            return

        canvas_tab = ttk.Frame(self.doc_nb)
        self.doc_nb.add(canvas_tab, text="FTA" if diagram_mode == "FTA" else diagram_mode)

        canvas = tk.Canvas(canvas_tab, bg=StyleManager.get_instance().canvas_bg)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        hbar = ttk.Scrollbar(canvas_tab, orient=tk.HORIZONTAL, command=canvas.xview)
        hbar.pack(side=tk.BOTTOM, fill=tk.X)
        vbar = ttk.Scrollbar(canvas_tab, orient=tk.VERTICAL, command=canvas.yview)
        vbar.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.config(xscrollcommand=hbar.set, yscrollcommand=vbar.set,
                      scrollregion=(0, 0, 2000, 2000))
        canvas.bind("<ButtonPress-3>", self.on_right_mouse_press)
        canvas.bind("<B3-Motion>", self.on_right_mouse_drag)
        canvas.bind("<ButtonRelease-3>", self.on_right_mouse_release)
        canvas.bind("<Button-1>", self.on_canvas_click)
        canvas.bind("<B1-Motion>", self.on_canvas_drag)
        canvas.bind("<ButtonRelease-1>", self.on_canvas_release)
        canvas.bind("<Double-1>", self.on_canvas_double_click)
        canvas.bind("<Control-MouseWheel>", self.on_ctrl_mousewheel)

        canvas.diagram_mode = diagram_mode
        self.analysis_tabs[diagram_mode] = {
            "tab": canvas_tab,
            "canvas": canvas,
            "hbar": hbar,
            "vbar": vbar,
        }
        self.canvas_tab = canvas_tab
        self.canvas_frame = canvas_tab
        self.canvas = canvas
        self.hbar = hbar
        self.vbar = vbar
        self.diagram_mode = diagram_mode
        self.doc_nb.select(canvas_tab)
        self._update_analysis_menus(diagram_mode)

    def create_fta_diagram(self):
        """Initialize an FTA diagram and its top-level event."""
        self._create_fta_tab("FTA")
        self.add_top_level_event()
        if getattr(self, "fta_root_node", None):
            self.window_controllers.open_page_diagram(self.fta_root_node)

    def create_cta_diagram(self):
        """Initialize a CTA diagram and its top-level event."""
        self.cta_manager.create_diagram()

    def enable_fta_actions(self, enabled: bool) -> None:
        """Enable or disable FTA-related menu actions."""
        mode = getattr(self, "diagram_mode", "FTA")
        if hasattr(self, "fta_menu"):
            state = tk.NORMAL if enabled else tk.DISABLED
            for key in (
                "add_gate",
                "add_basic_event",
                "add_gate_from_failure_mode",
                "add_fault_event",
            ):
                self.fta_menu.entryconfig(self._fta_menu_indices[key], state=state)
                
    def enable_paa_actions(self, enabled: bool) -> None:
        """Delegate to :class:`Pages_and_PAA` to toggle PAA actions."""
        self.pages_and_paa.enable_paa_actions(enabled)
                
    def _update_analysis_menus(self,mode=None):
        """Enable or disable node-adding menu items based on diagram mode."""
        if mode is None:
            mode = getattr(self, "diagram_mode", "FTA")
        self.enable_fta_actions(mode == "FTA")
        self.cta_manager.enable_actions(mode == "CTA")
        self.validation_consistency.enable_paa_actions(mode == "PAA")

    def _create_paa_tab(self) -> None:
        """Delegate PAA tab creation to helper."""
        self.pages_and_paa._create_paa_tab()

    def create_paa_diagram(self) -> None:
        """Delegate PAA diagram setup to helper."""
        self.pages_and_paa.create_paa_diagram()

    @property
    def paa_manager(self) -> PrototypeAssuranceManager:
        """Lazily create and return the PAA manager."""
        if not hasattr(self, "_paa_manager"):
            self._paa_manager = PrototypeAssuranceManager(self)
        return self._paa_manager

    @property
    def pages_and_paa(self):
        """Lazily create and return the Pages_and_PAA helper."""
        if not hasattr(self, "_pages_and_paa"):
            from .pages_and_paa import Pages_and_PAA

            self._pages_and_paa = Pages_and_PAA(self)
        return self._pages_and_paa

    def _reset_fta_state(self):
        """Clear references to the FTA tab and its canvas."""
        self.canvas_tab = None
        self.canvas_frame = None
        self.canvas = None
        self.hbar = None
        self.vbar = None
        self.page_diagram = None

    def ensure_fta_tab(self):  # pragma: no cover - delegation
        return self.validation_consistency.ensure_fta_tab()

    def _format_diag_title(self, diag) -> str:
        """Return SysML style title for a diagram tab."""
        if diag.name:
            return f"\N{LEFT-POINTING DOUBLE ANGLE QUOTATION MARK}{diag.diag_type}\N{RIGHT-POINTING DOUBLE ANGLE QUOTATION MARK} {diag.name}"
        return f"\N{LEFT-POINTING DOUBLE ANGLE QUOTATION MARK}{diag.diag_type}\N{RIGHT-POINTING DOUBLE ANGLE QUOTATION MARK}"

    def open_use_case_diagram(self):
        self.window_controllers.open_use_case_diagram()

    def open_activity_diagram(self):
        self.window_controllers.open_activity_diagram()

    def open_block_diagram(self):
        self.window_controllers.open_block_diagram()

    def open_internal_block_diagram(self):
        self.window_controllers.open_internal_block_diagram()

    def open_control_flow_diagram(self):
        self.window_controllers.open_control_flow_diagram()

    def open_causal_bayesian_network_window(self):
        self.window_controllers.open_causal_bayesian_network_window()

    def open_gsn_diagram(self, diagram):
        self.window_controllers.open_gsn_diagram(diagram)

    def open_arch_window(self, diag_id: str) -> None:
        self.window_controllers.open_arch_window(diag_id)

    def open_page_diagram(self, node, push_history=True):
        self.window_controllers.open_page_diagram(node, push_history)

    def manage_architecture(self):
        return self.open_windows_features.manage_architecture()

    def manage_gsn(self):
        self.gsn_manager.manage_gsn()

    def manage_safety_management(self):
        return self.open_windows_features.manage_safety_management()

    def manage_safety_cases(self):
        if not hasattr(self, "safety_case_library"):
            from analysis import SafetyCaseLibrary as _SCL

            self.safety_case_library = _SCL()
        if hasattr(self, "_safety_case_exp_tab") and self._safety_case_exp_tab.winfo_exists():
            self.doc_nb.select(self._safety_case_exp_tab)
        else:
            self._safety_case_exp_tab = self.lifecycle_ui._new_tab("Safety & Security Case Explorer")
            self._safety_case_window = SafetyCaseExplorer(
                self._safety_case_exp_tab, self, self.safety_case_library
            )
            self._safety_case_window.pack(fill=tk.BOTH, expand=True)
        self.refresh_all()


    def _diagram_copy_strategy1(self):
        win = self.window_controllers._focused_cbn_window()
        if win and getattr(win, "selected_node", None) and getattr(win, "copy_selected", None):
            self.selected_node = None
            self.clipboard_node = None
            self.cut_mode = False
            win.copy_selected()
            return True
        return False

    def _diagram_copy_strategy2(self):
        win = self.window_controllers._focused_gsn_window()
        if win and getattr(win, "selected_node", None) and getattr(win, "copy_selected", None):
            self.selected_node = None
            self.clipboard_node = None
            self.cut_mode = False
            win.copy_selected()
            return True
        return False

    def _diagram_copy_strategy3(self):
        win = getattr(self, "active_arch_window", None)
        if win and getattr(win, "selected_obj", None) and getattr(win, "copy_selected", None):
            self.selected_node = None
            self.clipboard_node = None
            self.cut_mode = False
            win.copy_selected()
            return True
        return False

    def _diagram_copy_strategy4(self):
        for ref in list(ARCH_WINDOWS):
            win = ref()
            if win and getattr(win, "selected_obj", None) and getattr(win, "copy_selected", None):
                self.selected_node = None
                self.clipboard_node = None
                self.cut_mode = False
                win.copy_selected()
                return True
        return False

    def _diagram_cut_strategy1(self):
        win = self.window_controllers._focused_cbn_window()
        if win and getattr(win, "selected_node", None) and getattr(win, "cut_selected", None):
            self.selected_node = None
            self.clipboard_node = None
            self.cut_mode = False
            win.cut_selected()
            return True
        return False

    def _diagram_cut_strategy2(self):
        win = self.window_controllers._focused_gsn_window()
        if win and getattr(win, "selected_node", None) and getattr(win, "cut_selected", None):
            self.selected_node = None
            self.clipboard_node = None
            self.cut_mode = False
            win.cut_selected()
            return True
        return False

    def _diagram_cut_strategy3(self):
        win = getattr(self, "active_arch_window", None)
        if win and getattr(win, "selected_obj", None) and getattr(win, "cut_selected", None):
            self.selected_node = None
            self.clipboard_node = None
            self.cut_mode = False
            win.cut_selected()
            return True
        return False

    def _diagram_cut_strategy4(self):
        for ref in list(ARCH_WINDOWS):
            win = ref()
            if win and getattr(win, "selected_obj", None) and getattr(win, "cut_selected", None):
                self.selected_node = None
                self.clipboard_node = None
                self.cut_mode = False
                win.cut_selected()
                return True
        return False

    def copy_node(self):
        self.diagram_clipboard.copy_node()

    def cut_node(self):
        self.diagram_clipboard.cut_node()

    # ------------------------------------------------------------------
    def _reset_gsn_clone(self, node):
        if isinstance(node, GSNNode):
            node.unique_id = str(uuid.uuid4())
            old_children = list(getattr(node, "children", []))
            node.children = []
            node.parents = []
            node.context_children = []
            node.is_primary_instance = False
            if getattr(node, "original", None) is None:
                node.original = node
            for child in old_children:
                self._reset_gsn_clone(child)

    # ------------------------------------------------------------------
    def _clone_for_paste_strategy1(self, node, parent=None):
        if hasattr(node, "clone"):
            if parent and getattr(node, "node_type", None) in {"Context", "Assumption", "Justification"}:
                return node.clone(parent)
            return node.clone()
        import copy
        clone = copy.deepcopy(node)
        self._reset_gsn_clone(clone)
        return clone

    def _clone_for_paste_strategy2(self, node, parent=None):
        import copy
        if isinstance(node, GSNNode):
            if parent and node.node_type in {"Context", "Assumption", "Justification"}:
                return node.clone(parent)
            return node.clone()
        clone = copy.deepcopy(node)
        self._reset_gsn_clone(clone)
        return clone

    def _clone_for_paste_strategy3(self, node, parent=None):
        try:
            if parent and getattr(node, "node_type", None) in {"Context", "Assumption", "Justification"}:
                return node.clone(parent)  # type: ignore[attr-defined]
            return node.clone()  # type: ignore[attr-defined]
        except Exception:
            import copy
            clone = copy.deepcopy(node)
            self._reset_gsn_clone(clone)
            return clone

    def _clone_for_paste_strategy4(self, node, parent=None):
        import copy
        clone = copy.deepcopy(node)
        self._reset_gsn_clone(clone)
        return clone

    def _clone_for_paste(self, node, parent=None):
        for strat in (
            self._clone_for_paste_strategy1,
            self._clone_for_paste_strategy2,
            self._clone_for_paste_strategy3,
            self._clone_for_paste_strategy4,
        ):
            try:
                clone = strat(node, parent)
                if clone is not None:
                    return clone
            except ValueError:
                messagebox.showwarning("Clone", "Cannot clone this node type.")
                return None
            except Exception:
                continue
        messagebox.showwarning("Clone", "Cannot clone this node type.")
        return None

    def _prepare_node_for_paste(self, target):
        """Return appropriate node instance when pasting."""
        if (
            isinstance(self.clipboard_node, GSNNode)
            and target in getattr(self.clipboard_node, "parents", [])
        ):
            return self._clone_for_paste(self.clipboard_node)
        from mainappsrc.models.fta.fault_tree_node import FaultTreeNode

        if (
            isinstance(self.clipboard_node, FaultTreeNode)
            or type(self.clipboard_node).__name__ == "FaultTreeNode"
        ):
            return self._clone_for_paste(self.clipboard_node)
        return self.clipboard_node

    def paste_node(self):
        self.diagram_clipboard.paste_node()

    def _get_diag_type(self, win):
        repo = getattr(win, "repo", None)
        diag_id = getattr(win, "diagram_id", None)
        if repo and diag_id:
            diag = repo.diagrams.get(diag_id)
            if diag:
                return diag.diag_type
        return None


    def clone_node_preserving_id(self, node, parent=None):
        """Delegate cloning to :class:`NodeCloneService`."""
        return self.node_clone_service.clone_node_preserving_id(node, parent)

    def _find_gsn_diagram(self, node):
        """Return the GSN diagram containing ``node`` if known.

        The application keeps GSN diagrams either in ``gsn_diagrams`` or nested
        inside modules.  When pasting a GSN node we must ensure the element is
        registered with its diagram's ``nodes`` list so it is rendered
        correctly.  This helper performs a recursive search through all known
        diagrams and modules to locate the owning diagram of ``node``.
        """

        for diag in getattr(self, "gsn_diagrams", []):
            if node in getattr(diag, "nodes", []):
                return diag

        def _search_modules(modules):
            for mod in modules:
                for diag in getattr(mod, "diagrams", []):
                    if node in getattr(diag, "nodes", []):
                        return diag
                result = _search_modules(getattr(mod, "modules", []))
                if result:
                    return result
            return None

        return _search_modules(getattr(self, "gsn_modules", []))

    def _copy_attrs_no_xy_strategy1(self, target, source, attrs):
        tx, ty = getattr(target, "x", None), getattr(target, "y", None)
        for attr in attrs:
            setattr(target, attr, getattr(source, attr, None))
        if tx is not None:
            target.x = tx
        if ty is not None:
            target.y = ty

    def _copy_attrs_no_xy_strategy2(self, target, source, attrs):
        tx, ty = getattr(target, "x", None), getattr(target, "y", None)
        values = {a: getattr(source, a, None) for a in attrs if hasattr(source, a)}
        for a, v in values.items():
            setattr(target, a, v)
        if tx is not None:
            target.x = tx
        if ty is not None:
            target.y = ty

    def _copy_attrs_no_xy_strategy3(self, target, source, attrs):
        tx, ty = getattr(target, "x", None), getattr(target, "y", None)
        for attr in attrs:
            if attr in {"x", "y"}:
                continue
            setattr(target, attr, getattr(source, attr, None))
        if tx is not None:
            target.x = tx
        if ty is not None:
            target.y = ty

    def _copy_attrs_no_xy_strategy4(self, target, source, attrs):
        tx, ty = getattr(target, "x", None), getattr(target, "y", None)
        for attr in attrs:
            try:
                setattr(target, attr, getattr(source, attr))
            except Exception:
                continue
        if tx is not None:
            target.x = tx
        if ty is not None:
            target.y = ty

    def _copy_attrs_no_xy(self, target, source, attrs):
        for strat in (
            self._copy_attrs_no_xy_strategy1,
            self._copy_attrs_no_xy_strategy2,
            self._copy_attrs_no_xy_strategy3,
            self._copy_attrs_no_xy_strategy4,
        ):
            try:
                strat(target, source, attrs)
                return
            except Exception:
                continue

    def sync_nodes_by_id(self, updated_node):  # pragma: no cover - simple delegation
        """Delegate to :class:`Syncing_And_IDs` helper."""

        return self.syncing_and_ids.sync_nodes_by_id(updated_node)

       




    def edit_severity(self):
        messagebox.showinfo(
            "Severity",
            "Severity is determined from the risk assessment and cannot be edited here.",
        )



    def set_last_saved_state(self):
        """Record the current model state for change detection."""
        self.last_saved_state = json.dumps(self.export_model_data(), sort_keys=True)

    def has_unsaved_changes(self):
        """Return True if the model differs from the last saved state."""
        current_state = json.dumps(self.export_model_data(), sort_keys=True)
        return current_state != getattr(self, "last_saved_state", None)

    # ------------------------------------------------------------
    # ------------------------------------------------------------
    # Undo support
    # ------------------------------------------------------------
    def push_undo_state(self, strategy: str = "v4", sync_repo: bool = True) -> None:
        self.undo_manager.push_undo_state(strategy=strategy, sync_repo=sync_repo)

    def _push_undo_state_v1(self, state: dict, stripped: dict) -> bool:
        return self.undo_manager._push_undo_state_v1(state, stripped)

    def _push_undo_state_v2(self, state: dict, stripped: dict) -> bool:
        return self.undo_manager._push_undo_state_v2(state, stripped)

    def _push_undo_state_v3(self, state: dict, stripped: dict) -> bool:
        return self.undo_manager._push_undo_state_v3(state, stripped)

    def _push_undo_state_v4(self, state: dict, stripped: dict) -> bool:
        return self.undo_manager._push_undo_state_v4(state, stripped)

    def _undo_hotkey(self, event):
        """Keyboard shortcut handler for undo."""
        self.undo_manager.undo()
        return "break"

    def _redo_hotkey(self, event):
        """Keyboard shortcut handler for redo."""
        self.undo_manager.redo()
        return "break"

    def undo(self, strategy: str = "v4"):
        self.undo_manager.undo(strategy=strategy)

    def redo(self, strategy: str = "v4"):
        self.undo_manager.redo(strategy=strategy)

    def clear_undo_history(self) -> None:
        """Remove all undo and redo history."""
        self.undo_manager.clear_history()
    def confirm_close(self):
        """Prompt to save if there are unsaved changes before closing."""
        if self.has_unsaved_changes():
            result = messagebox.askyesnocancel("Unsaved Changes", "Save changes before exiting?")
            if result is None:
                return
            if result:
                self.save_model()
        # Previously, any loaded model paths were deleted on close, which could
        # remove user data. Avoid deleting files that were explicitly opened by
        # the user so their project files remain intact.
        # Ensure the Tk event loop terminates and all windows are destroyed
        self.root.quit()
        self.root.destroy()


    def export_model_data(self, include_versions=True):
        return self.reporting_export.export_model_data(include_versions)

    def _load_project_properties(self, data: dict) -> None:
        """Delegate project property loading to the manager."""
        self.project_properties = self.project_properties_manager.load_project_properties(
            data
        )

    def _load_fault_tree_events(self, data: dict, ensure_root: bool) -> None:
        return self.safety_analysis._load_fault_tree_events(data, ensure_root)
          
    def apply_model_data(self, data: dict, ensure_root: bool = True):
        """Load model state from a dictionary."""

        # Clear any previously enabled work products so new governance can be
        # applied deterministically.  Minimal test instances may not define the
        # ``enabled_work_products`` attribute so ``getattr`` is used with a
        # default.
        current = list(getattr(self, "enabled_work_products", set()))
        for name in current:
            try:
                self.validation_consistency.disable_work_product(name)
            except Exception:
                pass
        self.enabled_work_products = set()
        self._load_project_properties(data)

        repo_data = data.get("sysml_repository")
        if repo_data:
            repo = SysMLRepository.get_instance()
            repo.from_dict(repo_data)

        self._load_fault_tree_events(data, ensure_root)

        global global_requirements
        global_requirements.clear()
        for rid, req in data.get("global_requirements", {}).items():
            global_requirements[rid] = ensure_requirement_defaults(req)

        self.gsn_modules = [
            GSNModule.from_dict(m) for m in data.get("gsn_modules", [])
        ]
        self.gsn_diagrams = [
            GSNDiagram.from_dict(d) for d in data.get("gsn_diagrams", [])
        ]

        self.safety_mgmt_toolbox = SafetyManagementToolbox.from_dict(
            data.get("safety_mgmt_toolbox", {})
        )
        toolbox = self.safety_mgmt_toolbox
        self.governance_manager.attach_toolbox(toolbox)
        # Refresh menus to expose phases from the loaded toolbox
        self._refresh_phase_requirements_menu()
        # Ensure the SysML repository knows about the active phase from the
        # loaded toolbox so diagrams and work products filter correctly.
        self.governance_manager.set_active_module(toolbox.active_module)
        for te in self.top_events:
            toolbox.register_loaded_work_product("FTA", te.user_name)
        for te in getattr(self, "cta_events", []):
            toolbox.register_loaded_work_product("CTA", te.user_name)
        for te in getattr(self, "paa_events", []):
            toolbox.register_loaded_work_product("Prototype Assurance Analysis", te.user_name)

        for name in data.get("enabled_work_products", []):
            try:
                self.validation_consistency.enable_work_product(name)
            except Exception:
                self.enabled_work_products.add(name)

        self.fmea_service.load_fmeas(data)

        self.fmedas = []
        for doc in data.get("fmedas", []):
            entries = [FaultTreeNode.from_dict(e) for e in doc.get("entries", [])]
            self.fmedas.append({
                "name": doc.get("name", "FMEDA"),
                "file": doc.get("file", f"fmeda_{len(self.fmedas)}.csv"),
                "entries": entries,
                "bom": doc.get("bom", ""),
                "created": doc.get("created", datetime.datetime.now().isoformat()),
                "author": doc.get("author", CURRENT_USER_NAME),
                "modified": doc.get("modified", datetime.datetime.now().isoformat()),
                "modified_by": doc.get("modified_by", CURRENT_USER_NAME),
            })

        self.update_failure_list()

        node_map = {}
        for te in self.top_events:
            for n in self.get_all_nodes(te):
                node_map[n.unique_id] = n
        for entry in self.get_all_fmea_entries():
            orig = node_map.get(entry.unique_id)
            if orig and entry is not orig:
                entry.is_primary_instance = False
                entry.original = orig

        self.mechanism_libraries = []
        for lib in data.get("mechanism_libraries", []):
            mechs = [DiagnosticMechanism(**m) for m in lib.get("mechanisms", [])]
            self.mechanism_libraries.append(MechanismLibrary(lib.get("name", ""), mechs))
        self.selected_mechanism_libraries = []
        for name in data.get("selected_mechanism_libraries", []):
            found = next((l for l in self.mechanism_libraries if l.name == name), None)
            if found:
                self.selected_mechanism_libraries.append(found)
        if not self.mechanism_libraries:
            self.load_default_mechanisms()

        self.mission_profiles = []
        for mp_data in data.get("mission_profiles", []):
            try:
                mp = MissionProfile(**mp_data)
                total = mp.tau_on + mp.tau_off
                mp.duty_cycle = mp.tau_on / total if total else 0.0
                self.mission_profiles.append(mp)
            except TypeError:
                pass

        self.reliability_analyses = []
        for ra in data.get("reliability_analyses", []):
            def load_comp(cdata):
                comp = ReliabilityComponent(
                    cdata.get("name", ""),
                    cdata.get("comp_type", ""),
                    cdata.get("quantity", 1),
                    cdata.get("attributes", {}),
                    cdata.get("qualification", cdata.get("safety_req", "")),
                    cdata.get("fit", 0.0),
                    cdata.get("is_passive", False),
                )
                comp.sub_boms = [
                    [load_comp(sc) for sc in bom]
                    for bom in cdata.get("sub_boms", [])
                ]
                return comp

            comps = [load_comp(c) for c in ra.get("components", [])]
            self.reliability_analyses.append(
                ReliabilityAnalysis(
                    ra.get("name", ""),
                    ra.get("standard", ""),
                    ra.get("profile", ""),
                    comps,
                    ra.get("total_fit", 0.0),
                    ra.get("spfm", 0.0),
                    ra.get("lpfm", 0.0),
                    ra.get("dc", 0.0),
                )
            )

        self.hazop_docs = []
        for d in data.get("hazops", []):
            entries = []
            for h in d.get("entries", []):
                h["safety"] = boolify(h.get("safety", False), False)
                h["covered"] = boolify(h.get("covered", False), False)
                entries.append(HazopEntry(**h))
            doc = HazopDoc(d.get("name", f"HAZOP {len(self.hazop_docs)+1}"), entries)
            self.hazop_docs.append(doc)
            toolbox.register_loaded_work_product("HAZOP", doc.name)
        if not self.hazop_docs and data.get("hazop_entries"):
            entries = []
            for h in hazop_entries:
                h["safety"] = boolify(h.get("safety", False), False)
                h["covered"] = boolify(h.get("covered", False), False)
                entries.append(HazopEntry(**h))
            doc = HazopDoc("Default", entries)
            self.hazop_docs.append(doc)
            toolbox.register_loaded_work_product("HAZOP", doc.name)
        self.active_hazop = self.hazop_docs[0] if self.hazop_docs else None
        self.hazop_entries = self.active_hazop.entries if self.active_hazop else []

        self.hara_docs = []
        for d in data.get("haras", []):
            entries = []
            for e in d.get("entries", []):
                cdata = e.get("cyber")
                cyber = self.cyber_manager.build_risk_entry(cdata)
                entries.append(
                    HaraEntry(
                        e.get("malfunction", ""),
                        e.get("hazard", ""),
                        e.get("scenario", ""),
                        e.get("severity", 1),
                        e.get("sev_rationale", ""),
                        e.get("controllability", 1),
                        e.get("cont_rationale", ""),
                        e.get("exposure", 1),
                        e.get("exp_rationale", ""),
                        e.get("asil", "QM"),
                        e.get("safety_goal", ""),
                        cyber,
                    )
                )
            hazops = d.get("hazops")
            if not hazops:
                hazop = d.get("hazop")
                hazops = [hazop] if hazop else []
                doc = HaraDoc(
                    d.get("name", f"Risk Assessment {len(self.hara_docs)+1}"),
                    hazops,
                    entries,
                    d.get("approved", False),
                    d.get("status", "draft"),
                    stpa=d.get("stpa", ""),
                    threat=d.get("threat", ""),
                    fi2tc=d.get("fi2tc", ""),
                    tc2fi=d.get("tc2fi", ""),
                )
            self.hara_docs.append(doc)
            toolbox.register_loaded_work_product("Risk Assessment", doc.name)
        if not self.hara_docs and data.get("hara_entries"):
            hazop_name = self.hazop_docs[0].name if self.hazop_docs else ""
            entries = []
            for e in data.get("hara_entries", []):
                cdata = e.get("cyber")
                cyber = self.cyber_manager.build_risk_entry(cdata)
                entries.append(
                    HaraEntry(
                        e.get("malfunction", ""),
                        e.get("hazard", ""),
                        e.get("scenario", ""),
                        e.get("severity", 1),
                        e.get("sev_rationale", ""),
                        e.get("controllability", 1),
                        e.get("cont_rationale", ""),
                        e.get("exposure", 1),
                        e.get("exp_rationale", ""),
                        e.get("asil", "QM"),
                        e.get("safety_goal", ""),
                        cyber,
                    )
                )
            doc = HaraDoc(
                "Default",
                [hazop_name] if hazop_name else [],
                entries,
                False,
                "draft",
                stpa="",
                threat="",
                fi2tc="",
                tc2fi="",
            )
            self.hara_docs.append(doc)
            toolbox.register_loaded_work_product("Risk Assessment", doc.name)
        self.active_hara = self.hara_docs[0] if self.hara_docs else None
        self.hara_entries = self.active_hara.entries if self.active_hara else []
        self.update_hazard_list()

        self.stpa_docs = []
        for d in data.get("stpas", []):
            entries = [
                StpaEntry(
                    e.get("action", ""),
                    e.get("not_providing", ""),
                    e.get("providing", ""),
                    e.get("incorrect_timing", ""),
                    e.get("stopped_too_soon", ""),
                    e.get("safety_constraints", []),
                )
                for e in d.get("entries", [])
            ]
            doc = StpaDoc(
                d.get("name", f"STPA {len(self.stpa_docs)+1}"),
                d.get("diagram", ""),
                entries,
            )
            self.stpa_docs.append(doc)
            toolbox.register_loaded_work_product("STPA", doc.name)
        if not self.stpa_docs and data.get("stpa_entries"):
            entries = [
                StpaEntry(
                    e.get("action", ""),
                    e.get("not_providing", ""),
                    e.get("providing", ""),
                    e.get("incorrect_timing", ""),
                    e.get("stopped_too_soon", ""),
                    e.get("safety_constraints", []),
                )
                for e in data.get("stpa_entries", [])
            ]
            doc = StpaDoc("Default", "", entries)
            self.stpa_docs.append(doc)
            toolbox.register_loaded_work_product("STPA", doc.name)
        self.active_stpa = self.stpa_docs[0] if self.stpa_docs else None
        self.stpa_entries = self.active_stpa.entries if self.active_stpa else []

        self.threat_docs = []
        for d in data.get("threat_docs", []):
            entries = []
            for e in d.get("entries", []):
                funcs = []
                raw_funcs = e.get("functions", [])
                if raw_funcs and isinstance(raw_funcs[0], dict):
                    for f in raw_funcs:
                        dmg_list = []
                        for ds in f.get("damage_scenarios", []):
                            threats = []
                            for t in ds.get("threats", []):
                                paths = [AttackPath(**p) for p in t.get("attack_paths", [])]
                                threats.append(
                                    ThreatScenario(
                                        t.get("stride", ""),
                                        t.get("scenario", ""),
                                        paths,
                                    )
                                )
                            dmg_list.append(
                                DamageScenario(
                                    ds.get("scenario", ""), ds.get("dtype", ""), threats
                                )
                            )
                        funcs.append(FunctionThreat(f.get("name", ""), dmg_list))
                else:
                    dmg_list = []
                    for ds in e.get("damage_scenarios", []):
                        threats = []
                        for t in ds.get("threats", []):
                            paths = [AttackPath(**p) for p in t.get("attack_paths", [])]
                            threats.append(
                                ThreatScenario(
                                    t.get("stride", ""),
                                    t.get("scenario", ""),
                                    paths,
                                )
                            )
                        dmg_list.append(
                            DamageScenario(ds.get("scenario", ""), ds.get("dtype", ""), threats)
                        )
                    func_names = raw_funcs
                    if func_names is None:
                        func = e.get("function")
                        func_names = [func] if func else []
                    for name in func_names:
                        funcs.append(FunctionThreat(name, dmg_list))
                entries.append(ThreatEntry(e.get("asset", ""), funcs))
            doc = ThreatDoc(
                d.get("name", f"Threat {len(self.threat_docs)+1}"),
                d.get("diagram", ""),
                entries,
            )
            self.threat_docs.append(doc)
            toolbox.register_loaded_work_product("Threat Analysis", doc.name)
        self.active_threat = self.threat_docs[0] if self.threat_docs else None
        self.threat_entries = self.active_threat.entries if self.active_threat else []

        self.fi2tc_docs = []
        for d in data.get("fi2tc_docs", []):
            doc = FI2TCDoc(
                d.get("name", f"FI2TC {len(self.fi2tc_docs)+1}"),
                d.get("entries", []),
            )
            self.fi2tc_docs.append(doc)
            toolbox.register_loaded_work_product("FI2TC", doc.name)
        if not self.fi2tc_docs and data.get("fi2tc_entries"):
            doc = FI2TCDoc("Default", data.get("fi2tc_entries", []))
            self.fi2tc_docs.append(doc)
            toolbox.register_loaded_work_product("FI2TC", doc.name)
        self.active_fi2tc = self.fi2tc_docs[0] if self.fi2tc_docs else None
        self.fi2tc_entries = self.active_fi2tc.entries if self.active_fi2tc else []

        self.tc2fi_docs = []
        for d in data.get("tc2fi_docs", []):
            doc = TC2FIDoc(
                d.get("name", f"TC2FI {len(self.tc2fi_docs)+1}"),
                d.get("entries", []),
            )
            self.tc2fi_docs.append(doc)
            toolbox.register_loaded_work_product("TC2FI", doc.name)
        if not self.tc2fi_docs and data.get("tc2fi_entries"):
            doc = TC2FIDoc("Default", data.get("tc2fi_entries", []))
            self.tc2fi_docs.append(doc)
            toolbox.register_loaded_work_product("TC2FI", doc.name)
        self.active_tc2fi = self.tc2fi_docs[0] if self.tc2fi_docs else None
        self.tc2fi_entries = self.active_tc2fi.entries if self.active_tc2fi else []

        self.cbn_docs = []
        for d in data.get("cbn_docs", []):
            net = CausalBayesianNetwork()
            net.nodes = d.get("nodes", [])
            net.parents = {k: list(v) for k, v in d.get("parents", {}).items()}
            raw_cpds = d.get("cpds", {})
            parsed_cpds = {}
            for var, cpd in raw_cpds.items():
                if isinstance(cpd, Mapping):
                    parsed_cpds[var] = {
                        tuple(ch == "1" for ch in key): val
                        for key, val in cpd.items()
                    }
                else:
                    parsed_cpds[var] = cpd
            net.cpds = parsed_cpds
            name = d.get("name", f"CBN {len(self.cbn_docs)+1}")
            positions = {k: tuple(v) for k, v in d.get("positions", {}).items()}
            types = {k: v for k, v in d.get("types", {}).items()}
            doc = CausalBayesianNetworkDoc(name, network=net, positions=positions, types=types)
            self.cbn_docs.append(doc)
            toolbox.register_loaded_work_product("Causal Bayesian Network Analysis", name)
        self.active_cbn = self.cbn_docs[0] if self.cbn_docs else None

        self.scenario_libraries = data.get("scenario_libraries", [])
        self.odd_libraries = data.get("odd_libraries", [])
        self.faults = data.get("faults", [])
        for be in self.get_all_basic_events():
            desc = be.description.strip()
            if desc and desc not in self.faults:
                self.faults.append(desc)
        mals = []
        for m in data.get("malfunctions", []):
            append_unique_insensitive(mals, m)
        self.malfunctions = mals
        self.hazards = data.get("hazards", [])
        self.failures = data.get("failures", [])
        if not self.odd_libraries and "odd_elements" in data:
            self.odd_libraries = [{"name": "Default", "elements": data.get("odd_elements", [])}]
        self.update_odd_elements()

        self.fmedas = []
        for doc in data.get("fmedas", []):
            entries = [FaultTreeNode.from_dict(e) for e in doc.get("entries", [])]
            self.fmedas.append({
                "name": doc.get("name", "FMEDA"),
                "file": doc.get("file", f"fmeda_{len(self.fmedas)}.csv"),
                "entries": entries,
                "bom": doc.get("bom", ""),
            })

        for event in self.top_events:
            AutoML_Helper.fix_clone_references(self.top_events)
        AutoML_Helper.update_unique_id_counter_for_top_events(self.top_events)
        for event in self.top_events:
            self.update_global_requirements_from_nodes(event)
        if hasattr(self, "hara_entries"):
            self.sync_hara_to_safety_goals()
        self.item_definition = data.get(
            "item_definition",
            getattr(self, "item_definition", {"description": "", "assumptions": ""}),
        )
        self.safety_concept = data.get(
            "safety_concept",
            getattr(
                self,
                "safety_concept",
                {"functional": "", "technical": "", "cybersecurity": ""},
            ),
        )
        self.reviews = []
        reviews_data = data.get("reviews")
        if reviews_data:
            for rd in reviews_data:
                participants = [ReviewParticipant(**p) for p in rd.get("participants", [])]
                comments = [ReviewComment(**c) for c in rd.get("comments", [])]
                moderators = [ReviewParticipant(**m) for m in rd.get("moderators", [])]
                if not moderators and rd.get("moderator"):
                    moderators = [ReviewParticipant(rd.get("moderator"), "", "moderator")]
                self.reviews.append(
                    ReviewData(
                        name=rd.get("name", ""),
                        description=rd.get("description", ""),
                        mode=rd.get("mode", "peer"),
                        moderators=moderators,
                        participants=participants,
                        comments=comments,
                        approved=rd.get("approved", False),
                        reviewed=rd.get("reviewed", False),
                        due_date=rd.get("due_date", ""),
                        closed=rd.get("closed", False),
                        fta_ids=rd.get("fta_ids", []),
                        fmea_names=rd.get("fmea_names", []),
                        fmeda_names=rd.get("fmeda_names", []),
                        hazop_names=rd.get("hazop_names", []),
                        hara_names=rd.get("hara_names", []),
                        stpa_names=rd.get("stpa_names", []),
                        fi2tc_names=rd.get("fi2tc_names", []),
                        tc2fi_names=rd.get("tc2fi_names", []),
                    )
                )
            current = data.get("current_review")
            self.review_data = None
            for r in self.reviews:
                if r.name == current:
                    self.review_data = r
                    break
        else:
            rd = data.get("review_data")
            if rd:
                participants = [ReviewParticipant(**p) for p in rd.get("participants", [])]
                comments = [ReviewComment(**c) for c in rd.get("comments", [])]
                moderators = [ReviewParticipant(**m) for m in rd.get("moderators", [])]
                if not moderators and rd.get("moderator"):
                    moderators = [ReviewParticipant(rd.get("moderator"), "", "moderator")]
                review = ReviewData(
                    name=rd.get("name", "Review 1"),
                    description=rd.get("description", ""),
                    mode=rd.get("mode", "peer"),
                    moderators=moderators,
                    participants=participants,
                    comments=comments,
                    approved=rd.get("approved", False),
                    reviewed=rd.get("reviewed", False),
                    due_date=rd.get("due_date", ""),
                    closed=rd.get("closed", False),
                    fta_ids=rd.get("fta_ids", []),
                    fmea_names=rd.get("fmea_names", []),
                    fmeda_names=rd.get("fmeda_names", []),
                    hazop_names=rd.get("hazop_names", []),
                    hara_names=rd.get("hara_names", []),
                    stpa_names=rd.get("stpa_names", []),
                    fi2tc_names=rd.get("fi2tc_names", []),
                    tc2fi_names=rd.get("tc2fi_names", []),
                )
                self.reviews = [review]
                self.review_data = review
            else:
                self.review_data = None

        self.update_hara_statuses()
        self.update_fta_statuses()
        self.versions = data.get("versions", [])

        self.selected_node = None
        if hasattr(self, "page_diagram") and self.page_diagram is not None:
            self.close_page_diagram()
        try:
            self.apply_governance_rules()
        except Exception:
            pass
        self.update_views()

    def _reset_on_load(self):
        """Close all open windows and clear state before loading a project."""

        if getattr(self, "page_diagram", None) is not None:
            self.close_page_diagram()

        for tab_id in list(getattr(self.doc_nb, "tabs", lambda: [])()):
            self.doc_nb._closing_tab = tab_id
            self.doc_nb.event_generate("<<NotebookTabClosed>>")
            if tab_id in getattr(self.doc_nb, "tabs", lambda: [])():
                try:
                    self.doc_nb.forget(tab_id)
                except Exception:
                    pass

        for win in (
            list(getattr(self, "use_case_windows", []))
            + list(getattr(self, "activity_windows", []))
            + list(getattr(self, "block_windows", []))
            + list(getattr(self, "ibd_windows", []))
        ):
            try:
                win.destroy()
            except Exception:
                pass
        self.use_case_windows = []
        self.activity_windows = []
        self.block_windows = []
        self.ibd_windows = []

        global AutoML_Helper, unique_node_id_counter
        SysMLRepository.reset_instance()
        AutoML_Helper = config_utils.AutoML_Helper = AutoMLHelper()
        unique_node_id_counter = config_utils.unique_node_id_counter = 1

        self.top_events = []
        self.cta_events = []
        self.root_node = None
        self.selected_node = None
        self.page_history = []
        self.undo_manager.clear_history()
        if getattr(self, "analysis_tree", None):
            self.analysis_tree.delete(*self.analysis_tree.get_children())

        self._reset_fta_state()

    def _prompt_save_before_load_v1(self):
        return messagebox.askyesnocancel(
            "Load Model", "Save current project before loading?"
        )

    def _prompt_save_before_load_v2(self):
        return messagebox.askyesnocancel(
            "Load Model", "Would you like to save before loading a new project?"
        )

    def _prompt_save_before_load_v3(self):
        message = "You have unsaved changes. Save before loading a project?"
        return messagebox.askyesnocancel("Load Model", message)

    def _prompt_save_before_load_v4(self):
        opts = {
            "title": "Load Model",
            "message": "Save changes before loading another project?",
        }
        return messagebox.askyesnocancel(**opts)

    def _prompt_save_before_load(self):
        return self._prompt_save_before_load_v3()


    def update_global_requirements_from_nodes(self,node):
        if hasattr(node, "safety_requirements"):
            for req in node.safety_requirements:
                # Use req["id"] as key; if already exists, you could update if needed.
                ensure_requirement_defaults(req)
                if req["id"] not in global_requirements:
                    global_requirements[req["id"]] = req
                else:
                    ensure_requirement_defaults(global_requirements[req["id"]])
        for child in node.children:
            self.update_global_requirements_from_nodes(child)

    def _generate_pdf_report(self):
        return self.reporting_export._generate_pdf_report()

    def generate_pdf_report(self):
        return self.reporting_export.generate_pdf_report()

    def generate_report(self):
        return self.reporting_export.generate_report()

    def build_html_report(self):
        return self.reporting_export.build_html_report()
    def resolve_original(self, node):
        return resolve_node_original(node)

    def go_back(self):
        return self.nav_input.go_back()

    def draw_page_subtree(self, page_root):
        return self.diagram_renderer.draw_page_subtree(page_root)

    def draw_page_grid(self):
        return self.diagram_renderer.draw_page_grid()

    def draw_page_connections_subtree(self, node, visited_ids):
        return self.diagram_renderer.draw_page_connections_subtree(node, visited_ids)

    def draw_page_nodes_subtree(self, node):
        return self.diagram_renderer.draw_page_nodes_subtree(node)

    def draw_node_on_page_canvas(self, *args, **kwargs):
        return self.diagram_renderer.draw_node_on_page_canvas(*args, **kwargs)

    def on_ctrl_mousewheel_page(self, event):
        return self.nav_input.on_ctrl_mousewheel_page(event)

    def close_page_diagram(self):
        return self.diagram_renderer.close_page_diagram()

    # --- Review Toolbox Methods ---
    def start_peer_review(self):
        return self.versioning_review.start_peer_review()

    def start_joint_review(self):
        return self.versioning_review.start_joint_review()

    def open_review_document(self, review):
        return self.versioning_review.open_review_document(review)

    def open_review_toolbox(self):
        return self.versioning_review.open_review_toolbox()

    def send_review_email(self, review):
        return self.versioning_review.send_review_email(review)

    def review_is_closed(self):
        return self.versioning_review.review_is_closed()

    def review_is_closed_for(self, review):
        return self.versioning_review.review_is_closed_for(review)

    def capture_diff_diagram(self, top_event):  # pragma: no cover - delegation
        return self.diagram_renderer.capture_diff_diagram(top_event)

    # --- End Review Toolbox Methods ---

    def compute_requirement_asil(self, req_id):
        """Return highest ASIL across all safety goals linked to the requirement."""
        goals = self.get_requirement_goal_names(req_id)
        asil = "QM"
        for g in goals:
            a = self.get_safety_goal_asil(g)
            if ASIL_ORDER.get(a, 0) > ASIL_ORDER.get(asil, 0):
                asil = a
        return asil

    def find_safety_goal_node(self, name):
        for te in self.top_events:
            if name in (te.safety_goal_description, te.user_name):
                return te
        return None

    def compute_validation_criteria(self, req_id):  # pragma: no cover - delegation
        return self.validation_consistency.compute_validation_criteria(req_id)

    def update_validation_criteria(self, req_id):  # pragma: no cover - delegation
        return self.validation_consistency.update_validation_criteria(req_id)

    def update_requirement_asil(self, req_id):
        req = global_requirements.get(req_id)
        if not req:
            return
        req["asil"] = self.compute_requirement_asil(req_id)

    def update_all_validation_criteria(self):  # pragma: no cover - delegation
        return self.validation_consistency.update_all_validation_criteria()

    def update_all_requirement_asil(self):
        for rid, req in global_requirements.items():
            if req.get("parent_id"):
                continue  # keep decomposition ASIL
            self.update_requirement_asil(rid)

    def update_base_event_requirement_asil(self):
        return self.safety_analysis.update_base_event_requirement_asil()

    def ensure_asil_consistency(self):
        """Sync safety goal ASILs from risk assessments and update requirement ASILs."""
        self.update_fta_statuses()
        self.sync_hara_to_safety_goals()
        self.update_hazard_list()
        self.update_all_requirement_asil()
        self.validation_consistency.update_all_validation_criteria()

    def invalidate_reviews_for_hara(self, name):
        return self.versioning_review.invalidate_reviews_for_hara(name)

    def invalidate_reviews_for_requirement(self, req_id):
        return self.versioning_review.invalidate_reviews_for_requirement(req_id)

    def add_version(self):
        return self.versioning_review.add_version()


    def compare_versions(self):
        return self.versioning_review.compare_versions()


    def merge_review_comments(self):
        return self.versioning_review.merge_review_comments()


    def calculate_diff_nodes(self, old_data):
        return self.review_manager.calculate_diff_nodes(old_data)


    def calculate_diff_between(self, data1, data2):
        return self.review_manager.calculate_diff_between(data1, data2)


    def node_map_from_data(self, top_events):
        return self.structure_tree_operations.node_map_from_data(top_events)


    def set_current_user(self):
        self.user_manager.set_current_user()

    def get_current_user_role(self):
        return self.data_access_queries.get_current_user_role()

    def focus_on_node(self, node):
        return self.nav_input.focus_on_node(node)

    def get_review_targets(self):
        return self.versioning_review.get_review_targets()


def load_user_data() -> tuple[dict, tuple[str, str]]:
    """Load cached users and last user config concurrently."""
    with ThreadPoolExecutor() as executor:
        users_future = executor.submit(load_all_users)
        config_future = executor.submit(load_user_config)
        return users_future.result(), config_future.result()


def main():
    root = tk.Tk()
    # Prevent the main window from being resized so small that
    # widgets and toolbars become unusable.
    root.minsize(1200, 700)
    enable_listbox_hover_highlight(root)
    # Hide the main window while prompting for user info
    root.withdraw()
    users, (last_name, last_email) = load_user_data()
    if users:
        dlg = UserSelectDialog(root, users, last_name)
        if dlg.result:
            name, email = dlg.result
            if name == "New User...":
                info = UserInfoDialog(root, "", "").result
                if info:
                    name, email = info
                    save_user_config(name, email)
            else:
                email = users.get(name, email)
                set_last_user(name)
    else:
        dlg = UserInfoDialog(root, last_name, last_email)
        if dlg.result:
            name, email = dlg.result
            save_user_config(name, email)
    set_current_user(name, email)
    # Create a fresh helper each session:
    global AutoML_Helper
    AutoML_Helper = config_utils.AutoML_Helper = AutoMLHelper()

    # Show and maximize the main window after login
    root.deiconify()
    try:
        root.state("zoomed")
    except tk.TclError:
        try:
            root.attributes("-zoomed", True)
        except tk.TclError:
            pass

    app = AutoMLApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
